import streamlit as st
import re
import json
import copy
import streamlit as st
import pandas as pd  # <--- Bắt buộc phải có dòng này
import threading

# =====================================================================
# ĐOẠN 6a - PHẦN 1: KHỞI TẠO BANNER VĂN BẢN ĐỈNH SIDEBAR CHUẨN ERP
# =====================================================================

# 1. Cấu hình trang rộng toàn màn hình chuẩn hệ thống SaaS/ERP Văn phòng
st.set_page_config(layout="wide", page_title="AI Fabric Consumption Matrix")

# Dùng lệnh Python tạo hộp chữ Banner vuông vắn ghim cứng lên đỉnh lề trái
with st.sidebar:
    st.markdown(
        """
        <div style="
            background: linear-gradient(135deg, #1e40af 0%, #1d4ed8 100%); 
            border-radius: 8px; 
            padding: 15px 10px; 
            text-align: center; 
            margin-top: -30px; /* Đẩy sát kịch trần lề trên */
            margin-bottom: 20px; 
            box-shadow: 0 4px 6px rgba(0, 0, 0, 0.15);
        ">
            <div style="font-family: 'Segoe UI', sans-serif; font-size: 16px; font-weight: 800; color: #ffffff; letter-spacing: 0.5px; line-height: 1.2;">
                PPJ GROUP
            </div>
            <div style="font-family: 'Segoe UI', sans-serif; font-size: 9px; font-weight: 600; color: #bfdbfe; letter-spacing: 0.3px; margin-top: 4px; text-transform: uppercase;">
                BOUNDLESS SOLUTIONS
            </div>
        </div>
        """,
        unsafe_allow_html=True
    )

# 2. Khởi tạo cấu trúc trạng thái bộ nhớ hệ thống (Session State) an toàn
if "bom_data" not in st.session_state: st.session_state.bom_data = None
if "chat_history" not in st.session_state: st.session_state.chat_history = []
if "pdf_bytes" not in st.session_state: st.session_state.pdf_bytes = None
if "pdf_name" not in st.session_state: st.session_state.pdf_name = ""
if "pdf_text_cache" not in st.session_state: st.session_state.pdf_text_cache = None
if "accumulated_bom_rows" not in st.session_state: st.session_state.accumulated_bom_rows = []



# =====================================================================
# ĐOẠN 6a - PHẦN 2: BỘ CẤU HÌNH CSS ĐỒNG BỘ MÀU SẮC & XỬ LÝ LỀ SIDEBAR - BẢN TINH GỌN CHỐNG TRỐNG
# =====================================================================
st.markdown("""
<style>
    /* 🎨 ÉP ĐỒNG BỘ TOÀN DIỆN MÀU NỀN XANH NGỌC MỊN CHO TẤT CẢ CÁC LỚP BAN NỀN */
    .stApp, header[data-testid="stHeader"], div[data-testid="stMainView"], section[data-testid="stSidebar"] + div { 
        background-color: #e6f4f1 !important; 
    }
    
    /* Ép khoảng cách lề của toàn vùng nội dung dạt sang phải, tránh so le */
    .block-container { 
        padding-top: 1.5rem !important; 
        padding-left: 2rem !important; 
        padding-right: 2rem !important; 
        margin-left: 300px !important; /* Tạo khoảng trống chuẩn tách biệt Sidebar */
        max-width: calc(100% - 300px) !important; 
        padding-bottom: 120px !important; 
    }
    
    /* Ép tất cả các khối ngang (Horizontal Blocks) dãn đều */
    div[data-testid="stHorizontalBlock"] { 
        margin-top: 0px !important; 
        padding-top: 0px !important; 
        width: 100% !important;
        max-width: 100% !important;
        background-color: transparent !important; 
    }

    /* NHUỘM XANH CÁC KHỐI CHỨA UPLOADER VÀ VISUALIZER CHO ĐỒNG BỘ */
    div[data-testid="stVerticalBlockBorderWrapper"] {
        background-color: #f0fdfa !important; 
        border: 1px solid #99f6e4 !important; 
        box-shadow: 0 4px 6px -1px rgba(15, 118, 110, 0.05) !important;
    }

    /* CẤU HÌNH THANH MENU ĐIỀU HƯỚNG ĐA TRANG GỌN GÀNG */
    [data-testid="stSidebarNav"] {
        padding-top: 5px !important; /* Thu hẹp lại lề vì khoảng trống đã được lấp đầy */
        background-color: transparent !important;
        position: relative !important;
    }
    
    /* Triệt tiêu hoàn toàn các cơ chế vẽ đè cũ tránh xung đột */
    [data-testid="stSidebarNav"]::before,
    [data-testid="stSidebarNav"]::after {
        content: "" !important;
        display: none !important;
    }

    /* ÉP TOÀN BỘ KHỐI CONTAINER CHAT VÀ FIELDSET PHÌNH TO 100% CHẠM BIÊN ĐÁY MÀN HÌNH */
    .stChatInput,
    .stChatInput > div,
    .stChatInput fieldset,
    div[data-testid="stChatInputContainer"] {
        position: fixed !important;
        bottom: 0 !important; 
        left: 300px !important; 
        right: 0 !important; 
        width: calc(100% - 300px) !important; 
        max-width: calc(100% - 300px) !important; 
        background-color: #ccfbf1 !important; 
        border: none !important;
        border-top: 1px solid #5eead4 !important; 
        border-radius: 0px !important; 
        box-shadow: 0 -4px 10px rgba(15, 118, 110, 0.06) !important;
        padding: 10px 2rem !important; 
    }

    /* Định dạng lõi nhập văn bản gõ chữ bên trong cho tiệp màu xanh ngọc */
    div[data-testid="stChatInputContainer"] textarea {
        background-color: #ccfbf1 !important; 
        color: #115e59 !important; 
        font-family: "Segoe UI", sans-serif !important;
        font-size: 13px !important;
        width: 100% !important;
        border: none !important;
    }

    /* Căn chỉnh nút gửi hình mũi tên gọn gàng bên góc phải dải băng chat */
    div[data-testid="stChatInputContainer"] button {
        background-color: #0f766e !important;
        color: #ffffff !important;
        border-radius: 6px !important;
    }

    /* 🛠️ MỞ KHÓA THANH CUỘN SIDEBAR CỐ ĐỊNH BÊN TRÁI HỆ THỐNG (SỬA LỖI LẤP TÍNH NĂNG) */
    [data-testid="stSidebar"] {
        background-color: #0f766e !important; 
        color: #ffffff !important;
        position: fixed !important;
        top: 0 !important;
        left: 0 !important;
        width: 300px !important; 
        min-width: 300px !important;
        max-width: 300px !important;
        height: 100vh !important;
        overflow-y: auto !important; /* Đã chuyển từ hidden thành auto để tự động xuất hiện thanh cuộn dọc */
        z-index: 99999 !important;
    }
    
    [data-testid="stSidebar"] > div:first-child {
        height: 100vh !important;
        overflow-y: auto !important; /* Đã mở khóa lớp cuộn bên trong phôi RAM */
    }

    div[data-testid="stSidebarNav"] {
        font-size: 11px !important;
    }

    /* Định dạng nút bấm xóa bộ nhớ tinh tế */
    [data-testid="stSidebar"] button {
        background-color: #115e59 !important;
        color: #fca5a5 !important;
        border: 1px solid #115e59 !important;
        font-weight: 600 !important;
        transition: all 0.2s ease !important;
    }
    [data-testid="stSidebar"] button:hover {
        background-color: #dc2626 !important;
        color: #ffffff !important;
    }

    .sidebar-sub-title {
        font-family: "Segoe UI", sans-serif !important; 
        font-size: 12px !important; 
        font-weight: 800 !important; 
        color: #fde047 !important; 
        text-shadow: 0px 1px 2px rgba(0,0,0,0.4) !important; 
        text-transform: uppercase !important; 
        letter-spacing: 0.8px !important; 
        margin-bottom: 6px !important;
        margin-top: 18px !important;
    }

    .sidebar-custom-card, .sidebar-custom-card-history {
        background: linear-gradient(135deg, #115e59 0%, #134e4a 100%) !important; 
        border: 1px solid #14b8a6 !important;
        border-radius: 6px !important; 
        padding: 12px !important; 
        box-shadow: 0 4px 6px rgba(0,0,0,0.15) !important;
        margin-bottom: 15px !important;
    }
    .sidebar-custom-card-history { padding: 6px 12px !important; }

    /* 🛠️ ÉP THU HẸP KHOẢNG CÁCH DUNG LƯỢNG KEY TRÊN SIDEBAR - CHỐNG LÃNG PHÍ KHÔNG GIAN VÙNG TRỐNG */
    [data-testid="stSidebar"] hr {
        margin-top: 8px !important;
        margin-bottom: 8px !important;
    }
    [data-testid="stSidebar"] h5 {
        margin-bottom: 2px !important;
        padding-bottom: 0px !important;
    }
    [data-testid="stSidebar"] div[data-testid="stProgress"] {
        margin-top: 2px !important;
        margin-bottom: 2px !important;
    }
    [data-testid="stSidebar"] .stCaptionContainer {
        margin-top: -4px !important;
        margin-bottom: 4px !important;
    }
    /* Kéo giật khối ngang chứa 2 Metric (Số lượt quét và số Tokens) lên sát dải băng pin phía trên */
    [data-testid="stSidebar"] div[data-testid="stVerticalBlock"] > div[data-testid="stHorizontalBlock"] {
        margin-top: -12px !important; 
        gap: 4px !important; 
    }
    [data-testid="stSidebar"] div[data-testid="stMetricWidget"] {
        padding-top: 0px !important;
        padding-bottom: 0px !important;
        margin-bottom: 0px !important;
    }
    [data-testid="stSidebar"] label[data-testid="stMetricLabel"] p {
        font-size: 11px !important;
    }
    .sidebar-divider { margin: 12px 0 8px 0 !important; border: 0 !important; border-top: 1px solid #115e59 !important; }

    .kpi-box-flat-matrix { border-radius: 6px 6px 0 0 !important; padding: 10px 12px !important; text-align: center !important; box-shadow: 0 2px 4px rgba(0,0,0,0.02) !important; box-sizing: border-box !important; }
    .kpi-num-flat-matrix { font-size: 16px !important; font-weight: 700 !important; color: #ffffff !important; font-family: 'Segoe UI', sans-serif !important; line-height: 1.2 !important; }
    .kpi-lbl-flat-matrix { font-size: 9px !important; font-weight: 600 !important; color: #ffffff !important; opacity: 0.95 !important; text-transform: uppercase !important; margin-top: 2px !important; }
    .bg-style-erp { background: linear-gradient(135deg, #334155 0%, #1e293b 100%) !important; }
    .bg-items-erp { background: linear-gradient(135deg, #0d9488 0%, #0f766e 100%) !important; }
    .bg-cons-erp  { background: linear-gradient(135deg, #ea580c 0%, #c2410c 100%) !important; }
    .bg-size-erp { background: linear-gradient(135deg, #16a34a 0%, #15803d 100%) !important; }

    .image-placeholder-box-flat { border: 1px solid #cbd5e1 !important; border-top: none !important; border-radius: 0 0 6px 6px !important; padding: 10px 5px !important; height: 140px !important; display: flex !important; align-items: center !important; justify-content: center !important; box-sizing: border-box !important; margin-bottom: 25px !important; background-color: #ffffff !important; overflow: hidden !important; }
    div[data-testid="stImage"] img { width: 100% !important; height: auto !important; }
    
    /* Khử và ẩn hoàn toàn cái icon rác ảnh vỡ nhỏ màu trắng ở giữa lề */
    [data-testid="stSidebar"] img, [data-testid="stSidebar"] div[data-testid="stImage"] { display: none !important; }
    .main-body-spacer, .sticky-top-container, div[smart-fixed-container], div[data-testid="stHorizontalBlock"]:empty { display: none !important; height: 0px !important; margin: 0 !important; padding: 0 !important; }
</style>
""", unsafe_allow_html=True)



import streamlit as st
import re

# =====================================================================
# KHỞI TẠO DỮ LIỆU ĐỂ TRÁNH LỖI BIẾN CHƯA ĐỊNH NGHĨA (NAMEERROR)
# =====================================================================
kpi_style_id = st.session_state.get("style_id", "N/A")
total_materials = 0
main_fabric_cons = "0.00"
active_size_kpi = "M"

# Khởi tạo các giá trị session state mặc định nếu chưa có
if "pdf_name" not in st.session_state: st.session_state.pdf_name = ""
if "pdf_bytes" not in st.session_state: st.session_state.pdf_bytes = None
if "pdf_text_cache" not in st.session_state: st.session_state.pdf_text_cache = None

# =====================================================================
# ĐOẠN B: GIAO DIỆN HIỂN THỊ KPIs MÀU SẮC ĐỘNG & GRID THÂN TRANG HỢP NHẤT
# =====================================================================

# 🌟 TIÊU ĐỀ ĐÃ ĐỔI SANG MÀU XANH THEME ERP SANG TRỌNG 🌟
st.markdown(
    """
    <div style="background: linear-gradient(135deg, #0f766e 0%, #115e59 100%); border-radius: 6px; padding: 14px 20px; margin-bottom: 20px; box-shadow: 0 4px 6px -1px rgba(15, 118, 110, 0.1), 0 2px 4px -1px rgba(15, 118, 110, 0.06); text-align: center;">
        <h2 style="font-family: 'Segoe UI', sans-serif; font-size: 16px; font-weight: 700; color: #ffffff; margin: 0; text-transform: uppercase; letter-spacing: 0.8px;">
            🚀 AUTOMATED CAD CONSUMPTION & INDUSTRIAL COSTING ENGINE
        </h2>
    </div>
    """, 
    unsafe_allow_html=True
)

# Phân bổ lưới 4 ô KPIs Native gốc của Streamlit
k_col1, k_col2, k_col3, k_col4 = st.columns(4)

# Cấu hình chung cho hiệu ứng Emoji: To rõ (70px), có đổ bóng mờ tạo độ nổi khối 3D cực đẹp
emoji_style = "font-size: 70px; display: inline-block; filter: drop-shadow(0px 4px 6px rgba(0, 0, 0, 0.15)); transform: scale(1); transition: all 0.2s ease-in-out;"

with k_col1: 
    st.markdown(f'<div class="kpi-box-flat-matrix bg-style-erp"><div class="kpi-num-flat-matrix">{kpi_style_id}</div><div class="kpi-lbl-flat-matrix">Mã hàng đang xử lý</div></div>', unsafe_allow_html=True)
    st.markdown(f'<div class="image-placeholder-box-flat"><span style="{emoji_style}">👕</span></div>', unsafe_allow_html=True)

with k_col2: 
    st.markdown(f'<div class="kpi-box-flat-matrix bg-items-erp"><div class="kpi-num-flat-matrix">{total_materials} Item(s)</div><div class="kpi-lbl-flat-matrix">Tổng số vật tư kết xuất</div></div>', unsafe_allow_html=True)
    st.markdown(f'<div class="image-placeholder-box-flat"><span style="{emoji_style}">👖</span></div>', unsafe_allow_html=True)

with k_col3: 
    st.markdown(f'<div class="kpi-box-flat-matrix bg-cons-erp"><div class="kpi-num-flat-matrix">{main_fabric_cons}</div><div class="kpi-lbl-flat-matrix">Định mức vải chính dự kiến</div></div>', unsafe_allow_html=True)
    st.markdown(f'<div class="image-placeholder-box-flat"><span style="{emoji_style}">✂️</span></div>', unsafe_allow_html=True)

with k_col4: 
    st.markdown(f'<div class="kpi-box-flat-matrix bg-size-erp"><div class="kpi-num-flat-matrix">{active_size_kpi}</div><div class="kpi-lbl-flat-matrix">Cỡ hạt tính định mức</div></div>', unsafe_allow_html=True)
    st.markdown(f'<div class="image-placeholder-box-flat"><span style="{emoji_style}">🧵</span></div>', unsafe_allow_html=True)


# --- BẢNG ĐIỀU KHIỂN SIDEBAR MÁY CHỦ MỚI (CỐ ĐỊNH CHUẨN XÁC CHẾ ĐỘ KEY MUA - PREMIUM KEY ĐỒNG BỘ) ---
st.sidebar.markdown("### ⚙️ ENGINE CONTROLS")
if st.sidebar.button("🗑️ CLEAR SYSTEM MEMORY", use_container_width=True):
    st.session_state.bom_data = {}
    st.session_state.chat_history = []
    st.session_state.pdf_bytes = None
    st.session_state.pdf_name = ""
    st.session_state.pdf_text_cache = None
    if "processed_display_rows" in st.session_state: st.session_state.processed_display_rows = []
    if "accumulated_bom_rows" in st.session_state: st.session_state.accumulated_bom_rows = []
    if "last_active_blueprint" in st.session_state: st.session_state.last_active_blueprint = None
    if "raw_ai_debug_payload" in st.session_state: st.session_state.raw_ai_debug_payload = None
    if "pdf_page_one_image" in st.session_state: st.session_state.pdf_page_one_image = None
    st.rerun()

st.sidebar.markdown("---")
st.sidebar.markdown("##### 🔑 API KEY CAPACITY (DUNG LƯỢNG KEY)")

# Khởi tạo bộ đếm dung lượng tiêu hao thực tế trong màng RAM hệ thống
if "api_calls_count" not in st.session_state:
    st.session_state["api_calls_count"] = 0
if "tokens_consumed" not in st.session_state:
    st.session_state["tokens_consumed"] = 0

# Tính toán dung lượng ước tính dựa trên dữ liệu văn bản thô Techpack đã quét
if st.session_state.get("raw_pdf_text_extracted"):
    current_tokens = len(str(st.session_state["raw_pdf_text_extracted"])) // 4
    if st.session_state["tokens_consumed"] == 0:
        st.session_state["tokens_consumed"] = current_tokens
        st.session_state["api_calls_count"] += 1

# 🚨 SỬA LỖI ĐỒNG BỘ: Ép cứng nhận diện chế độ Key mua (Premium) cho công ty
# - Gán cứng True vì đây là Key mua đã nạp tiền/liên kết thẻ thanh toán của công ty bạn.
# - Gán False nếu sau này bạn muốn đổi sang một mã Key cá nhân dùng thử miễn phí nào khác.
is_paid_key = True  

if is_paid_key:
    # ➔ CHẾ ĐỘ KEY MUA: Mở khóa băng thông, màn hình luôn báo đầy pin 100% màu xanh ngọc lam sáng rõ
    st.sidebar.progress(1.0) 
    st.sidebar.caption("CN⚙️ **TÀI KHOẢN TRẢ PHÍ (PREMIUM KEY)** | Trạng thái: `BĂNG THÔNG MỞ RỘNG`")
else:
    # ➔ CHẾ ĐỘ KEY FREE: Tự động co giãn theo giới hạn 1500 lượt của Google
    max_daily_calls = 1500
    capacity_percentage = (max(0, max_daily_calls - st.session_state["api_calls_count"]) / max_daily_calls) * 100
    st.sidebar.progress(capacity_percentage / 100)
    st.sidebar.caption(f"🔋 Dung lượng khả dụng (Free Tier): `{capacity_percentage:.1f}%`")

# Hiển thị bộ đôi số liệu trực quan khít sát lề dải pin, chống lãng phí khoảng trống dư thừa
col_cap1, col_cap2 = st.sidebar.columns(2)
with col_cap1:
    st.metric("🔄 Lượt đã quét", f"{st.session_state['api_calls_count']}")
with col_cap2:
    st.metric("📊 Đã dùng (Tokens)", f"{st.session_state['tokens_consumed']:,}")

# CẤU HÌNH CANH SỢI SƠ ĐỒ (CAD) CO GIÃN ĐỊNH MỨC THEO GERBER CỦA BẠN VẪN GIỮ NGUYÊN PHÍA DƯỚI
st.sidebar.markdown("---")
st.sidebar.markdown("##### 📏 CẤU HÌNH CANH SỢI SƠ ĐỒ (CAD)")

st.sidebar.checkbox(
    "🔄 Cắt tự do (Xoay ngược 180°)", 
    key="allow_rotation_90", 
    value=True,
    help="Cho phép chi tiết xoay ngược đầu đuôi tự do. Trong một bộ không nhất thiết phải cùng chiều. Tối ưu sơ đồ khít nhất, định mức thấp nhất."
)
st.sidebar.checkbox(
    "✂️ Cắt mỗi bộ 1 chiều (Nap)", 
    key="is_nap_layout",
    help="Tất cả chi tiết trong 1 bộ rập phải xoay cùng 1 chiều dọc thớ vải."
)
st.sidebar.checkbox(
    "🧵 Tất cả size 1 chiều (One-Way)", 
    key="is_one_way_fabric",
    help="Ép toàn bộ chi tiết rập của mọi cỡ size quay chung về 1 hướng duy nhất (vải tuyết/nhung)."
)




# --- TÍCH HỢP 3 Ý TƯỞNG TIỆN ÍCH DƯỚI NÚT CLEAR (MÀU XANH NGỌC LAM) ---
with st.sidebar:
    # -----------------------------------------------------------------
    # KHỐI 1: THÔNG TIN HỆ THỐNG (SYSTEM STATUS)
    # -----------------------------------------------------------------
    st.markdown("<hr style='margin: 20px 0 12px 0; border: 0; border-top: 1px solid #115e59;'>", unsafe_allow_html=True)
    st.markdown("<div style='font-family:\"Segoe UI\", sans-serif; font-size: 12px; font-weight: 700; color: #ffffff; text-transform: uppercase; letter-spacing: 0.5px; margin-bottom: 8px;'>⚙️ SYSTEM STATUS</div>", unsafe_allow_html=True)
    st.markdown(
        """
        <div style="background: linear-gradient(135deg, #115e59 0%, #134e4a 100%); border: 1px solid #14b8a6; border-radius: 6px; padding: 12px; box-shadow: 0 4px 6px rgba(0,0,0,0.15); margin-bottom: 15px;">
            <div style="display: flex; justify-content: space-between; margin-bottom: 6px; font-size: 11px; color: #ccfbf1; font-family: 'Segoe UI', sans-serif;">
                <span>Core Engine:</span>
                <span style="color: #ffffff; font-weight: 700;">v2.4.1-AI</span>
            </div>
            <div style="display: flex; justify-content: space-between; margin-bottom: 6px; font-size: 11px; color: #ccfbf1; font-family: 'Segoe UI', sans-serif;">
                <span>AI CAD Status:</span>
                <span style="color: #4ade80; font-weight: 700;">● Connected</span>
            </div>
            <div style="display: flex; justify-content: space-between; font-size: 11px; color: #ccfbf1; font-family: 'Segoe UI', sans-serif;">
                <span>Response Time:</span>
                <span style="color: #fde047; font-weight: 700;">&lt; 1.2s</span>
            </div>
        </div>
        """,
        unsafe_allow_html=True
    )

    # -----------------------------------------------------------------
    # KHỐI 2: HƯỚNG DẪN SỬ DỤNG NHANH (QUICK USER GUIDE)
    # -----------------------------------------------------------------
    st.markdown("<div style='font-family:\"Segoe UI\", sans-serif; font-size: 12px; font-weight: 700; color: #ffffff; text-transform: uppercase; letter-spacing: 0.5px; margin-bottom: 8px;'>📖 QUICK USER GUIDE</div>", unsafe_allow_html=True)
    st.markdown(
        """
        <div style="background: linear-gradient(135deg, #115e59 0%, #134e4a 100%); border: 1px solid #14b8a6; border-radius: 6px; padding: 12px; box-shadow: 0 4px 6px rgba(0,0,0,0.15); margin-bottom: 15px;">
            <div style="display: flex; align-items: flex-start; margin-bottom: 10px; font-size: 11px; color: #ffffff; font-family: 'Segoe UI', sans-serif;">
                <div style="background-color: #14b8a6; color: #ffffff; font-weight: 700; width: 18px; height: 18px; border-radius: 50%; display: flex; align-items: center; justify-content: center; margin-right: 8px; flex-shrink: 0;">1</div>
                <div style="line-height: 1.4;"><span style="font-weight: 700; color: #2dd4bf;">Tải tài liệu:</span> Upload file Techpack PDF.</div>
            </div>
            <div style="display: flex; align-items: flex-start; margin-bottom: 10px; font-size: 11px; color: #ffffff; font-family: 'Segoe UI', sans-serif;">
                <div style="background-color: #14b8a6; color: #ffffff; font-weight: 700; width: 18px; height: 18px; border-radius: 50%; display: flex; align-items: center; justify-content: center; margin-right: 8px; flex-shrink: 0;">2</div>
                <div style="line-height: 1.4;"><span style="font-weight: 700; color: #2dd4bf;">Định mức:</span> Xem dữ liệu ở bảng đinh mức.</div>
            </div>
            <div style="display: flex; align-items: flex-start; font-size: 11px; color: #ffffff; font-family: 'Segoe UI', sans-serif;">
                <div style="background-color: #14b8a6; color: #ffffff; font-weight: 700; width: 18px; height: 18px; border-radius: 50%; display: flex; align-items: center; justify-content: center; margin-right: 8px; flex-shrink: 0;">3</div>
                <div style="line-height: 1.4;"><span style="font-weight: 700; color: #2dd4bf;">Xuất bảng:</span> Lưu bảng BOM Matrix.</div>
            </div>
        </div>
        """,
        unsafe_allow_html=True
    )

    # -----------------------------------------------------------------
    # KHỐI 3: DANH SÁCH LỊCH SỬ MÃ HÀNG ĐỘNG (RECENT CODE HISTORY)
    # -----------------------------------------------------------------
    st.markdown("<div style='font-family:\"Segoe UI\", sans-serif; font-size: 12px; font-weight: 700; color: #ffffff; text-transform: uppercase; letter-spacing: 0.5px; margin-bottom: 8px;'>🕒 RECENT CODE HISTORY</div>", unsafe_allow_html=True)
    if "history_list" not in st.session_state:
        st.session_state.history_list = ["PPJ-K12-200451", "PPJ-M04-330129"]
    if kpi_style_id != "N/A" and kpi_style_id not in st.session_state.history_list:
        st.session_state.history_list.insert(0, kpi_style_id)
        st.session_state.history_list = st.session_state.history_list[:3]

    history_html = '<div style="background: linear-gradient(135deg, #115e59 0%, #134e4a 100%); border: 1px solid #14b8a6; border-radius: 6px; padding: 6px 12px; box-shadow: 0 4px 6px rgba(0,0,0,0.15); margin-bottom: 15px;">'
    for index, style_code in enumerate(st.session_state.history_list):
        border_style = 'border-bottom: 1px solid #14b8a6;' if index < len(st.session_state.history_list) - 1 else ''
        if index == 0 and kpi_style_id != "N/A":
            history_html += (
                '<div style="display: flex; justify-content: space-between; align-items: center; padding: 7px 0; ' + border_style + ' font-size: 11px; font-family: \'Segoe UI\', sans-serif;">'
                '    <span style="color: #ffffff; font-weight: 700;">📦 ' + style_code + '</span>'
                '    <span style="color: #ffffff; font-size: 10px; font-weight: 700; background-color: #14b8a6; padding: 2px 8px; border-radius: 10px; box-shadow: 0 1px 3px rgba(0,0,0,0.2);">Active</span>'
                '</div>'
            )
        else:
            history_html += (
                '<div style="display: flex; justify-content: space-between; align-items: center; padding: 7px 0; ' + border_style + ' font-size: 11px; font-family: \'Segoe UI\', sans-serif;">'
                '    <span style="color: #ccfbf1; font-weight: 600;">📦 ' + style_code + '</span>'
                '    <span style="color: #5eead4; font-size: 10px;">Processed</span>'
                '</div>'
            )
    history_html += '</div>'
    
    st.markdown(history_html, unsafe_allow_html=True)
    st.markdown("<div style='font-size: 10px; color: #ccfbf1; font-family: \"Segoe UI\", sans-serif; text-align: center; margin-top: 15px; opacity: 0.8;'>© 2026 PPJ Digital Transformation</div>", unsafe_allow_html=True)




import streamlit as st
import re
import fitz  # Thư viện PyMuPDF để trích xuất văn bản và ảnh tự động từ file PDF

# ------------------------------------------------------------------------------
# LƯỚI CHIA ĐÔI CỘT CHÍNH THỰC TẾ (ĐÃ ĐÓNG KHUNG VIỀN ĐẸP MẮT & SỬA LỖI HIỂN THỊ)
# ------------------------------------------------------------------------------
col_left, col_right = st.columns(2)

# --- CỘT TRÁI: BỘ TẢI FILE & HỒ SƠ TÓM TẮT MÃ HÀNG MÀU XANH ---
with col_left:
    with st.container(border=True, height=520):
        st.markdown("### 📂 TECHPACK UPLOADER & PROFILE SUMMARY")
        
        uploaded_file = st.file_uploader("Upload PDF", type=["pdf"], label_visibility="collapsed")
        
        if uploaded_file is not None:
            # Nếu phát hiện người dùng tải lên một file hoàn toàn mới
            if st.session_state.pdf_name != uploaded_file.name:
                st.session_state.pdf_text_cache = None
                st.session_state.pdf_page_one_image = None
                if "accumulated_bom_rows" in st.session_state: st.session_state.accumulated_bom_rows = []
                
            st.session_state.pdf_bytes = uploaded_file.read()
            st.session_state.pdf_name = uploaded_file.name

            # Tự động bóc tách Văn Bản và chuyển đổi PDF thành hình ảnh Sketch ngay lập tức
            if st.session_state.pdf_text_cache is None or st.session_state.pdf_page_one_image is None:
                with st.spinner("🤖 AI đang đọc tài liệu và trích xuất hình ảnh phác thảo..."):
                    try:
                        # Mở file PDF trực tiếp từ bộ nhớ bytes
                        doc = fitz.open(stream=st.session_state.pdf_bytes, filetype="pdf")
                        
                        # 1. Trích xuất toàn bộ text từ tất cả các trang phục vụ Regex tìm mã hàng
                        full_text = ""
                        for page in doc:
                            full_text += page.get_text()
                        st.session_state.pdf_text_cache = full_text
                        
                        # 2. Chuyển đổi trang đầu tiên (Trang 0) thành hình ảnh PNG chất lượng cao
                        if len(doc) > 0:
                            page_one = doc[0]
                            pix = page_one.get_pixmap(matrix=fitz.Matrix(2, 2)) # Zoom x2 để ảnh nét hơn
                            image_bytes = pix.tobytes("png")
                            st.session_state.pdf_page_one_image = image_bytes
                            
                        doc.close()
                    except Exception as e:
                        st.error(f"Lỗi khi đọc file PDF kĩ thuật: {e}")
                
                # Khởi động lại luồng giao diện để cập nhật ngay lập tức dữ liệu mới lên màn hình
                st.rerun()

        # Hiển thị thông tin hồ sơ tóm tắt sau khi đã trích xuất văn bản thành công
        if st.session_state.pdf_text_cache is not None:
            st.markdown("<div style='margin-top: 15px;'></div>", unsafe_allow_html=True)
            txt = st.session_state.pdf_text_cache
            
            def get_meta(pattern, default="N/A"):
                m = re.search(pattern, txt, re.IGNORECASE)
                return m.group(1).strip() if m else default

            # Thực hiện quét thông tin kĩ thuật bằng biểu thức chính quy (Regex)
            style_id = get_meta(r'(?:Style ID|Style_ID|Mã hàng)\s*[:\-=\s]*([\w\d\-]+)', st.session_state.pdf_name.replace(".pdf",""))
            short_desc = get_meta(r'(?:Short Desc|Description|Tên sản phẩm)\s*[:\-=\s]*([^\n]+)', "THE BAGGY JEANS")
            customer = get_meta(r'(?:Customer|Khách hàng|Brand)\s*[:\-=\s]*([^\n]+)', "FACTORY STANDARD")
            season = get_meta(r'(?:Season|Mùa hàng)\s*[:\-=\s]*([^\n]+)', "Fall 2025 Apparel Reitmans")
            fabric_type = get_meta(r'(?:Long Description|Chất liệu gốc)\s*[:\-=\s]*([^\n]+)', "LIGHT ORANGE - MID RISE - POPLIN FABRIC")

            # Ghim mã hàng vào bộ nhớ toàn cục để đồng bộ lên các khối KPIs trần trang và Lịch sử
            st.session_state.style_id = style_id

            # Bộ cấu hình CSS cao cấp đóng khung hộp viền mịn màng, đổ bóng 3D độc lập
            box_style = (
                "background-color: #f8fafc; "
                "border: 1px solid #e2e8f0; "
                "border-radius: 6px; "
                "padding: 12px 14px; "
                "margin-bottom: 12px; "
                "box-shadow: 0 1px 3px 0 rgba(0, 0, 0, 0.05);"
            )
            lbl_style = "font-family: 'Segoe UI', sans-serif; font-size: 11px; font-weight: 600; color: #64748b; text-transform: uppercase; letter-spacing: 0.5px; margin-bottom: 4px;"
            val_style = "font-family: 'Segoe UI', sans-serif; font-size: 14px; font-weight: 700; color: #1e293b; white-space: nowrap; overflow: hidden; text-overflow: ellipsis;"

            # Chia lưới cột nhỏ bên trong khung Techpack Uploader
            m_col1, m_col2 = st.columns(2)
            with m_col1:
                st.markdown(f'<div style="{box_style}"><div style="{lbl_style}">Style Code / Mã hàng</div><div style="{val_style}; color: #0f766e;">{style_id}</div></div>', unsafe_allow_html=True)
                st.markdown(f'<div style="{box_style}"><div style="{lbl_style}">Customer / Đối tác</div><div style="{val_style}">{customer}</div></div>', unsafe_allow_html=True)
                st.markdown(f'<div style="{box_style}"><div style="{lbl_style}">Season / Mùa sản xuất</div><div style="{val_style}">{season}</div></div>', unsafe_allow_html=True)
            with m_col2:
                st.markdown(f'<div style="{box_style}"><div style="{lbl_style}">Garment Type / Kiểu dáng</div><div style="{val_style}">{short_desc}</div></div>', unsafe_allow_html=True)
                st.markdown(f'<div style="{box_style}"><div style="{lbl_style}">Material Spec / Mô tả vải</div><div style="{val_style}" title="{fabric_type}">{fabric_type[:28]}...</div></div>', unsafe_allow_html=True)
                st.markdown(
                    f'<div style="{box_style} background-color: #f0fdf4; border-color: #bbf7d0;">'
                    f'  <div style="{lbl_style} color: #166534;">Techpack Status</div>'
                    f'  <div style="{val_style} color: #15803d; display: flex; align-items: center; gap: 6px;">🟢 READY TO BOM</div>'
                    f'</div>', 
                    unsafe_allow_html=True
                )
        else:
            if st.session_state.pdf_bytes is None:
                st.markdown("<div style='margin-top: 40px; text-align: center; color: #64748b; font-size: 13px; font-style: italic;'>Bảng tóm tắt hồ sơ trống. Vui lòng tải tài liệu lên hệ thống.</div>", unsafe_allow_html=True)

# --- CỘT PHẢI: KHÔNG GIAN HIỂN THỊ HÌNH ẢNH SKETCH ---
with col_right:
    with st.container(border=True, height=520):
        st.markdown("### 🎨 TECHPACK SKETCH VISUALIZER")
        
        # Hình ảnh phác thảo dạng bytes sau khi trích xuất từ PDF sẽ hiển thị sắc nét tại đây
        if "pdf_page_one_image" in st.session_state and st.session_state.pdf_page_one_image is not None:
            st.image(st.session_state.pdf_page_one_image, caption=f"Bản vẽ phác thảo trích xuất: {st.session_state.get('pdf_name', '')}", use_container_width=True)
        else:
            st.markdown("<div style='margin-top: 60px; text-align: center; color: #64748b; font-size: 13px; font-style: italic;'>Chưa có hình ảnh phác thảo. Vui lòng tải Techpack PDF để trích xuất hệ thống.</div>", unsafe_allow_html=True)






# =====================================================================
# 🧠 ĐOẠN A (VERSION V24.0): AI PURE SCAN - STRICT PIECE COUNT SYNC
# =====================================================================
@st.cache_data(
    show_spinner=False,
    ttl=3600,
    hash_funcs={bytes: lambda b: hashlib.sha256(b).hexdigest()},
)
def execute_final_gerber_pure_scan(
    pdf_bytes,
    current_query,
    active_width,
    target_size_cmd,
    raw_json_schema,
    prompt_agent_2,
):
    import copy
    import hashlib
    import json
    import re
    import fitz
    import google.generativeai as genai

    # =================================================================
    # 1. CHUẨN HÓA PDF INPUT
    # =================================================================
    if hasattr(pdf_bytes, "getvalue"):
        pdf_bytes = pdf_bytes.getvalue()

    if not isinstance(pdf_bytes, bytes):
        raise TypeError("Dữ liệu PDF đầu vào không đúng định dạng bytes hợp lệ!")

    full_pdf_raw_text = ""
    image_payloads = []

    with fitz.open(stream=pdf_bytes, filetype="pdf") as doc_recovery:

        total_pages = len(doc_recovery)

        for page_idx in range(total_pages):

            page_text = doc_recovery[page_idx].get_text("text")

            full_pdf_raw_text += (
                f"\n--- PAGE {page_idx + 1} ---\n"
                f"{page_text}"
            )

            # Chỉ gửi tối đa 2 trang hình ảnh để kiểm soát token
            if len(image_payloads) < 2:
                try:
                    pix = doc_recovery[page_idx].get_pixmap(
                        dpi=72,
                        colorspace=fitz.csRGB
                    )

                    image_payloads.append({
                        "mime_type": "image/jpeg",
                        "data": pix.tobytes("jpeg")
                    })

                except Exception:
                    continue

    # =================================================================
    # 2. GHÉP INPUT CHO GEMINI
    # =================================================================
    gemini_inputs = list(image_payloads)

    gemini_inputs.insert(
        0,
        f"""
=== USER CHAT COMMAND ===
{current_query}

=== TARGET SIZE ===
{target_size_cmd}

=== TECHPACK TEXT ===
{full_pdf_raw_text}
"""
    )

    extended_prompt = prompt_agent_2 + """

=====================================================================
🚨 MASTER PIECE COUNT RULE - ABSOLUTE
=====================================================================

For EVERY valid pattern component:

1. "cut_quantity" MUST represent the TOTAL PHYSICAL NUMBER OF PIECES
   required for ONE finished garment.

2. "piece_count" MUST be IDENTICAL to "cut_quantity".

3. NEVER default piece_count to 1 when cut_quantity is available.

4. Examples:
   - Front Left + Front Right = cut_quantity 2
   - Back Left + Back Right = cut_quantity 2
   - Sleeve pair = cut_quantity 2
   - Collar pair if two physical pieces = cut_quantity 2
   - Waistband cut as one continuous piece = cut_quantity 1
   - Pocket pair = cut_quantity 2
   - Single pocket = cut_quantity 1

5. If a component is explicitly marked:
   mirror_piece = true
   OR
   is_left_right_pair = true

   then normally cut_quantity should be at least 2,
   UNLESS the Tech Pack clearly states that the pattern is cut on fold
   or one physical piece produces both sides.

6. "mirror_piece" describes geometry relationship.
   It does NOT mean that the row itself should be duplicated.

7. DO NOT combine left/right physical pieces into one geometry width.

8. DO NOT reduce cut_quantity to 1 merely because only one CAD
   geometry description is shown.

=====================================================================
🚨 GEOMETRY RULE
=====================================================================

"bounding_box_width" = width of ONE physical pattern piece.

Never combine left + right pieces into one width.

Example:
If left front = 12"
and right front = 12"

Then:
bounding_box_width = 12"
cut_quantity = 2

NOT:
bounding_box_width = 24"
cut_quantity = 1

=====================================================================
🚨 IMPORTANT
=====================================================================

Always return cut_quantity whenever a physical pattern component
can be identified.

Do not omit cut_quantity.
Do not use 1 as a generic fallback if the Tech Pack indicates a pair.
"""

    gemini_inputs.append(extended_prompt)

    # =================================================================
    # 3. GỌI GEMINI
    # =================================================================
    model = genai.GenerativeModel("gemini-2.5-flash")

    response = model.generate_content(
        gemini_inputs,
        generation_config={
            "response_mime_type": "application/json",
            "response_schema": raw_json_schema,
            "temperature": 0.0,
        },
        request_options={"timeout": 120.0},
    )

    if not response or not response.text:
        raise RuntimeError("Mô hình Gemini trả về kết quả rỗng!")

    txt = response.text.strip()

    if txt.startswith("```"):
        txt = re.sub(r"^```json\s*", "", txt)
        txt = re.sub(r"^```\s*", "", txt)
        txt = re.sub(r"\s*```$", "", txt)

    txt = txt.strip()

    try:
        blueprint_worker = json.loads(txt)

    except json.JSONDecodeError as json_err:
        raise RuntimeError(
            "Mô hình Gemini trả về JSON không hợp lệ:\n\n"
            + txt
        ) from json_err

    # =================================================================
    # 4. NORMALIZE BOM
    # =================================================================
    if blueprint_worker and "bom_rows" in blueprint_worker:

        blueprint_worker["calculated_on_size"] = target_size_cmd

        for row in blueprint_worker.get("bom_rows", []):

            # ---------------------------------------------------------
            # COMPONENT NAME
            # ---------------------------------------------------------
            if "component_name" in row:
                row["component_name"] = " ".join(
                    str(row["component_name"]).upper().split()
                )

            # ---------------------------------------------------------
            # GEOMETRY
            # ---------------------------------------------------------
            try:
                row["bounding_box_length"] = round(
                    float(row.get("bounding_box_length", 0.0)),
                    2
                )
            except Exception:
                row["bounding_box_length"] = 0.0

            try:
                row["bounding_box_width"] = round(
                    float(row.get("bounding_box_width", 0.0)),
                    2
                )
            except Exception:
                row["bounding_box_width"] = 0.0

            try:
                row["polygon_net_area"] = float(
                    row.get("polygon_net_area", 0.0)
                )
            except Exception:
                row["polygon_net_area"] = 0.0

            # ---------------------------------------------------------
            # MATERIAL CLASS
            # ---------------------------------------------------------
            comp_name = str(
                row.get("component_name", "")
            ).upper().strip()

            mat_class = str(
                row.get("material_class", "FABRIC")
            ).upper().strip()

            if any(k in comp_name for k in [
                "FUSING",
                "INTERLINING",
                "MEX",
                "MẾCH",
                "DỰNG",
                "KEO",
                "KEO LOT"
            ]):
                mat_class = "FUSING"

            elif any(k in comp_name for k in [
                "LINING",
                "POCKET BAG",
                "LOT TUI",
                "LÓT",
                "RIB",
                "BO GÂN"
            ]):
                mat_class = "LINING"

            elif any(k in comp_name for k in [
                "CONTRAST",
                "PHỐI"
            ]):
                mat_class = "CONTRAST"

            else:
                mat_class = "FABRIC"

            row["material_class"] = mat_class

            # ---------------------------------------------------------
            # 🔥 MASTER PIECE COUNT
            # ---------------------------------------------------------
            # cut_quantity là nguồn dữ liệu chính
            raw_cut_qty = row.get("cut_quantity", None)

            # Nếu AI có piece_count nhưng thiếu cut_quantity
            if raw_cut_qty is None:
                raw_cut_qty = row.get("piece_count", None)

            try:
                cut_qty = int(float(raw_cut_qty))
            except Exception:
                cut_qty = 0

            # ---------------------------------------------------------
            # Nếu AI không trả quantity, kiểm tra pair/mirror
            # ---------------------------------------------------------
            if cut_qty <= 0:

                is_pair = bool(
                    row.get("is_left_right_pair", False)
                )

                is_mirror = bool(
                    row.get("mirror_piece", False)
                )

                fold_type = str(
                    row.get("fold_type", "")
                ).upper().strip()

                # Có pair/mirror → 2
                if (is_pair or is_mirror) and fold_type not in [
                    "ON_FOLD",
                    "CENTER_FOLD"
                ]:
                    cut_qty = 2

                else:
                    cut_qty = 1

            # Không cho quantity < 1
            cut_qty = max(1, cut_qty)

            # 🔒 KHÓA ĐỒNG BỘ TUYỆT ĐỐI
            row["cut_quantity"] = cut_qty
            row["piece_count"] = cut_qty

            # ---------------------------------------------------------
            # 🔥 SINGLE PIECE WIDTH
            # ---------------------------------------------------------
            # KHÔNG nhân/chia piece_count dựa trên width.
            # Width và quantity là 2 khái niệm độc lập.
            #
            # Ví dụ:
            # Front L/R = width 12", quantity 2
            #
            # Không được biến:
            # width 24", quantity 1
            # ---------------------------------------------------------

            bbox_area = (
                row["bounding_box_length"]
                * row["bounding_box_width"]
            )

            if (
                row["polygon_net_area"] > bbox_area
                and bbox_area > 0
            ):
                row["polygon_net_area"] = (
                    bbox_area *
                    (0.76 if mat_class == "FABRIC" else 0.85)
                )

            # ---------------------------------------------------------
            # GROSS CONSUMPTION
            # ---------------------------------------------------------
            try:
                row["gross_consumption"] = round(
                    float(row.get("gross_consumption", 0.0)),
                    4
                )
            except Exception:
                row["gross_consumption"] = 0.0

            # ---------------------------------------------------------
            # MARKER EFFICIENCY
            # ---------------------------------------------------------
            try:
                row["marker_efficiency"] = str(
                    row.get(
                        "marker_efficiency",
                        "82.5%"
                    )
                ).strip()

            except Exception:
                row["marker_efficiency"] = "82.5%"

            # ---------------------------------------------------------
            # FABRIC WIDTH
            # ---------------------------------------------------------
            try:

                forced_width = float(active_width)

                if current_query:

                    width_match = re.search(
                        r"(khổ\s*vải|khổ)\s*(\d+(\.\d+)?)",
                        str(current_query),
                        re.IGNORECASE
                    )

                    if width_match:
                        forced_width = float(
                            width_match.group(2)
                        )

                row["fabric_width_inch"] = forced_width

            except Exception:
                row["fabric_width_inch"] = float(active_width)

    # =================================================================
    # 5. API COUNTER
    # =================================================================
    if "api_calls_count" not in st.session_state:
        st.session_state["api_calls_count"] = 0

    if "tokens_consumed" not in st.session_state:
        st.session_state["tokens_consumed"] = 0

    st.session_state["api_calls_count"] += 1

    st.session_state["tokens_consumed"] += (
        len(str(full_pdf_raw_text)) // 4
    )

    return blueprint_worker

import io
import re
import hashlib
import numpy as np
import pandas as pd
import streamlit as st

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# =====================================================================
# 🟩 MASTER CHAT WORKSPACE & PARAMS SYNC
# VERSION V30.0
#
# FLOW:
#
# USER CHAT
#    ↓
# last_submitted_query
#    ↓
# PARAMETER EXTRACTION
#    ↓
# MASTER SESSION STATE
#    ↓
# BOM DATA
#    ↓
# IE CALCULATION ENGINE
#
# ĐOẠN NÀY LÀ ĐOẠN DUY NHẤT XỬ LÝ CHAT + PARAMETER SYNC
# =====================================================================


# =====================================================================
# 1. SESSION STATE SAFETY
# =====================================================================

if "chat_history" not in st.session_state:
    st.session_state["chat_history"] = []

if "ai_processing" not in st.session_state:
    st.session_state["ai_processing"] = False

if "last_submitted_query" not in st.session_state:
    st.session_state["last_submitted_query"] = ""

if (
    "bom_data" not in st.session_state
    or not isinstance(st.session_state.get("bom_data"), dict)
):
    st.session_state["bom_data"] = {}

ctx = st.session_state["bom_data"]


# =====================================================================
# 2. SAFE NUMBER EXTRACTOR
# =====================================================================

def extract_number(pattern, text):
    """
    Trả về số được tìm thấy.
    Không tìm thấy -> None.
    """

    if not text:
        return None

    match = re.search(
        pattern,
        text,
        re.IGNORECASE
    )

    if not match:
        return None

    try:
        return float(match.group(1))
    except (TypeError, ValueError):
        return None


# =====================================================================
# 3. CHAT HISTORY DISPLAY
# =====================================================================

chat_history_container = st.container()

with chat_history_container:

    st.markdown(
        '<br>'
        '<div class="cad-card">'
        '<div class="cad-header">'
        '💬 CHATGPT IE COLLABORATION WORKSPACE'
        '</div>'
        '</div>',
        unsafe_allow_html=True
    )

    if st.session_state.get("chat_history"):

        for msg in st.session_state["chat_history"]:

            if not isinstance(msg, dict):
                continue

            user_msg = str(msg.get("user", "") or "")
            ai_msg = str(msg.get("ai", "") or "")

            if user_msg:

                with st.chat_message("user"):
                    st.write(user_msg)

            if ai_msg:

                with st.chat_message("assistant"):
                    st.write(ai_msg)


# =====================================================================
# 4. CHAT INPUT
# =====================================================================

safe_user_prompt = st.chat_input(
    "Gõ lệnh tính toán "
    "(Ví dụ: tính định mức cỡ 32 khổ 56 co rút dọc 3 ngang 14)...",
    key="ie_workspace_fixed_dynamic_chat_final_patch_v30_0"
)


# =====================================================================
# 5. KHI USER NHẬP LỆNH
# =====================================================================

if safe_user_prompt:

    query_text = str(
        safe_user_prompt
    ).strip()

    if not query_text:
        st.stop()

    query_lower = query_text.lower()

    st.session_state["last_submitted_query"] = query_text
    st.session_state["ai_processing"] = True


    # ================================================================
    # A. FABRIC WIDTH
    #
    # Hỗ trợ:
    #   khổ 56
    #   khổ vải 56
    #   khổ chính 56
    #   width 56
    # ================================================================

    width_from_chat = extract_number(
        r'\b(?:khổ\s*vải|khổ\s*chính|khổ|kho|width)'
        r'\s*[:=-]?\s*(\d+(?:\.\d+)?)\b',
        query_lower
    )

    if (
        width_from_chat is not None
        and width_from_chat > 0
    ):

        detected_width = width_from_chat

    else:

        detected_width = float(
            st.session_state.get(
                "current_active_width",
                st.session_state.get(
                    "fabric_width_inch",
                    ctx.get(
                        "fabric_width_inch",
                        58.0
                    )
                )
            )
        )

        if detected_width <= 0:
            detected_width = 58.0


    # MASTER FABRIC WIDTH
    st.session_state["current_active_width"] = detected_width
    st.session_state["fabric_width_inch"] = detected_width

    ctx["fabric_width_inch"] = detected_width


    # ================================================================
    # B. FUSING WIDTH
    # ================================================================

    fusing_from_chat = extract_number(
        r'\b(?:khổ\s*keo|keo\s*khổ|khổ\s*dựng)'
        r'\s*[:=-]?\s*(\d+(?:\.\d+)?)\b',
        query_lower
    )

    if (
        fusing_from_chat is not None
        and fusing_from_chat > 0
    ):

        fusing_width = fusing_from_chat

    else:

        fusing_width = float(
            st.session_state.get(
                "fusing_width_inch",
                ctx.get(
                    "fusing_width_inch",
                    59.0
                )
            )
        )

        if fusing_width <= 0:
            fusing_width = 59.0


    st.session_state["fusing_width_inch"] = fusing_width
    ctx["fusing_width_inch"] = fusing_width


    # ================================================================
    # C. LINING WIDTH
    # ================================================================

    lining_from_chat = extract_number(
        r'\b(?:khổ\s*lót|lót\s*khổ|vải\s*lót\s*khổ)'
        r'\s*[:=-]?\s*(\d+(?:\.\d+)?)\b',
        query_lower
    )

    if (
        lining_from_chat is not None
        and lining_from_chat > 0
    ):

        lining_width = lining_from_chat

    else:

        lining_width = float(
            st.session_state.get(
                "lining_width_inch",
                ctx.get(
                    "lining_width_inch",
                    57.0
                )
            )
        )

        if lining_width <= 0:
            lining_width = 57.0


    st.session_state["lining_width_inch"] = lining_width
    ctx["lining_width_inch"] = lining_width


    # ================================================================
    # D. SIZE
    #
    # Hỗ trợ:
    #   cỡ 32
    #   size 32
    #   kích cỡ 32
    # ================================================================

    size_match = re.search(
        r'\b(?:cỡ|size|kích\s*cỡ)'
        r'\s*[:=-]?\s*([a-zA-Z0-9]+)\b',
        query_lower,
        re.IGNORECASE
    )

    if size_match:

        detected_size_code = (
            str(size_match.group(1))
            .upper()
            .strip()
        )

    else:

        size_candidates = [
            st.session_state.get("current_active_size"),
            st.session_state.get("target_size"),
            ctx.get("detected_base_size"),
            ctx.get("base_size"),
            ctx.get("calculated_on_size"),
            "32"
        ]

        detected_size_code = "32"

        for candidate in size_candidates:

            if candidate is None:
                continue

            candidate_text = (
                str(candidate)
                .strip()
            )

            if candidate_text:

                detected_size_code = (
                    candidate_text
                    .upper()
                    .strip()
                )

                break


    # ================================================================
    # SIZE COMPLEX
    #
    # 32X33 -> 32
    # 32 x 33 -> 32
    # ================================================================

    detected_size_code = re.split(
        r'\s*[xX×]\s*',
        detected_size_code
    )[0].strip()


    # MASTER SIZE
    st.session_state["current_active_size"] = detected_size_code
    st.session_state["target_size"] = detected_size_code
    st.session_state["detected_base_size"] = detected_size_code

    ctx["calculated_on_size"] = detected_size_code
    ctx["detected_base_size"] = detected_size_code


    # ================================================================
    # E. WARP / VERTICAL SHRINKAGE
    #
    # Hỗ trợ:
    #   co rút dọc 3
    #   co dọc 3
    #   dọc 3
    #   vertical 3
    # ================================================================

    warp_from_chat = extract_number(
        r'\b(?:co\s*rút\s*dọc|co\s*dọc|dọc|shrink_v|vertical)'
        r'\s*[:=-]?\s*(-?\d+(?:\.\d+)?)\s*%?\b',
        query_lower
    )

    if warp_from_chat is not None:

        warp_shrink = warp_from_chat

    else:

        # KHÔNG RESET VỀ 0
        warp_shrink = float(
            st.session_state.get(
                "warp_shrinkage",
                st.session_state.get(
                    "current_warp_shrinkage",
                    st.session_state.get(
                        "shrinkage_vertical",
                        0.0
                    )
                )
            )
        )


    # ================================================================
    # F. WEFT / HORIZONTAL SHRINKAGE
    # ================================================================

    weft_from_chat = extract_number(
        r'\b(?:co\s*rút\s*ngang|co\s*ngang|ngang|shrink_h|horizontal)'
        r'\s*[:=-]?\s*(-?\d+(?:\.\d+)?)\s*%?\b',
        query_lower
    )

    if weft_from_chat is not None:

        weft_shrink = weft_from_chat

    else:

        # KHÔNG RESET VỀ 0
        weft_shrink = float(
            st.session_state.get(
                "weft_shrinkage",
                st.session_state.get(
                    "current_weft_shrinkage",
                    st.session_state.get(
                        "shrinkage_horizontal",
                        0.0
                    )
                )
            )
        )


    # ================================================================
    # G. ĐỒNG BỘ TOÀN BỘ TÊN BIẾN SHRINKAGE
    # ================================================================

    # Vertical / Warp
    st.session_state["warp_shrinkage"] = warp_shrink
    st.session_state["current_warp_shrinkage"] = warp_shrink
    st.session_state["shrinkage_vertical"] = warp_shrink

    # Horizontal / Weft
    st.session_state["weft_shrinkage"] = weft_shrink
    st.session_state["current_weft_shrinkage"] = weft_shrink
    st.session_state["shrinkage_horizontal"] = weft_shrink


    # ================================================================
    # H. MASTER SNAPSHOT
    # ================================================================

    st.session_state["ie_master_params"] = {

        "size": detected_size_code,

        "fabric_width": detected_width,

        "fusing_width": fusing_width,

        "lining_width": lining_width,

        "warp_shrinkage": warp_shrink,

        "weft_shrinkage": weft_shrink,
    }


    # ================================================================
    # I. ĐỒNG BỘ BOM DATA
    # ================================================================

    ctx["fabric_width_inch"] = detected_width
    ctx["fusing_width_inch"] = fusing_width
    ctx["lining_width_inch"] = lining_width

    ctx["calculated_on_size"] = detected_size_code
    ctx["detected_base_size"] = detected_size_code

    ctx["warp_shrinkage"] = warp_shrink
    ctx["weft_shrinkage"] = weft_shrink

    st.session_state["bom_data"] = ctx


    # ================================================================
    # J. TẠO THÔNG BÁO AUDIT
    # ================================================================

    parameter_summary = (
        f"Size = {detected_size_code} | "
        f"Khổ = {detected_width:g}\" | "
        f"Co dọc = {warp_shrink:g}% | "
        f"Co ngang = {weft_shrink:g}%"
    )

    st.session_state["chat_history"].append({
        "user": query_text,
        "ai": (
            "⚙️ IE Engine đã nhận và đồng bộ:\n\n"
            f"{parameter_summary}\n\n"
            "🔄 Đang kích hoạt tái tính định mức."
        )
    })


    # ================================================================
    # K. RESET PIPELINE
    # ================================================================

    st.session_state["pipeline_auto_run_executed"] = False

    st.session_state["ie_parameter_sync_complete"] = True


    # ================================================================
    # L. RERUN
    # ================================================================

    st.rerun()



# =====================================================================
# 🟩 ĐOẠN 2 - VERSION V28.5
# MASTER AI SCAN + PRODUCT TYPE VALIDATION + WIDTH/SHRINK SYNC
# =====================================================================

import re
import streamlit as st


if st.session_state.ai_processing:

    current_query = st.session_state["last_submitted_query"]

    active_pdf = (
        st.session_state.get("pdf_bytes")
        or st.session_state.get("uploaded_file")
        or st.session_state.get("current_pdf")
        or st.session_state.get("pdf_data")
    )

    # ================================================================
    # 1. ĐỌC THÔNG SỐ TỪ CHAT
    # ================================================================

    dynamic_width = 58.0
    target_size = "32"

    warp_shrinkage = 0.0
    weft_shrinkage = 0.0

    if current_query:

        query_text = str(current_query)

        # ------------------------------------------------------------
        # KHỔ VẢI
        # ------------------------------------------------------------
        w_m = re.search(
            r"(?:khổ\s*vải|khổ)\s*[:=]?\s*(\d+(?:\.\d+)?)",
            query_text,
            re.IGNORECASE
        )

        if w_m:
            dynamic_width = float(w_m.group(1))

        # ------------------------------------------------------------
        # SIZE
        # ------------------------------------------------------------
        s_m = re.search(
            r"(?:cỡ|size)\s*[:=]?\s*([a-zA-Z0-9]+)",
            query_text,
            re.IGNORECASE
        )

        if s_m:
            target_size = str(
                s_m.group(1)
            ).upper().strip()

        # ------------------------------------------------------------
        # CO DỌC
        # ------------------------------------------------------------
        warp_m = re.search(
            r"(?:co\s*rút\s*dọc|co\s*dọc|độ\s*co\s*dọc)\s*[:=]?\s*(\d+(?:\.\d+)?)\s*%?",
            query_text,
            re.IGNORECASE
        )

        if warp_m:
            val = float(warp_m.group(1))

            if 0 <= val <= 15:
                warp_shrinkage = val

        # ------------------------------------------------------------
        # CO NGANG
        # ------------------------------------------------------------
        weft_m = re.search(
            r"(?:co\s*rút\s*ngang|co\s*ngang|độ\s*co\s*ngang)\s*[:=]?\s*(\d+(?:\.\d+)?)\s*%?",
            query_text,
            re.IGNORECASE
        )

        if weft_m:
            val = float(weft_m.group(1))

            if 0 <= val <= 15:
                weft_shrinkage = val


    # ================================================================
    # 2. AI SCAN
    # ================================================================

    if active_pdf is not None:

        with st.spinner(
            "🧠 AI Vision đang quét phôi rập Nguyên Liệu..."
        ):

            try:

                # ====================================================
                # 3. JSON SCHEMA
                # ====================================================

                raw_json_schema = {

                    "type": "OBJECT",

                    "properties": {

                        "detected_product_type": {
                            "type": "STRING"
                        },

                        "detected_base_size": {
                            "type": "STRING"
                        },

                        "bom_rows": {

                            "type": "ARRAY",

                            "items": {

                                "type": "OBJECT",

                                "properties": {

                                    "component_name": {
                                        "type": "STRING"
                                    },

                                    "bounding_box_length": {
                                        "type": "NUMBER"
                                    },

                                    "bounding_box_width": {
                                        "type": "NUMBER"
                                    },

                                    "piece_shape": {
                                        "type": "STRING"
                                    },

                                    "piece_function": {
                                        "type": "STRING"
                                    },

                                    "fold_type": {
                                        "type": "STRING"
                                    },

                                    "material_zone": {
                                        "type": "STRING",
                                        "enum": [
                                            "SELF",
                                            "LINING",
                                            "FUSING",
                                            "RIB",
                                            "CONTRAST"
                                        ]
                                    },

                                    "grain_constraint": {
                                        "type": "STRING"
                                    },

                                    "packing_priority": {
                                        "type": "INTEGER"
                                    },

                                    "convex_fill_ratio": {
                                        "type": "NUMBER"
                                    },

                                    "seam_allowance": {
                                        "type": "STRING"
                                    },

                                    "mirror_piece": {
                                        "type": "BOOLEAN"
                                    },

                                    "is_left_right_pair": {
                                        "type": "BOOLEAN"
                                    },

                                    "requires_matching": {
                                        "type": "BOOLEAN"
                                    },

                                    "critical_alignment": {
                                        "type": "STRING"
                                    },

                                    "cut_quantity": {
                                        "type": "INTEGER"
                                    },

                                    "grain_direction": {
                                        "type": "STRING"
                                    },

                                    "rotation_allowed": {
                                        "type": "STRING"
                                    },

                                    "edge_curvature": {
                                        "type": "STRING"
                                    },

                                    "shape_complexity": {
                                        "type": "STRING"
                                    },

                                    "inference_source": {
                                        "type": "STRING"
                                    },

                                    "cad_reconstruction_score": {
                                        "type": "INTEGER"
                                    },

                                    "field_confidence": {

                                        "type": "OBJECT",

                                        "properties": {

                                            "dimensions": {
                                                "type": "STRING"
                                            },

                                            "geometry_shape": {
                                                "type": "STRING"
                                            },

                                            "grain_alignment": {
                                                "type": "STRING"
                                            }
                                        },

                                        "required": [
                                            "dimensions",
                                            "geometry_shape",
                                            "grain_alignment"
                                        ]
                                    },

                                    "shape_parameters": {

                                        "type": "OBJECT",

                                        "properties": {

                                            "estimated_corner_points": {
                                                "type": "INTEGER"
                                            },

                                            "dominant_axis": {
                                                "type": "STRING"
                                            },

                                            "top_width_ratio": {
                                                "type": "NUMBER"
                                            },

                                            "bottom_width_ratio": {
                                                "type": "NUMBER"
                                            },

                                            "left_edge_profile": {
                                                "type": "STRING"
                                            },

                                            "right_edge_profile": {
                                                "type": "STRING"
                                            },

                                            "waist_curve_depth": {
                                                "type": "NUMBER"
                                            },

                                            "hem_curve_depth": {
                                                "type": "NUMBER"
                                            },

                                            "crotch_projection_ratio": {
                                                "type": "NUMBER"
                                            }
                                        }
                                    }
                                },

                                "required": [
                                    "component_name",
                                    "bounding_box_length",
                                    "bounding_box_width",
                                    "piece_shape",
                                    "piece_function",
                                    "fold_type",
                                    "material_zone",
                                    "packing_priority",
                                    "convex_fill_ratio",
                                    "mirror_piece"
                                ]
                            }
                        }
                    },

                    "required": [
                        "detected_product_type",
                        "detected_base_size",
                        "bom_rows"
                    ]
                }


                # ====================================================
                # 4. MASTER PRODUCT TYPE PROMPT
                # ====================================================

                prompt_agent_2 = f"""

You are a senior Industrial Garment IE & CAD Pattern Engineering Intelligence.

============================================================
PRODUCT TYPE IDENTIFICATION - ABSOLUTE PRIORITY
============================================================

You MUST identify the ACTUAL GARMENT TYPE from the Tech Pack,
garment sketch, style description and construction.

DO NOT infer product type from the word "JEAN" alone.

The following distinctions are mandatory:

------------------------------------------------------------
SHORT
------------------------------------------------------------

If the garment is a SHORT / SHORTS / BERMUDA:

Output:

SHORT

Examples:

JEAN SHORT
DENIM SHORT
DENIM SHORTS
JEAN SHORTS
WALK SHORT
BERMUDA SHORT

ALL OF THESE MUST BE:

SHORT

NOT:

JEAN
JEAN_LONG
PANT
TROUSER

------------------------------------------------------------
JEAN_LONG
------------------------------------------------------------

Use JEAN_LONG ONLY when the garment is clearly a
FULL-LENGTH / LONG-LEG JEAN.

Examples:

LONG JEANS
FULL LENGTH JEANS
LONG DENIM PANTS
JEAN LONG PANTS

These may be:

JEAN_LONG

------------------------------------------------------------
JEAN
------------------------------------------------------------

Use JEAN only when the product is identified as
JEAN/DENIM but the document does not clearly establish
that it is SHORT or FULL-LENGTH.

------------------------------------------------------------
PANT
------------------------------------------------------------

Use PANT for ordinary long pants when it is not specifically
identified as JEAN/DENIM.

------------------------------------------------------------
CRITICAL RULE
------------------------------------------------------------

The word "JEAN" by itself does NOT mean JEAN_LONG.

If the product name contains both:

JEAN + SHORT

or:

DENIM + SHORT

the final product type MUST be:

SHORT

============================================================

Reconstruct the multi-layered CAD metadata for EVERY valid
fabric/fusing piece in the Tech Pack for Size {target_size}.

============================================================
ACCESSORY OMISSION
============================================================

NEVER extract:

buttons,
sewing threads,
zippers,
sliders,
rivets,
labels,
care labels,
size tabs,
hangtags,
polybags,
metal accessories,
plastic accessories.

ONLY extract:

SELF,
LINING,
FUSING,
RIB,
CONTRAST.

============================================================
SINGLE PIECE RULE
============================================================

'bounding_box_width' MUST represent ONE SINGLE physical piece.

NEVER combine left and right pieces.

============================================================
GEOMETRY
============================================================

Extract:

bounding_box_length
bounding_box_width
piece_shape
piece_function
fold_type
material_zone
packing_priority
convex_fill_ratio
mirror_piece
cut_quantity
grain_direction
rotation_allowed
edge_curvature
shape_complexity

All dimensions must be in inches.

NEVER output zero dimensions for a valid pattern piece.

============================================================
VALIDATION
============================================================

Skip rows that are not actual pattern pieces.

The output must contain the actual product type.
"""


                # ====================================================
                # 5. CALL AI
                # ====================================================

                bom_data = execute_final_gerber_pure_scan(

                    pdf_bytes=active_pdf,

                    current_query=current_query,

                    active_width=dynamic_width,

                    target_size_cmd=target_size,

                    raw_json_schema=raw_json_schema,

                    prompt_agent_2=prompt_agent_2
                )


                # ====================================================
                # 6. MASTER PRODUCT TYPE VALIDATION
                # ====================================================

                if not isinstance(bom_data, dict):
                    raise RuntimeError(
                        "AI không trả về bom_data dạng dictionary."
                    )


                ai_product_raw = str(
                    bom_data.get(
                        "detected_product_type",
                        ""
                    )
                ).upper().strip()


                # ----------------------------------------------------
                # Lấy thêm tên style / query để kiểm tra SHORT
                # ----------------------------------------------------

                product_scan_text = " ".join(
                    [
                        ai_product_raw,
                        str(current_query or "").upper(),
                        str(
                            bom_data.get(
                                "style_code",
                                ""
                            )
                        ).upper()
                    ]
                )


                # ====================================================
                # 7. PRODUCT TYPE ALIAS
                # ====================================================

                product_alias = {

                    "SHORTS": "SHORT",
                    "BERMUDA": "SHORT",
                    "BERMUDAS": "SHORT",

                    "WALK SHORT": "SHORT",
                    "WALK SHORTS": "SHORT",

                    "DENIM SHORT": "SHORT",
                    "DENIM SHORTS": "SHORT",

                    "JEAN SHORT": "SHORT",
                    "JEAN SHORTS": "SHORT",

                    "JEANS SHORT": "SHORT",
                    "JEANS SHORTS": "SHORT",

                    "LONG JEAN": "JEAN_LONG",
                    "LONG JEANS": "JEAN_LONG",

                    "JEAN LONG": "JEAN_LONG",
                    "JEANS LONG": "JEAN_LONG",

                    "FULL LENGTH JEAN": "JEAN_LONG",
                    "FULL LENGTH JEANS": "JEAN_LONG",

                    "PANTS": "PANT",
                    "TROUSERS": "TROUSER",

                    "T-SHIRT": "TSHIRT",
                    "T SHIRT": "TSHIRT",
                    "TEE SHIRT": "TSHIRT"
                }


                normalized_product_type = product_alias.get(
                    ai_product_raw,
                    ai_product_raw
                )


                # ====================================================
                # 8. SHORT OVERRIDE - CAO NHẤT
                # ====================================================

                short_keywords = [

                    "SHORT",

                    "SHORTS",

                    "BERMUDA",

                    "WALK SHORT",

                    "WALKING SHORT",

                    "DENIM SHORT",

                    "DENIM SHORTS",

                    "JEAN SHORT",

                    "JEAN SHORTS",

                    "JEANS SHORT",

                    "JEANS SHORTS"
                ]


                is_short = any(
                    keyword in product_scan_text
                    for keyword in short_keywords
                )


                if is_short:

                    final_product_type = "SHORT"

                elif normalized_product_type in [
                    "JEAN_LONG",
                    "JEAN",
                    "PANT",
                    "TROUSER",
                    "KHAKI",
                    "JACKET",
                    "COAT",
                    "BLAZER",
                    "SUIT",
                    "SHIRT",
                    "BLOUSE",
                    "POLO",
                    "TEE",
                    "TSHIRT",
                    "TANK",
                    "DRESS",
                    "SKIRT",
                    "OVERALL",
                    "COVERALL",
                    "BIB",
                    "JUMPSUIT",
                    "DUNGAREE"
                ]:

                    final_product_type = normalized_product_type

                else:

                    # Không được fallback JEAN_LONG
                    final_product_type = "PANT"


                # ====================================================
                # 9. MASTER AI DECISION
                # ====================================================

                if "ai_expert_decision" not in bom_data:
                    bom_data["ai_expert_decision"] = {}

                if not isinstance(
                    bom_data["ai_expert_decision"],
                    dict
                ):
                    bom_data["ai_expert_decision"] = {}


                bom_data[
                    "ai_expert_decision"
                ][
                    "ai_product_type_raw"
                ] = ai_product_raw


                bom_data[
                    "ai_expert_decision"
                ][
                    "detected_product_type"
                ] = final_product_type


                bom_data[
                    "detected_product_type"
                ] = final_product_type


                bom_data[
                    "product_type"
                ] = final_product_type


                # ====================================================
                # 10. MASTER WIDTH / SHRINKAGE COMMIT
                # ====================================================

                bom_data[
                    "fabric_width_inch"
                ] = float(dynamic_width)

                bom_data[
                    "usable_width_inch"
                ] = float(dynamic_width)

                bom_data[
                    "calculated_on_size"
                ] = str(target_size)

                bom_data[
                    "warp_shrinkage_percent"
                ] = float(warp_shrinkage)

                bom_data[
                    "weft_shrinkage_percent"
                ] = float(weft_shrinkage)


                bom_data[
                    "ai_expert_decision"
                ][
                    "detected_base_size"
                ] = str(target_size)

                bom_data[
                    "ai_expert_decision"
                ][
                    "fabric_width"
                ] = float(dynamic_width)

                bom_data[
                    "ai_expert_decision"
                ][
                    "warp_shrinkage_percent"
                ] = float(warp_shrinkage)

                bom_data[
                    "ai_expert_decision"
                ][
                    "weft_shrinkage_percent"
                ] = float(weft_shrinkage)


                # ====================================================
                # 11. SESSION MASTER SYNC
                # ====================================================

                st.session_state[
                    "bom_data"
                ] = bom_data

                st.session_state[
                    "current_active_width"
                ] = float(dynamic_width)

                st.session_state[
                    "current_active_size"
                ] = str(target_size)

                st.session_state[
                    "current_warp_shrinkage"
                ] = float(warp_shrinkage)

                st.session_state[
                    "current_weft_shrinkage"
                ] = float(weft_shrinkage)

                # ====================================================
                # 12. DEBUG
                # ====================================================

                print(
                    "\n"
                    "====================================================\n"
                    "[AI PRODUCT TYPE MASTER]\n"
                    f"AI RAW        : {ai_product_raw}\n"
                    f"NORMALIZED    : {normalized_product_type}\n"
                    f"SHORT CHECK   : {is_short}\n"
                    f"FINAL TYPE    : {final_product_type}\n"
                    f"SIZE          : {target_size}\n"
                    f"WIDTH         : {dynamic_width}\"\n"
                    f"SHRINK WARP   : {warp_shrinkage}%\n"
                    f"SHRINK WEFT   : {weft_shrinkage}%\n"
                    "===================================================="
                )


                # ====================================================
                # 13. FINISH
                # ====================================================

                st.session_state.ai_processing = False

                st.rerun()


            except Exception as e:

                st.error(
                    f"❌ Lỗi xử lý luồng AI Execute (Đoạn 2): {str(e)}"
                )

                st.session_state.ai_processing = False

                st.rerun()


# =====================================================================
# 🧠 MASTER PARAMETER CONTROLLER V27.5
# 🔒 CHAT COMMAND = NGUỒN SỰ THẬT TUYỆT ĐỐI
# 🔒 KHỔ VẢI / SIZE / CO DỌC / CO NGANG KHÔNG ĐƯỢC CACHE AI GHI ĐÈ
# =====================================================================

import re
import streamlit as st


def initialize_and_sync_parameters():
    """
    MASTER CONTROLLER

    Ưu tiên tuyệt đối:
        CHAT COMMAND
            ↓
        current_active_width
        current_active_size
        current_warp_shrinkage
        current_weft_shrinkage
            ↓
        bom_data
            ↓
        IE ENGINE 5.2B1/B2

    Không cho giá trị cũ 58" từ bom_data/cache AI ghi đè
    giá trị mới người dùng nhập trong Chat.
    """

    # ================================================================
    # 0. KIỂM TRA NGUỒN BOM
    # ================================================================
    if not (
        st.session_state.get("bom_data")
        or st.session_state.get("accumulated_bom_rows")
    ):
        return None, None

    bom_source = st.session_state.get("bom_data", {})

    if not isinstance(bom_source, dict):
        bom_source = {}

    # ================================================================
    # 1. GIỮ NGUYÊN BỘ NÃO VIRTUAL PIECES
    # ================================================================
    ai_expert_decision = bom_source.get(
        "ai_expert_decision",
        {}
    )

    if not isinstance(ai_expert_decision, dict):
        ai_expert_decision = {}

    virtual_pieces_layer = ai_expert_decision.get(
        "virtual_pieces_layer",
        {}
    )

    if not isinstance(virtual_pieces_layer, dict):
        virtual_pieces_layer = {}

    # ================================================================
    # 2. LẤY CHAT COMMAND
    # ================================================================
    user_query_text = str(
        st.session_state.get(
            "last_submitted_query",
            ""
        )
    ).strip()

    # ================================================================
    # 3. LẤY GIÁ TRỊ HIỆN TẠI
    #    CHỈ DÙNG LÀM FALLBACK
    # ================================================================
    try:
        fabric_width = float(
            st.session_state.get(
                "current_active_width",
                bom_source.get(
                    "fabric_width_inch",
                    58.0
                )
            )
        )
    except:
        fabric_width = 58.0

    try:
        warp_shrinkage = float(
            st.session_state.get(
                "current_warp_shrinkage",
                bom_source.get(
                    "warp_shrinkage_percent",
                    0.0
                )
            )
        )
    except:
        warp_shrinkage = 0.0

    try:
        weft_shrinkage = float(
            st.session_state.get(
                "current_weft_shrinkage",
                bom_source.get(
                    "weft_shrinkage_percent",
                    0.0
                )
            )
        )
    except:
        weft_shrinkage = 0.0

    detected_size = st.session_state.get(
        "current_active_size",
        bom_source.get(
            "detected_base_size",
            bom_source.get(
                "calculated_on_size",
                "32"
            )
        )
    )

    target_size = str(
        detected_size
    ).strip().upper()

    if not target_size:
        target_size = "32"

    # ================================================================
    # 4. CHAT PARSER - MASTER
    #
    # Hỗ trợ:
    #   khổ 62
    #   khổ vải 62
    #   khổ=62
    #   khổ vải=62
    #   khổ 56
    #
    # Và tránh bắt nhầm số của co rút / size.
    # ================================================================
    if user_query_text:

        query = " ".join(
            user_query_text.lower().split()
        )

        # ------------------------------------------------------------
        # KHỔ VẢI
        # ------------------------------------------------------------
        width_match = re.search(
            r"\bkhổ(?:\s+vải)?\s*[:=]?\s*(\d+(?:\.\d+)?)\b",
            query,
            re.IGNORECASE
        )

        if width_match:

            parsed_width = float(
                width_match.group(1)
            )

            # Khổ vải kỹ thuật hợp lệ
            if 20.0 <= parsed_width <= 100.0:

                fabric_width = parsed_width

                print(
                    f"[MASTER CHAT SYNC] "
                    f"KHỔ CHAT = {fabric_width:.2f}\""
                )

        # ------------------------------------------------------------
        # CO DỌC
        # ------------------------------------------------------------
        warp_match = re.search(
            r"\b(?:co\s*rút\s*dọc|co\s*dọc|độ\s*co\s*dọc)"
            r"\s*[:=]?\s*(\d+(?:\.\d+)?)",
            query,
            re.IGNORECASE
        )

        if warp_match:

            parsed_warp = float(
                warp_match.group(1)
            )

            if 0.0 <= parsed_warp <= 15.0:
                warp_shrinkage = parsed_warp

        # ------------------------------------------------------------
        # CO NGANG
        # ------------------------------------------------------------
        weft_match = re.search(
            r"\b(?:co\s*rút\s*ngang|co\s*ngang|độ\s*co\s*ngang)"
            r"\s*[:=]?\s*(\d+(?:\.\d+)?)",
            query,
            re.IGNORECASE
        )

        if weft_match:

            parsed_weft = float(
                weft_match.group(1)
            )

            if 0.0 <= parsed_weft <= 15.0:
                weft_shrinkage = parsed_weft

        # ------------------------------------------------------------
        # SIZE
        # Hỗ trợ:
        # size 30
        # size M
        # cỡ 32
        # cỡ XL
        # ------------------------------------------------------------
        size_match = re.search(
            r"\b(?:size|cỡ)\s*[:=]?\s*([a-zA-Z0-9]+)\b",
            query,
            re.IGNORECASE
        )

        if size_match:

            target_size = str(
                size_match.group(1)
            ).strip().upper()

    # ================================================================
    # 5. 🔒 MASTER COMMIT VÀO SESSION STATE
    # ================================================================
    st.session_state[
        "current_active_width"
    ] = float(fabric_width)

    st.session_state[
        "current_active_size"
    ] = str(target_size)

    st.session_state[
        "current_warp_shrinkage"
    ] = float(warp_shrinkage)

    st.session_state[
        "current_weft_shrinkage"
    ] = float(weft_shrinkage)

    # ================================================================
    # 6. 🔒 MASTER COMMIT VÀO BOM DATA
    # ================================================================
    bom_source[
        "fabric_width_inch"
    ] = float(fabric_width)

    bom_source[
        "usable_width_inch"
    ] = float(fabric_width)

    bom_source[
        "warp_shrinkage_percent"
    ] = float(warp_shrinkage)

    bom_source[
        "weft_shrinkage_percent"
    ] = float(weft_shrinkage)

    bom_source[
        "calculated_on_size"
    ] = str(target_size)

    bom_source[
        "detected_base_size"
    ] = str(target_size)

    # ================================================================
    # 7. KHÓA LỚP AI KHÔNG ĐƯỢC GHI ĐÈ THÔNG SỐ CHAT
    # ================================================================
    ai_expert_decision[
        "virtual_pieces_layer"
    ] = virtual_pieces_layer

    ai_expert_decision[
        "detected_base_size"
    ] = str(target_size)

    ai_expert_decision[
        "fabric_width"
    ] = float(fabric_width)

    ai_expert_decision[
        "fabric_width_inch"
    ] = float(fabric_width)

    ai_expert_decision[
        "warp_shrinkage_percent"
    ] = float(warp_shrinkage)

    ai_expert_decision[
        "weft_shrinkage_percent"
    ] = float(weft_shrinkage)

    bom_source[
        "ai_expert_decision"
    ] = ai_expert_decision

    # ================================================================
    # 8. GHI BOM DATA CUỐI CÙNG
    # ================================================================
    st.session_state[
        "bom_data"
    ] = bom_source

    # ================================================================
    # 9. DEBUG MASTER
    # ================================================================
    print(
        "\n"
        "============================================================\n"
        "[MASTER PARAMETER SYNC]\n"
        f"CHAT        = {user_query_text}\n"
        f"SIZE        = {target_size}\n"
        f"WIDTH       = {fabric_width:.2f}\"\n"
        f"WARP SHRINK  = {warp_shrinkage:.2f}%\n"
        f"WEFT SHRINK  = {weft_shrinkage:.2f}%\n"
        "============================================================\n"
    )

    return bom_source, user_query_text
# =====================================================================
# 🧠 CUTTING INSTRUCTION ENGINE (VERSION V26.0)
# 🔒 STRICT PIECE QUANTITY RECOVERY
# =====================================================================
import re
import streamlit as st


def extract_cutting_instructions_from_pdf(
    component_name,
    raw_pdf_text,
    current_inferred_pcs=1.0
):
    """
    Quét Callout kỹ thuật trong PDF để xác định số lượng mảnh cắt thực tế.

    NGUYÊN TẮC:
    1. Giữ nguyên số lượng AI đã xác định nếu PDF không có Callout rõ ràng.
    2. Nếu PDF có CUT / QTY / PCS / X2 / PAIR thì dùng dữ liệu PDF.
    3. Không tự động ép mọi PANEL / FRONT / BACK / SLEEVE thành 2.
    4. Không nhân layer_multiplier lần nữa ở các tầng sau.
    5. final_validated_pcs là số lượng mảnh vật lý cuối cùng.
    """

    # =================================================================
    # 1. SAFE BASE QUANTITY
    # =================================================================
    try:
        base_pcs = int(float(current_inferred_pcs))
    except (TypeError, ValueError):
        base_pcs = 1

    base_pcs = max(base_pcs, 1)

    final_validated_pcs = base_pcs

    layer_multiplier = 1
    is_paired = False

    calc_log = (
        f"AI Base Quantity: giữ nguyên {base_pcs} mảnh "
        f"nếu PDF không có Callout xác nhận khác."
    )

    # =================================================================
    # 2. KHÔNG CÓ PDF TEXT → GIỮ NGUYÊN AI QUANTITY
    # =================================================================
    if not raw_pdf_text:
        return {
            "layer_multiplier": 1,
            "final_validated_pcs": final_validated_pcs,
            "is_paired": False,
            "calc_log": (
                "CAD Fallback: Không có PDF text. "
                f"Giữ nguyên AI quantity = {final_validated_pcs}."
            )
        }

    # =================================================================
    # 3. CHUẨN HÓA TEXT
    # =================================================================
    text_clean = " ".join(
        str(raw_pdf_text).lower().split()
    )

    comp_clean = " ".join(
        str(component_name).lower().split()
    ).strip()

    if not comp_clean:
        return {
            "layer_multiplier": 1,
            "final_validated_pcs": final_validated_pcs,
            "is_paired": False,
            "calc_log": (
                "Không có Component Name. "
                f"Giữ AI quantity = {final_validated_pcs}."
            )
        }

    # =================================================================
    # 4. TÌM COMPONENT TRONG PDF
    # =================================================================
    match_positions = [
        m.start()
        for m in re.finditer(
            re.escape(comp_clean),
            text_clean,
            re.IGNORECASE
        )
    ]

    if not match_positions:
        return {
            "layer_multiplier": 1,
            "final_validated_pcs": final_validated_pcs,
            "is_paired": False,
            "calc_log": (
                f"Không tìm thấy Callout cho [{component_name}]. "
                f"Giữ AI quantity = {final_validated_pcs}."
            )
        }

    # =================================================================
    # 5. QUÉT TỪNG VÙNG QUANH COMPONENT
    # =================================================================
    best_detected_qty = None
    best_priority = 999
    best_window = ""

    for match_index in match_positions:

        window_start = max(
            0,
            match_index - 100
        )

        window_end = min(
            len(text_clean),
            match_index + 180
        )

        scan_window = text_clean[
            window_start:window_end
        ]

        # =============================================================
        # PRIORITY 1
        # Explicit CUT / QTY / PCS
        # =============================================================
        explicit_patterns = [
            r"\bcut\s*(?:qty\s*)?(?:x\s*)?[:=]?\s*(\d+)\b",
            r"\bqty\s*[:=]?\s*(\d+)\b",
            r"\bquantity\s*[:=]?\s*(\d+)\b",
            r"\b(\d+)\s*pcs\b",
            r"\b(\d+)\s*pieces\b",
            r"\bcut\s+(\d+)\b",
            r"\bcắt\s+(\d+)\b",
            r"\bsố\s*lượng\s*[:=]?\s*(\d+)\b",
            r"\bsl\s*[:=]?\s*(\d+)\b",
        ]

        detected_qty = None
        detected_priority = 999

        for pattern in explicit_patterns:

            m = re.search(
                pattern,
                scan_window,
                re.IGNORECASE
            )

            if m:

                try:
                    qty = int(m.group(1))
                except (TypeError, ValueError):
                    continue

                if qty >= 1:

                    detected_qty = qty
                    detected_priority = 1
                    break

        # =============================================================
        # PRIORITY 2
        # X2 / X3 / X4
        # =============================================================
        if detected_qty is None:

            x_match = re.search(
                r"(?:\bx\s*|\*\s*)([2-9])\b",
                scan_window,
                re.IGNORECASE
            )

            if x_match:

                try:
                    qty = int(x_match.group(1))
                except (TypeError, ValueError):
                    qty = None

                if qty and qty >= 2:

                    detected_qty = qty
                    detected_priority = 2

        # =============================================================
        # PRIORITY 3
        # PAIR / 1 PAIR / DOUBLE
        # =============================================================
        pair_detected = bool(
            re.search(
                r"\b(pair|1\s*pair|double|mirror|"
                r"đối\s*xứng|cặp|đôi)\b",
                scan_window,
                re.IGNORECASE
            )
        )

        if pair_detected:

            is_paired = True

            if detected_qty is None:

                # Chỉ suy ra 2 khi PDF thực sự nói PAIR.
                detected_qty = 2
                detected_priority = 3

        # =============================================================
        # CHỌN KẾT QUẢ CÓ ĐỘ ƯU TIÊN CAO NHẤT
        # =============================================================
        if (
            detected_qty is not None
            and detected_priority < best_priority
        ):
            best_detected_qty = detected_qty
            best_priority = detected_priority
            best_window = scan_window

    # =================================================================
    # 6. COMMIT QUANTITY TỪ PDF
    # =================================================================
    if best_detected_qty is not None:

        final_validated_pcs = max(
            int(best_detected_qty),
            1
        )

        if best_priority == 1:

            calc_log = (
                f"PDF Callout xác nhận CUT/QTY = "
                f"{final_validated_pcs} mảnh. "
                f"AI Base = {base_pcs}."
            )

        elif best_priority == 2:

            calc_log = (
                f"PDF Callout xác nhận X{final_validated_pcs}. "
                f"AI Base = {base_pcs}."
            )

        else:

            calc_log = (
                "PDF Callout xác nhận PAIR/MIRROR. "
                f"Khôi phục = {final_validated_pcs} mảnh."
            )

    else:

        # =============================================================
        # KHÔNG CÓ CALLOUT → TUYỆT ĐỐI KHÔNG ÉP VỀ 2
        # =============================================================
        final_validated_pcs = base_pcs

        calc_log = (
            f"Không có Callout quantity rõ ràng. "
            f"Giữ nguyên AI Base = {base_pcs}."
        )

    # =================================================================
    # 7. 🔒 MASTER GUARD
    # =================================================================
    final_validated_pcs = max(
        int(final_validated_pcs),
        1
    )

    # Không cho layer multiplier nhân lại quantity.
    layer_multiplier = 1

    # =================================================================
    # 8. RETURN MASTER QUANTITY
    # =================================================================
    return {
        "layer_multiplier": 1,

        # SỐ LƯỢNG MẢNH VẬT LÝ CUỐI CÙNG
        "final_validated_pcs": int(
            final_validated_pcs
        ),

        "is_paired": bool(is_paired),

        "calc_log": calc_log,

        # Debug để các tầng sau biết nguồn quantity
        "base_pcs": int(base_pcs),

        "quantity_source": (
            "PDF_CALLOUT"
            if best_detected_qty is not None
            else "AI_BASE"
        ),

        "detected_pdf_quantity": (
            int(best_detected_qty)
            if best_detected_qty is not None
            else None
        )
    }



# =====================================================================
# 🟩 ĐOẠN 2 - DATA CLEANING & MASTER PARAMETER READ
# VERSION V27.0 - MASTER LOCK / NO OVERRIDE
# =====================================================================

import re
import pandas as pd


# =====================================================================
# 1. LẤY BOM ROWS
# =====================================================================

rows = ctx.get("bom_rows", [])

if not rows:
    rows = st.session_state.get("processed_display_rows", [])


# =====================================================================
# 2. KIỂM TRA DỮ LIỆU BOM
# =====================================================================

_has_rows = (
    isinstance(rows, list) and len(rows) > 0
) or (
    isinstance(rows, pd.DataFrame) and not rows.empty
)


if _has_rows:

    # =================================================================
    # 3. TẠO DATAFRAME SẠCH
    # =================================================================

    df_bom = (
        pd.DataFrame(rows)
        if isinstance(rows, list)
        else rows.copy()
    )

    # Loại bỏ column trùng tên
    df_bom = df_bom.loc[
        :,
        ~df_bom.columns.duplicated()
    ].copy()


    # =================================================================
    # 4. NHẬN DIỆN PRODUCT TYPE THỰC TẾ
    #    CHỐNG AI BẮT NHẦM JEAN_LONG
    # =================================================================

    ai_decision = st.session_state.get(
        "bom_data",
        {}
    ).get(
        "ai_expert_decision",
        {}
    )

    if not isinstance(ai_decision, dict):
        ai_decision = {}


    style_code_upper = str(
        ai_decision.get(
            "style_code",
            ""
        )
    ).upper().strip()


    material_spec_upper = str(
        ai_decision.get(
            "material_spec",
            ""
        )
    ).upper().strip()


    p_type_friendly = str(
        ai_decision.get(
            "product_type_friendly",
            "JEAN_LONG"
        )
    ).upper().strip()


    # =================================================================
    # 5. CHUỖI TỔNG HỢP PHỤC VỤ QUÉT PRODUCT TYPE
    # =================================================================

    combined_search_text = (
        f"{style_code_upper} | "
        f"{material_spec_upper} | "
        f"{p_type_friendly}"
    )


    # =================================================================
    # 6. LẤY PRODUCT TYPE TỪ CONTEXT
    # =================================================================

    prod = str(
        ctx.get(
            "detected_product_type",
            ctx.get(
                "product_segmented",
                "JEAN_LONG"
            )
        )
    ).upper().strip()


    # =================================================================
    # 7. ƯU TIÊN NHẬN DIỆN PRODUCT TYPE
    # =================================================================

    if "DRESS" in combined_search_text:

        prod = "DRESS"

    elif "SKIRT" in combined_search_text:

        prod = "SKIRT"

    elif "SHORT" in combined_search_text:

        prod = "SHORT"

    elif (
        "JACKET" in combined_search_text
        or "COAT" in combined_search_text
    ):

        prod = "JACKET"

    elif "SHIRT" in combined_search_text:

        prod = "SHIRT"


    # =================================================================
    # 8. ĐỒNG BỘ PRODUCT TYPE
    # =================================================================

    ctx["detected_product_type"] = prod
    ctx["product_segmented"] = prod


    # =================================================================
    # 9. FABRIC PATTERN
    # =================================================================

    fabric_pattern_raw = str(
        ctx.get(
            "fabric_pattern",
            "SOLID"
        )
    ).upper()


    # =================================================================
    # 10. XÁC ĐỊNH CỘT MATERIAL
    # =================================================================

    m_col = next(
        (
            c
            for c in [
                "Material Class",
                "material_class"
            ]
            if c in df_bom.columns
        ),
        "material_class"
    )


    # Nếu column chưa tồn tại thì tạo
    if m_col not in df_bom.columns:
        df_bom[m_col] = ""


    # =================================================================
    # 11. XÁC ĐỊNH CỘT PIECE COUNT
    # =================================================================

    pcs_col = next(
        (
            c
            for c in [
                "Số lượng rập",
                "piece_count"
            ]
            if c in df_bom.columns
        ),
        "piece_count"
    )


    # Nếu column chưa tồn tại thì tạo
    if pcs_col not in df_bom.columns:
        df_bom[pcs_col] = 1


    # =================================================================
    # 12. XÁC ĐỊNH BOUNDING BOX LENGTH
    # =================================================================

    orig_l_col = next(
        (
            c
            for c in [
                "bounding_box_length",
                "Dài (L-inch)"
            ]
            if c in df_bom.columns
        ),
        "bounding_box_length"
    )


    if orig_l_col not in df_bom.columns:
        df_bom[orig_l_col] = 0.0


    # =================================================================
    # 13. XÁC ĐỊNH BOUNDING BOX WIDTH
    # =================================================================

    orig_w_col = next(
        (
            c
            for c in [
                "bounding_box_width",
                "Rộng (W-inch)"
            ]
            if c in df_bom.columns
        ),
        "bounding_box_width"
    )


    if orig_w_col not in df_bom.columns:
        df_bom[orig_w_col] = 0.0


    # =================================================================
    # 14. CLEAN GEOMETRY DIMENSIONS
    # =================================================================

    df_bom[orig_l_col] = pd.to_numeric(
        df_bom[orig_l_col],
        errors="coerce"
    ).fillna(0.0)


    df_bom[orig_w_col] = pd.to_numeric(
        df_bom[orig_w_col],
        errors="coerce"
    ).fillna(0.0)


    # =================================================================
    # 15. GIỮ LẠI GROSS CONSUMPTION GỐC
    # =================================================================

    target_orig_gross_col = next(
        (
            c
            for c in [
                "Gross Consumption",
                "gross_consumption",
                "allocated_gross"
            ]
            if c in df_bom.columns
        ),
        None
    )


    if target_orig_gross_col:

        df_bom["original_raw_gross"] = pd.to_numeric(
            df_bom[target_orig_gross_col],
            errors="coerce"
        ).fillna(0.0)

    else:

        df_bom["original_raw_gross"] = 0.0


    # =================================================================
    # 16. USER EDIT BUFFER
    # =================================================================

    if "user_edited_materials" not in st.session_state:
        st.session_state["user_edited_materials"] = {}


    if "user_edited_pieces" not in st.session_state:
        st.session_state["user_edited_pieces"] = {}


    # =================================================================
    # 17. APPLY USER MATERIAL EDIT
    # =================================================================

    for idx, row in df_bom.iterrows():

        if idx in st.session_state["user_edited_materials"]:

            df_bom.at[
                idx,
                m_col
            ] = st.session_state[
                "user_edited_materials"
            ][idx]


    # =================================================================
    # 18. CLEAN PIECE COUNT
    # =================================================================

    def clean_precise_piece_count(row):

        comp_name = str(
            row.get(
                "component_name",
                row.get(
                    "Component Name",
                    ""
                )
            )
        ).upper().strip()


        pcs_raw_str = str(
            row.get(
                pcs_col,
                "1"
            )
        )


        pcs_extracted = re.search(
            r"(\d+(?:\.\d+)?)",
            pcs_raw_str
        )


        if pcs_extracted:

            try:
                pcs_val = float(
                    pcs_extracted.group(1)
                )

            except (TypeError, ValueError):

                pcs_val = 1.0

        else:

            pcs_val = 1.0


        if pcs_val <= 0:
            pcs_val = 1.0


        return pcs_val


    # =================================================================
    # 19. CREATE NUMERIC PIECE COUNT
    # =================================================================

    clean_piece_values = []

    for idx, row in df_bom.iterrows():

        if idx in st.session_state[
            "user_edited_pieces"
        ]:

            try:

                edited_value = float(
                    st.session_state[
                        "user_edited_pieces"
                    ][idx]
                )

                if edited_value > 0:
                    clean_piece_values.append(
                        edited_value
                    )
                else:
                    clean_piece_values.append(1.0)

            except (TypeError, ValueError):

                clean_piece_values.append(
                    clean_precise_piece_count(row)
                )

        else:

            clean_piece_values.append(
                clean_precise_piece_count(row)
            )


    df_bom["pcs_numeric"] = clean_piece_values


    df_bom[pcs_col] = pd.to_numeric(
        df_bom["pcs_numeric"],
        errors="coerce"
    ).fillna(1.0)


    # =====================================================================
    # 🚨 20. MASTER PARAMETER READ
    #
    # ĐOẠN 2 KHÔNG TỰ TÍNH PARAMETER.
    # ĐOẠN 1 LÀ MASTER DUY NHẤT.
    # =====================================================================

    master_params = st.session_state.get(
        "ie_master_params",
        {}
    )


    if not isinstance(master_params, dict):
        master_params = {}


    # =================================================================
    # 21. FABRIC WIDTH
    # =================================================================

    fabric_width = master_params.get(
        "fabric_width",
        st.session_state.get(
            "current_active_width",
            st.session_state.get(
                "fabric_width_inch",
                ctx.get(
                    "fabric_width_inch",
                    58.0
                )
            )
        )
    )


    try:

        fabric_width = float(
            fabric_width
        )

    except (TypeError, ValueError):

        fabric_width = 58.0


    if fabric_width <= 0:
        fabric_width = 58.0


    # =================================================================
    # 22. WARP / VERTICAL SHRINKAGE
    # =================================================================

    warp_shrink = master_params.get(
        "warp_shrinkage",
        st.session_state.get(
            "current_warp_shrinkage",
            st.session_state.get(
                "warp_shrinkage",
                st.session_state.get(
                    "shrinkage_vertical",
                    ctx.get(
                        "warp_shrinkage",
                        0.0
                    )
                )
            )
        )
    )


    try:

        warp_shrink = float(
            warp_shrink
        )

    except (TypeError, ValueError):

        warp_shrink = 0.0


    # =================================================================
    # 23. WEFT / HORIZONTAL SHRINKAGE
    # =================================================================

    weft_shrink = master_params.get(
        "weft_shrinkage",
        st.session_state.get(
            "current_weft_shrinkage",
            st.session_state.get(
                "weft_shrinkage",
                st.session_state.get(
                    "shrinkage_horizontal",
                    ctx.get(
                        "weft_shrinkage",
                        0.0
                    )
                )
            )
        )
    )


    try:

        weft_shrink = float(
            weft_shrink
        )

    except (TypeError, ValueError):

        weft_shrink = 0.0


    # =================================================================
    # 24. ĐỒNG BỘ MASTER → SESSION ALIAS
    #
    # Các tầng cũ vẫn có thể đọc các tên biến cũ.
    # =================================================================

    st.session_state[
        "fabric_width_inch"
    ] = fabric_width


    st.session_state[
        "current_active_width"
    ] = fabric_width


    st.session_state[
        "warp_shrinkage"
    ] = warp_shrink


    st.session_state[
        "weft_shrinkage"
    ] = weft_shrink


    st.session_state[
        "current_warp_shrinkage"
    ] = warp_shrink


    st.session_state[
        "current_weft_shrinkage"
    ] = weft_shrink


    st.session_state[
        "shrinkage_vertical"
    ] = warp_shrink


    st.session_state[
        "shrinkage_horizontal"
    ] = weft_shrink


    # =================================================================
    # 25. MASTER → BOM CONTEXT
    # =================================================================

    ctx[
        "fabric_width_inch"
    ] = fabric_width


    ctx[
        "warp_shrinkage_percent"
    ] = warp_shrink


    ctx[
        "weft_shrinkage_percent"
    ] = weft_shrink


    ctx[
        "warp_shrinkage"
    ] = warp_shrink


    ctx[
        "weft_shrinkage"
    ] = weft_shrink


    # =================================================================
    # 26. LƯU MASTER PARAMETER SNAPSHOT VÀO BOM
    # =================================================================

    ctx[
        "ie_master_params"
    ] = {
        "size": master_params.get(
            "size",
            st.session_state.get(
                "current_active_size",
                ctx.get(
                    "calculated_on_size",
                    "32"
                )
            )
        ),
        "fabric_width": fabric_width,
        "fusing_width": master_params.get(
            "fusing_width",
            st.session_state.get(
                "fusing_width_inch",
                ctx.get(
                    "fusing_width_inch",
                    59.0
                )
            )
        ),
        "lining_width": master_params.get(
            "lining_width",
            st.session_state.get(
                "lining_width_inch",
                ctx.get(
                    "lining_width_inch",
                    57.0
                )
            )
        ),
        "warp_shrinkage": warp_shrink,
        "weft_shrinkage": weft_shrink,
    }


    # =================================================================
    # 27. LƯU BOM DATA CUỐI CÙNG
    # =================================================================

    st.session_state[
        "bom_data"
    ] = ctx


    # =================================================================
    # 28. MASTER SYNC STATUS
    # =================================================================

    st.session_state[
        "ie_parameter_sync_complete"
    ] = True

    # =====================================================================
    # 🟩 ĐOẠN 3.1 - AI PRODUCT CLASSIFIER
    # VERSION V28.1 - COMPACT / CAD SAFE
    # =====================================================================

    import re

    COMPANY_DENSITY_PRIOR = {
        "SHIRT": 0.82,
        "JEAN_LONG": 0.795,
        "SHORT": 0.83,
        "JACKET": 0.68,
        "VEST": 0.82,
        "TOPS_KNIT": 0.78,
        "SKIRT": 0.82,
        "DRESS_FLARE": 0.72,
    }

    if not isinstance(st.session_state.get("bom_data"), dict):
        st.session_state["bom_data"] = {}

    ctx = st.session_state["bom_data"]

    if not isinstance(ctx.get("ai_expert_decision"), dict):
        ctx["ai_expert_decision"] = {}

    ai_decision = ctx["ai_expert_decision"]

    comp_col = next(
        (
            c for c in [
                "Component Name",
                "component_name",
                "Component_Name"
            ]
            if c in df_bom.columns
        ),
        None
    )

    all_components_text = (
        " ".join(
            df_bom[comp_col]
            .fillna("")
            .astype(str)
            .str.upper()
        )
        if comp_col
        else ""
    )

    style_code = str(
        ai_decision.get("style_code", "")
    ).upper().strip()

    material_spec = str(
        ai_decision.get("material_spec", "")
    ).upper().strip()

    prod_master = str(
        ctx.get(
            "detected_product_type",
            ctx.get(
                "product_segmented",
                prod if "prod" in locals() else ""
            )
        )
    ).upper().strip()

    combined_text = (
        f"{style_code} "
        f"{material_spec} "
        f"{prod_master} "
        f"{all_components_text}"
    ).upper()

    product_category = None

    if re.search(
        r"\b(SHIRT|SƠ MI|SO MI|BLOUSE)\b",
        combined_text
    ):
        product_category = "SHIRT"

    elif re.search(
        r"\b(SKIRT|CHÂN VÁY|CHAN VAY)\b",
        combined_text
    ):
        product_category = "SKIRT"

    elif re.search(
        r"\b(DRESS|ĐẦM|DAM|FLARE DRESS|SHIFT DRESS|MAXI DRESS)\b",
        combined_text
    ):
        product_category = "DRESS_FLARE"

    elif re.search(
        r"\b(JACKET|COAT|BLAZER|ÁO KHOÁC|AO KHOAC|BOMBER|PARKA)\b",
        combined_text
    ):
        product_category = "JACKET"

    elif re.search(
        r"\b(SHORT|SHORTS|QUẦN SHORT|QUAN SHORT)\b",
        combined_text
    ):
        product_category = "SHORT"

    elif re.search(
        r"\b(TROUSER|TROUSERS|PANTS|PANT|JEAN|JEANS|QUẦN|QUAN|WAISTBAND|FLY|CẠP|CAP|POCKET FACING)\b",
        all_components_text
    ):
        product_category = "JEAN_LONG"

    elif re.search(
        r"\b(VEST|WAISTCOAT)\b",
        combined_text
    ):
        product_category = "VEST"

    elif re.search(
        r"\b(POLO|T-SHIRT|TSHIRT|TEE|KNIT TOP|HOODIE|SWEATSHIRT)\b",
        combined_text
    ):
        product_category = "TOPS_KNIT"

    if product_category is None:

        master_map = {
            "SHIRT": "SHIRT",
            "SKIRT": "SKIRT",
            "SHORT": "SHORT",
            "JACKET": "JACKET",
            "COAT": "JACKET",
            "VEST": "VEST",
            "TOPS_KNIT": "TOPS_KNIT",
            "POLO": "TOPS_KNIT",
            "T-SHIRT": "TOPS_KNIT",
            "TSHIRT": "TOPS_KNIT",
            "DRESS": "DRESS_FLARE",
            "DRESS_FLARE": "DRESS_FLARE",
            "JEAN": "JEAN_LONG",
            "JEANS": "JEAN_LONG",
            "PANT": "JEAN_LONG",
            "PANTS": "JEAN_LONG",
            "TROUSER": "JEAN_LONG",
            "JEAN_LONG": "JEAN_LONG",
        }

        product_category = master_map.get(
            prod_master,
            "JEAN_LONG"
        )

    density_prior = COMPANY_DENSITY_PRIOR.get(
        product_category,
        COMPANY_DENSITY_PRIOR["JEAN_LONG"]
    )

    friendly_map = {
        "SHIRT": "SHIRT (Áo sơ mi)",
        "SKIRT": "SKIRT (Chân váy)",
        "SHORT": "SHORT (Quần short)",
        "JACKET": "JACKET (Áo khoác Jacket)",
        "VEST": "VEST (Áo Vest/Blazer)",
        "TOPS_KNIT": "TOPS_KNIT (Áo thun/Polo)",
        "DRESS_FLARE": "DRESS_FLARE (Đầm)",
        "JEAN_LONG": "JEAN_LONG (Quần dài Jeans/Pants)",
    }

    ai_product_type = friendly_map.get(
        product_category,
        friendly_map["JEAN_LONG"]
    )

    virtual_pieces = ai_decision.get(
        "virtual_pieces_layer",
        {}
    )

    if not isinstance(virtual_pieces, dict):
        virtual_pieces = {}

    ai_decision["product_category"] = product_category
    ai_decision["product_type_friendly"] = ai_product_type
    ai_decision["estimated_density_prior"] = density_prior
    ai_decision["virtual_pieces_layer"] = virtual_pieces

    ctx["detected_product_type"] = product_category
    ctx["product_segmented"] = product_category
    ctx["estimated_density_prior"] = density_prior

    st.session_state["current_product_category"] = product_category
    st.session_state["current_product_type_friendly"] = ai_product_type
    st.session_state["current_estimated_density_prior"] = density_prior

    st.session_state["bom_data"] = ctx
         # =====================================================================
    # 🟩 ĐOẠN 4 - MASTER GEOMETRY & STRICT MATERIAL CLASSIFIER
    # VERSION V28.1 - COMPACT / CAD SAFE
    # =====================================================================

    import pandas as pd
    import numpy as np

    # ---------------------------------------------------------------------
    # 1. MASTER CONTEXT
    # ---------------------------------------------------------------------

    if not isinstance(st.session_state.get("bom_data"), dict):
        st.session_state["bom_data"] = {}

    ctx = st.session_state["bom_data"]

    if not isinstance(ctx.get("ai_expert_decision"), dict):
        ctx["ai_expert_decision"] = {}

    comp_col_check = next(
        (
            c for c in [
                "Component Name",
                "component_name",
                "Component_Name"
            ]
            if c in df_bom.columns
        ),
        "component_name"
    )

    m_col_check = next(
        (
            c for c in [
                "Material Class",
                "material_class"
            ]
            if c in df_bom.columns
        ),
        "material_class"
    )

    user_edited_materials = st.session_state.get(
        "user_edited_materials",
        {}
    )

    user_edited_pieces = st.session_state.get(
        "user_edited_pieces",
        {}
    )

    virtual_pieces_layer = {}


    # ---------------------------------------------------------------------
    # 2. SAFE NUMBER CONVERTER
    # ---------------------------------------------------------------------

    def safe_float(value, default=0.0):
        try:
            value = pd.to_numeric(value, errors="coerce")
            return default if pd.isna(value) else float(value)
        except Exception:
            return default


    # ---------------------------------------------------------------------
    # 3. MATERIAL CLASSIFIER + GEOMETRY
    # ---------------------------------------------------------------------

    for idx, row in df_bom.iterrows():

        idx_str = str(idx).strip()

        comp_name_raw = str(
            row.get(
                comp_col_check,
                row.get("component_name", "")
            )
        ).strip()

        comp_name_upper = comp_name_raw.upper()

        mat_str = str(
            row.get(m_col_check, "")
        ).upper().strip()


        # ---------------------------------------------------------------
        # MATERIAL CLASS
        # ---------------------------------------------------------------

        if idx in user_edited_materials:
            p_class = str(
                user_edited_materials[idx]
            ).upper().strip()

        elif idx_str in user_edited_materials:
            p_class = str(
                user_edited_materials[idx_str]
            ).upper().strip()

        elif any(
            k in comp_name_upper or k in mat_str
            for k in [
                "THREAD",
                "CHỈ",
                "BUTTON",
                "NÚT",
                "ZIP",
                "ZIPPER",
                "ACCESSORY"
            ]
        ):
            p_class = "ACCESSORY"

        elif any(
            k in comp_name_upper or k in mat_str
            for k in [
                "FUSING",
                "MEC",
                "MẾCH",
                "KEO",
                "INTERLINING",
                "DỰNG",
                "WAISTBAND FUSING"
            ]
        ):
            p_class = "FUSING"

        elif any(
            k in comp_name_upper or k in mat_str
            for k in [
                "RIB",
                "BO GÂN",
                "BO CO",
                "BO TAY",
                "BO LAI",
                "BO LƯNG",
                "BO LUNG",
                "BO TĂM"
            ]
        ):
            p_class = "RIB"

        elif any(
            k in comp_name_upper or k in mat_str
            for k in [
                "LINING",
                "LÓT",
                "POCKET BAG",
                "POCKETING",
                "VẢI LÓT",
                "POCKET FACING",
                "POCKETING FABRIC"
            ]
        ):
            p_class = "LINING"

        elif any(
            k in comp_name_upper or k in mat_str
            for k in [
                "CONTRAST",
                "PHỐI",
                "VẢI PHỐI",
                "MATCHING"
            ]
        ):
            p_class = "CONTRAST"

        elif any(
            k in comp_name_upper or k in mat_str
            for k in [
                "PADDING",
                "GÒN",
                "WADDING",
                "BÔNG LOT"
            ]
        ):
            p_class = "PADDING"

        else:
            p_class = "FABRIC"


        # ---------------------------------------------------------------
        # GEOMETRY INPUT
        # ---------------------------------------------------------------

        l_orig = safe_float(
            row.get(
                "bounding_box_length",
                row.get(
                    "Dài (L-inch)",
                    row.get(
                        "Chiều dài rập (inch)",
                        0.0
                    )
                )
            )
        )

        w_orig = safe_float(
            row.get(
                "bounding_box_width",
                row.get(
                    "Rộng (W-inch)",
                    row.get(
                        "Chiều rộng rập (inch)",
                        0.0
                    )
                )
            )
        )

        net_area_real = safe_float(
            row.get(
                "polygon_net_area",
                0.0
            )
        )

        if l_orig <= 0 or w_orig <= 0:
            continue


        # ---------------------------------------------------------------
        # NORMALIZE L / W
        # ---------------------------------------------------------------

        if w_orig > l_orig:
            l_orig, w_orig = w_orig, l_orig

        bbox_area = l_orig * w_orig


        # ---------------------------------------------------------------
        # NET AREA CONTROL
        # ---------------------------------------------------------------

        if net_area_real <= 0:
            net_area_real = bbox_area * 0.74

        elif net_area_real > bbox_area:
            net_area_real = bbox_area * 0.85


        # ---------------------------------------------------------------
        # CAD OBB EFFICIENCY CONTROL
        # ---------------------------------------------------------------

        if net_area_real > 0 and bbox_area > 0:

            current_factor = (
                net_area_real / bbox_area
            )

            aspect_ratio = (
                l_orig / max(w_orig, 0.001)
            )

            log_aspect = np.log1p(
                aspect_ratio
            )

            target_obb_eff = max(
                0.6400,
                min(
                    0.9200,
                    0.88
                    - (0.05 * log_aspect)
                    + (0.15 * current_factor)
                )
            )

            if (
                current_factor < target_obb_eff
                and target_obb_eff > 0
            ):

                optimized_area = (
                    net_area_real /
                    target_obb_eff
                )

                w_new = np.sqrt(
                    optimized_area /
                    max(aspect_ratio, 0.001)
                )

                l_new = (
                    w_new *
                    aspect_ratio
                )

                # Chỉ cập nhật nếu geometry hợp lệ
                if l_new > 0 and w_new > 0:
                    l_orig = l_new
                    w_orig = w_new


        # ---------------------------------------------------------------
        # PIECE COUNT
        # ---------------------------------------------------------------

        raw_pcs = safe_float(
            row.get(
                "pcs_numeric",
                row.get(
                    "Số lượng rập",
                    1.0
                )
            ),
            1.0
        )

        raw_pcs = max(
            raw_pcs,
            1.0
        )

        if idx in user_edited_pieces:
            final_pcs = safe_float(
                user_edited_pieces[idx],
                raw_pcs
            )

        elif idx_str in user_edited_pieces:
            final_pcs = safe_float(
                user_edited_pieces[idx_str],
                raw_pcs
            )

        else:
            final_pcs = raw_pcs

        final_pcs = max(
            final_pcs,
            1.0
        )


        # ---------------------------------------------------------------
        # SAVE VIRTUAL CAD PIECE
        # ---------------------------------------------------------------

        virtual_pieces_layer[idx_str] = {
            "material_class": p_class,
            "production_l": round(l_orig, 2),
            "production_w": round(w_orig, 2),
            "production_net_area": round(
                net_area_real,
                2
            ),
            "polygon_net_area": round(
                net_area_real,
                2
            ),
            "active_user_pieces": int(
                final_pcs
            ),
            "component_name": comp_name_raw
        }


    # ---------------------------------------------------------------------
    # 4. WRITE BACK TO MASTER DATAFRAME
    # ---------------------------------------------------------------------

    for idx_str, vp in virtual_pieces_layer.items():

        target_idx = None

        if idx_str in df_bom.index:
            target_idx = idx_str

        else:
            try:
                idx_int = int(idx_str)

                if idx_int in df_bom.index:
                    target_idx = idx_int

            except Exception:
                pass

        if target_idx is None:
            continue

        df_bom.at[
            target_idx,
            "Chiều dài rập (inch)"
        ] = vp["production_l"]

        df_bom.at[
            target_idx,
            "Chiều rộng rập (inch)"
        ] = vp["production_w"]

        df_bom.at[
            target_idx,
            "polygon_net_area"
        ] = vp["production_net_area"]

        df_bom.at[
            target_idx,
            "Material Class"
        ] = vp["material_class"]


    # ---------------------------------------------------------------------
    # 5. MASTER OUTPUT
    # ---------------------------------------------------------------------

    ctx["ai_expert_decision"][
        "virtual_pieces_layer"
    ] = virtual_pieces_layer

    st.session_state["bom_data"] = ctx
        # =====================================================================
    # 🟩 ĐOẠN 5.1 - PIECE NORMALIZE PIPELINE
    # VERSION V28.6 - STRICT USER SYNC / CAD SAFE
    # =====================================================================

    import pandas as pd

    # ---------------------------------------------------------------------
    # 1. MASTER WIDTH PARAMETERS
    # ---------------------------------------------------------------------

    current_fabric_width = float(
        st.session_state.get(
            "current_active_width",
            st.session_state.get(
                "fabric_width_inch",
                58.0
            )
        )
    )

    fusing_width = float(
        st.session_state.get(
            "fusing_width_inch",
            st.session_state.get(
                "fusing_width",
                59.0
            )
        )
    )

    lining_width = float(
        st.session_state.get(
            "lining_width_inch",
            st.session_state.get(
                "lining_width",
                57.0
            )
        )
    )


    # ---------------------------------------------------------------------
    # 2. COLUMN RESOLVER
    # ---------------------------------------------------------------------

    l_col = next(
        (
            c for c in [
                "Chiều dài rập (inch)",
                "bounding_box_length"
            ]
            if c in df_bom.columns
        ),
        None
    )

    w_col = next(
        (
            c for c in [
                "Chiều rộng rập (inch)",
                "bounding_box_width"
            ]
            if c in df_bom.columns
        ),
        None
    )

    pcs_col = next(
        (
            c for c in [
                "pcs_numeric",
                "Số lượng rập"
            ]
            if c in df_bom.columns
        ),
        "Số lượng rập"
    )


    # ---------------------------------------------------------------------
    # 3. USER EDIT BUFFER
    # ---------------------------------------------------------------------

    user_edited_materials = st.session_state.get(
        "user_edited_materials",
        {}
    )

    user_edited_pieces = st.session_state.get(
        "user_edited_pieces",
        {}
    )

    raw_unpaired_pieces = []
    list_lengths = []
    list_widths = []


    # ---------------------------------------------------------------------
    # 4. SAFE NUMBER CONVERTER
    # ---------------------------------------------------------------------

    def safe_piece_float(value, default=0.0):

        try:
            value = pd.to_numeric(
                value,
                errors="coerce"
            )

            if pd.isna(value):
                return default

            return float(value)

        except Exception:
            return default


    # ---------------------------------------------------------------------
    # 5. NORMALIZE EVERY CAD PIECE
    # ---------------------------------------------------------------------

    for idx, r in df_bom.iterrows():

        idx_str = str(idx).strip()

        # ---------------------------------------------------------------
        # VIRTUAL PIECE
        # ---------------------------------------------------------------

        v_piece = virtual_pieces_layer.get(
            idx_str,
            {}
        )

        if not isinstance(v_piece, dict):
            v_piece = {}

        virtual_pieces_layer[idx_str] = v_piece


        # ---------------------------------------------------------------
        # GEOMETRY
        # ---------------------------------------------------------------

        p_len = safe_piece_float(
            v_piece.get(
                "production_l",
                r.get(
                    l_col,
                    0.0
                ) if l_col else 0.0
            )
        )

        p_wid = safe_piece_float(
            v_piece.get(
                "production_w",
                r.get(
                    w_col,
                    0.0
                ) if w_col else 0.0
            )
        )

        net_area = safe_piece_float(
            v_piece.get(
                "polygon_net_area",
                r.get(
                    "polygon_net_area",
                    0.0
                )
            )
        )


        # ---------------------------------------------------------------
        # MATERIAL CLASS - USER OVERRIDE FIRST
        # ---------------------------------------------------------------

        if idx in user_edited_materials:

            p_class_check = str(
                user_edited_materials[idx]
            ).upper().strip()

        elif idx_str in user_edited_materials:

            p_class_check = str(
                user_edited_materials[idx_str]
            ).upper().strip()

        else:

            p_class_check = str(
                v_piece.get(
                    "material_class",
                    r.get(
                        "Material Class",
                        "FABRIC"
                    )
                )
            ).upper().strip()

        if not p_class_check:
            p_class_check = "FABRIC"

        v_piece["material_class"] = p_class_check


        # ---------------------------------------------------------------
        # NET AREA FALLBACK
        # ---------------------------------------------------------------

        if (
            net_area <= 0.0
            and p_len > 0.0
            and p_wid > 0.0
        ):
            net_area = p_len * p_wid


        # ---------------------------------------------------------------
        # PIECE COUNT
        # ---------------------------------------------------------------

        raw_pcs = safe_piece_float(
            v_piece.get(
                "active_user_pieces",
                r.get(
                    pcs_col,
                    1.0
                ) if pcs_col else 1.0
            ),
            1.0
        )

        raw_pcs = max(
            raw_pcs,
            1.0
        )


        # ---------------------------------------------------------------
        # USER PIECE OVERRIDE
        # ---------------------------------------------------------------

        if idx in user_edited_pieces:

            pcs = safe_piece_float(
                user_edited_pieces[idx],
                raw_pcs
            )

        elif idx_str in user_edited_pieces:

            pcs = safe_piece_float(
                user_edited_pieces[idx_str],
                raw_pcs
            )

        else:

            pcs = raw_pcs

        pcs = max(
            pcs,
            1.0
        )


        # ---------------------------------------------------------------
        # WRITE PIECE COUNT BACK
        # ---------------------------------------------------------------

        df_bom.at[
            idx,
            pcs_col
        ] = int(round(pcs))

        v_piece["active_user_pieces"] = int(
            round(pcs)
        )


        # ---------------------------------------------------------------
        # WRITE GEOMETRY
        # ---------------------------------------------------------------

        list_lengths.append(
            round(p_len, 2)
            if p_len > 0
            else 0.0
        )

        list_widths.append(
            round(p_wid, 2)
            if p_wid > 0
            else 0.0
        )

        df_bom.at[
            idx,
            "polygon_net_area"
        ] = round(
            net_area,
            2
        )

        v_piece["polygon_net_area"] = round(
            net_area,
            2
        )


        # ---------------------------------------------------------------
        # CREATE UNPAIRED PIECES FOR NESTING
        # ---------------------------------------------------------------

        if (
            p_class_check in [
                "FABRIC",
                "FUSING",
                "LINING",
                "RIB",
                "CONTRAST",
                "PADDING"
            ]
            and p_len > 0.0
            and p_wid > 0.0
        ):

            loop_pcs = max(
                1,
                int(round(pcs))
            )

            for _ in range(loop_pcs):

                raw_unpaired_pieces.append({
                    "idx": idx_str,
                    "l": p_len,
                    "w": p_wid,
                    "area": net_area,
                    "material_class": p_class_check,
                    "priority": 3
                })


    # ---------------------------------------------------------------------
    # 6. SORT PIECES FOR NESTING
    # ---------------------------------------------------------------------

    raw_unpaired_pieces.sort(
        key=lambda x: (
            x.get(
                "priority",
                3
            ),
            -x.get(
                "area",
                0.0
            )
        )
    )


    # ---------------------------------------------------------------------
    # 7. WRITE NORMALIZED GEOMETRY BACK
    # ---------------------------------------------------------------------

    df_bom[
        "Chiều dài rập (inch)"
    ] = list_lengths

    df_bom[
        "Chiều rộng rập (inch)"
    ] = list_widths


    # ---------------------------------------------------------------------
    # 8. MASTER SYNC
    # ---------------------------------------------------------------------

    ctx[
        "ai_expert_decision"
    ][
        "virtual_pieces_layer"
    ] = virtual_pieces_layer

    ctx[
        "raw_unpaired_pieces"
    ] = raw_unpaired_pieces

    ctx[
        "fabric_width_inch"
    ] = current_fabric_width

    ctx[
        "fusing_width_inch"
    ] = fusing_width

    ctx[
        "lining_width_inch"
    ] = lining_width

    st.session_state[
        "bom_data"
    ] = ctx


   
   # =====================================================================
# 🟩 ĐOẠN 5.2 - PHẦN A
# VERSION V29.0 - MASTER PRODUCT TYPE + MARKER EFFICIENCY ROUTER
# =====================================================================

import pandas as pd
import streamlit as st
import re

# =====================================================================
# 1. MASTER BOM CONTEXT
# =====================================================================

if "bom_data" not in st.session_state or not isinstance(
    st.session_state["bom_data"], dict
):
    st.session_state["bom_data"] = {}

ctx = st.session_state["bom_data"]

if "ai_expert_decision" not in ctx or not isinstance(
    ctx.get("ai_expert_decision"),
    dict
):
    ctx["ai_expert_decision"] = {}

ai_decision = ctx["ai_expert_decision"]


# =====================================================================
# 2. MASTER CONFIG MATRIX
# =====================================================================

CONFIG_MATRIX = {

    # ---------------------------------------------------------------
    # OVERALL
    # ---------------------------------------------------------------
    "OVERALL":  [0.71, "OVERALLS (Quần yếm/Quần bảo hộ)"],
    "COVERALL": [0.71, "OVERALLS (Quần yếm/Quần bảo hộ)"],
    "BIB":      [0.71, "OVERALLS (Quần yếm/Quần bảo hộ)"],
    "JUMPSUIT": [0.70, "OVERALLS (Quần yếm/Quần bảo hộ)"],
    "DUNGAREE": [0.71, "OVERALLS (Quần yếm/Quần bảo hộ)"],

    # ---------------------------------------------------------------
    # BOTTOMS
    # ---------------------------------------------------------------
    "DRESS":     [0.75, "DRESS (Đầm xòe/suông)"],
    "SKIRT":     [0.66, "SKIRT (Chân váy)"],

    "SHORT":     [0.68, "SHORT (Quần short)"],

    "JEAN":      [0.80, "JEAN (Vải Denim/Jean)"],
    "JEAN_LONG": [0.80, "JEAN_LONG (Quần Jeans dài chuẩn)"],

    "KHAKI":     [0.60, "KHAKI (Quần Khaki)"],
    "TROUSER":   [0.71, "TROUSER (Quần tây công sở)"],
    "PANT":      [0.72, "PANT (Quần dài dáng suông)"],

    # ---------------------------------------------------------------
    # OUTERWEAR
    # ---------------------------------------------------------------
    "JACKET": [0.78, "JACKET (Áo khoác gió/Jeans)"],
    "COAT":   [0.60, "COAT (Áo măng tô/Áo choàng)"],
    "BLAZER": [0.65, "BLAZER (Áo Vest mỏng/Blazer)"],
    "SUIT":   [0.65, "SUIT (Bộ Comple/Suit)"],

    # ---------------------------------------------------------------
    # TOPS
    # ---------------------------------------------------------------
    "SHIRT":  [0.60, "SHIRT (Áo sơ mi vải dệt)"],
    "BLOUSE": [0.78, "BLOUSE (Áo kiểu/Blouse)"],
    "POLO":   [0.76, "POLO (Áo thun cổ bẻ)"],
    "TEE":    [0.76, "TEE/TSHIRT (Áo thun cổ tròn)"],
    "TSHIRT": [0.76, "TEE/TSHIRT (Áo thun cổ tròn)"],
    "TANK":   [0.74, "TANK (Áo ba lỗ/Sát nách)"],
}


# =====================================================================
# 3. ALIAS NORMALIZATION
# =====================================================================

PRODUCT_TYPE_ALIAS = {

    # -------------------------------
    # SHORT
    # -------------------------------
    "SHORTS": "SHORT",
    "BERMUDA": "SHORT",
    "BERMUDAS": "SHORT",
    "WALK SHORT": "SHORT",
    "WALKING SHORT": "SHORT",
    "DENIM SHORT": "SHORT",
    "DENIM SHORTS": "SHORT",
    "JEAN SHORT": "SHORT",
    "JEAN SHORTS": "SHORT",
    "JEANS SHORT": "SHORT",
    "JEANS SHORTS": "SHORT",

    # -------------------------------
    # LONG JEANS
    # -------------------------------
    "LONG JEAN": "JEAN_LONG",
    "LONG JEANS": "JEAN_LONG",
    "JEAN LONG": "JEAN_LONG",
    "JEANS LONG": "JEAN_LONG",
    "FULL LENGTH JEAN": "JEAN_LONG",
    "FULL LENGTH JEANS": "JEAN_LONG",

    # -------------------------------
    # PANTS
    # -------------------------------
    "PANTS": "PANT",
    "LONG PANT": "PANT",
    "LONG PANTS": "PANT",
    "TROUSERS": "TROUSER",

    # -------------------------------
    # T-SHIRT
    # -------------------------------
    "T-SHIRT": "TSHIRT",
    "T SHIRT": "TSHIRT",
    "TEE SHIRT": "TSHIRT",

    # -------------------------------
    # JACKET
    # -------------------------------
    "JACKETS": "JACKET",

    # -------------------------------
    # SHIRT
    # -------------------------------
    "SHIRTS": "SHIRT",

    # -------------------------------
    # POLO
    # -------------------------------
    "POLOS": "POLO",
}


# =====================================================================
# 4. LẤY PRODUCT TYPE THÔ TỪ AI
# =====================================================================

raw_type_candidates = [

    ai_decision.get("ai_product_type_raw"),

    ai_decision.get("detected_product_type"),

    ctx.get("detected_product_type"),

    ctx.get("product_type"),

    ctx.get("product_type_raw"),

    ctx.get("ie_detected_type"),

    st.session_state.get("ai_product_type_raw"),

    st.session_state.get("detected_product_type"),
]


inherited_raw_type = ""

for candidate in raw_type_candidates:

    if candidate is None:
        continue

    candidate_clean = str(candidate).upper().strip()

    if candidate_clean and candidate_clean not in [
        "NONE",
        "NULL",
        "UNKNOWN",
        "N/A",
        "NAN",
    ]:
        inherited_raw_type = candidate_clean
        break


# =====================================================================
# 5. LẤY THÊM THÔNG TIN TỪ STYLE / QUERY / TECHPACK
#    DÙNG ĐỂ CHỐNG AI NHẬN SAI SHORT THÀNH JEAN_LONG
# =====================================================================

style_code_text = str(
    ctx.get("style_code", "")
).upper().strip()

user_query_text = str(
    st.session_state.get(
        "last_submitted_query",
        ""
    )
).upper().strip()

detected_text_blob = " ".join(
    [
        inherited_raw_type,
        style_code_text,
        user_query_text,
        str(ctx.get("product_type_friendly", "")).upper(),
        str(ctx.get("ai_product_type_friendly", "")).upper(),
    ]
).strip()


# =====================================================================
# 6. SHORT OVERRIDE - ƯU TIÊN TUYỆT ĐỐI
#
# Nếu Tech Pack / Chat / AI có dấu hiệu SHORT thì phải là SHORT.
# JEAN SHORT KHÔNG ĐƯỢC PHÉP CHẠY THÀNH JEAN_LONG.
# =====================================================================

SHORT_PATTERNS = [
    r"\bSHORT\b",
    r"\bSHORTS\b",
    r"\bBERMUDA\b",
    r"\bWALK\s*SHORT\b",
    r"\bWALKING\s*SHORT\b",
    r"\bDENIM\s*SHORT\b",
    r"\bDENIM\s*SHORTS\b",
    r"\bJEAN\s*SHORT\b",
    r"\bJEAN\s*SHORTS\b",
    r"\bJEANS\s*SHORT\b",
    r"\bJEANS\s*SHORTS\b",
]


is_short_detected = any(
    re.search(pattern, detected_text_blob, re.IGNORECASE)
    for pattern in SHORT_PATTERNS
)


# =====================================================================
# 7. NORMALIZE PRODUCT TYPE
# =====================================================================

normalized_type = PRODUCT_TYPE_ALIAS.get(
    inherited_raw_type,
    inherited_raw_type
)


# =====================================================================
# 8. MASTER PRODUCT TYPE DECISION
# =====================================================================

if is_short_detected:

    # 🔒 SHORT LUÔN ƯU TIÊN HƠN JEAN / PANT / TROUSER
    ie_detected_type = "SHORT"

elif normalized_type in CONFIG_MATRIX:

    ie_detected_type = normalized_type

else:

    # ---------------------------------------------------------------
    # KHÔNG CÒN ÉP FALLBACK VỀ JEAN_LONG
    #
    # JEAN_LONG chỉ được dùng khi AI thực sự nhận diện là JEAN_LONG.
    # Nếu AI trả dữ liệu không hợp lệ -> PANT là fallback trung tính.
    # ---------------------------------------------------------------
    ie_detected_type = "PANT"


# =====================================================================
# 9. EXTRA SAFETY:
#    JEAN SHORT / DENIM SHORT TUYỆT ĐỐI KHÔNG ĐƯỢC CHẠY JEAN_LONG
# =====================================================================

if any(
    x in detected_text_blob
    for x in [
        "JEAN SHORT",
        "JEAN SHORTS",
        "DENIM SHORT",
        "DENIM SHORTS",
        "JEANS SHORT",
        "JEANS SHORTS",
        "BERMUDA",
        "WALK SHORT",
    ]
):

    ie_detected_type = "SHORT"


# =====================================================================
# 10. LẤY HIỆU SUẤT CƠ SỞ
# =====================================================================

dynamic_marker_efficiency = float(
    CONFIG_MATRIX[ie_detected_type][0]
)

product_type_friendly = CONFIG_MATRIX[
    ie_detected_type
][1]


# =====================================================================
# 11. NAP / ONE-WAY ROUTER
# =====================================================================

is_nap_mode = bool(
    st.session_state.get(
        "is_nap_fabric",
        False
    )
)

is_one_way_mode = bool(
    st.session_state.get(
        "is_one_way_fabric",
        False
    )
)


# ONE-WAY ưu tiên cao hơn NAP
if is_one_way_mode:

    dynamic_marker_efficiency -= 0.05

elif is_nap_mode:

    dynamic_marker_efficiency -= 0.03


# =====================================================================
# 12. GIỚI HẠN AN TOÀN HIỆU SUẤT
# =====================================================================

dynamic_marker_efficiency = max(
    0.52,
    min(
        round(dynamic_marker_efficiency, 4),
        0.95
    )
)


# =====================================================================
# 13. MASTER COMMIT
# =====================================================================

ctx["ie_detected_type"] = ie_detected_type

ctx["ie_product_type_friendly"] = product_type_friendly

ctx["detected_product_type"] = ie_detected_type

ctx["product_type"] = ie_detected_type


ai_decision["ai_product_type_raw"] = inherited_raw_type

ai_decision["detected_product_type"] = ie_detected_type

ai_decision["product_type_friendly"] = product_type_friendly

ai_decision["marker_efficiency"] = dynamic_marker_efficiency


# =====================================================================
# 14. SESSION MASTER SYNC
# =====================================================================

st.session_state[
    "active_marker_efficiency_value"
] = float(
    dynamic_marker_efficiency
)

st.session_state[
    "ie_detected_type"
] = ie_detected_type

st.session_state[
    "ie_product_type_friendly"
] = product_type_friendly


# =====================================================================
# 15. DEBUG LOG - KIỂM TRA CHÍNH XÁC AI ĐÃ NHẬN DIỆN GÌ
# =====================================================================

print(
    "\n"
    "============================================================\n"
    "[PRODUCT TYPE ROUTER V29]\n"
    f"AI RAW TYPE       = {inherited_raw_type}\n"
    f"NORMALIZED TYPE   = {normalized_type}\n"
    f"SHORT DETECTED    = {is_short_detected}\n"
    f"FINAL PRODUCT     = {ie_detected_type}\n"
    f"EFFICIENCY        = {dynamic_marker_efficiency:.4f}\n"
    f"NAP MODE          = {is_nap_mode}\n"
    f"ONE-WAY MODE      = {is_one_way_mode}\n"
    "============================================================"
)


# =====================================================================
# 16. LƯU MASTER BOM
# =====================================================================

ctx["ai_expert_decision"] = ai_decision

st.session_state["bom_data"] = ctx
# =====================================================================
# 🟩 ĐOẠN 5.2 - PHẦN B1 + B2
# VERSION V30.0
# MASTER COMMERCIAL CONSUMPTION ENGINE
#
# 🔥 V30.0 FIX HIGH CONSUMPTION
#
# 1. GROUP MARKER BY MATERIAL CLASS
# 2. NO DOUBLE WASTAGE
# 3. CAD POLYGON AREA DEFAULT = SEAM INCLUDED
# 4. SHRINKAGE CONTROLLED
# 5. PIECE QUANTITY PRESERVED
# 6. ROW GROSS DM ALLOCATED FROM MASTER MARKER
# 7. TOTAL ROW DM = MASTER COMMERCIAL DM
# =====================================================================

import pandas as pd
import streamlit as st


# =====================================================================
# 🟢 B1 - INITIALIZATION & MASTER PARAMETER RECOVERY
# =====================================================================

if (
    "bom_data" not in st.session_state
    or not isinstance(
        st.session_state["bom_data"],
        dict
    )
):
    st.session_state["bom_data"] = {}


ctx = st.session_state["bom_data"]


if (
    "ai_expert_decision" not in ctx
    or not isinstance(
        ctx["ai_expert_decision"],
        dict
    )
):
    ctx["ai_expert_decision"] = {}


ai_decision = ctx["ai_expert_decision"]


# =====================================================================
# 1. RECOVERY VIRTUAL PIECES
# =====================================================================

stored_virtual_pieces = ai_decision.get(
    "virtual_pieces_layer",
    {}
)


if not isinstance(
    stored_virtual_pieces,
    dict
):
    stored_virtual_pieces = {}


# =====================================================================
# 2. RECOVERY DATAFRAME
# =====================================================================

if (
    "df_bom" not in locals()
    or df_bom is None
    or (
        isinstance(df_bom, pd.DataFrame)
        and df_bom.empty
    )
):

    rows_backup = ctx.get(
        "bom_rows",
        st.session_state.get(
            "processed_display_rows",
            []
        )
    )

    if rows_backup:
        df_bom = pd.DataFrame(
            rows_backup
        )
    else:
        df_bom = pd.DataFrame()


# =====================================================================
# 3. MASTER FABRIC WIDTH
# =====================================================================

raw_width = st.session_state.get(
    "current_active_width",
    None
)


if raw_width in [
    None,
    "",
    0,
    "0"
]:
    raw_width = ctx.get(
        "fabric_width_inch",
        None
    )


if raw_width in [
    None,
    "",
    0,
    "0"
]:
    raw_width = ctx.get(
        "usable_width_inch",
        None
    )


try:

    parsed_width = float(
        raw_width
    )

except (
    TypeError,
    ValueError
):

    parsed_width = 58.0


if parsed_width <= 0:
    parsed_width = 58.0


# ---------------------------------------------------------------------
# MASTER WIDTH COMMIT
# ---------------------------------------------------------------------

st.session_state[
    "current_active_width"
] = parsed_width


ctx[
    "fabric_width_inch"
] = parsed_width


ctx[
    "usable_width_inch"
] = parsed_width


# =====================================================================
# 4. SHRINKAGE
# =====================================================================

raw_shrink_v = st.session_state.get(
    "current_warp_shrinkage",
    ctx.get(
        "warp_shrinkage_percent",
        0.0
    )
)


raw_shrink_h = st.session_state.get(
    "current_weft_shrinkage",
    ctx.get(
        "weft_shrinkage_percent",
        0.0
    )
)


try:

    shrink_v_percent = float(
        raw_shrink_v
    )

except (
    TypeError,
    ValueError
):

    shrink_v_percent = 0.0


try:

    shrink_h_percent = float(
        raw_shrink_h
    )

except (
    TypeError,
    ValueError
):

    shrink_h_percent = 0.0


shrink_v_percent = max(
    0.0,
    min(
        shrink_v_percent,
        30.0
    )
)


shrink_h_percent = max(
    0.0,
    min(
        shrink_h_percent,
        30.0
    )
)


shrink_v = (
    shrink_v_percent / 100.0
)


shrink_h = (
    shrink_h_percent / 100.0
)


# ---------------------------------------------------------------------
# MASTER SHRINK COMMIT
# ---------------------------------------------------------------------

st.session_state[
    "current_warp_shrinkage"
] = shrink_v_percent


st.session_state[
    "current_weft_shrinkage"
] = shrink_h_percent


ctx[
    "warp_shrinkage_percent"
] = shrink_v_percent


ctx[
    "weft_shrinkage_percent"
] = shrink_h_percent


# =====================================================================
# 5. PRODUCT TYPE
# =====================================================================

product_type = str(
    ctx.get(
        "ie_detected_type",
        ai_decision.get(
            "ai_product_type_raw",
            ctx.get(
                "detected_product_type",
                "JEAN_LONG"
            )
        )
    )
).upper().strip()


# ---------------------------------------------------------------------
# PRODUCT NORMALIZATION
# ---------------------------------------------------------------------

if "JACKET" in product_type:

    product_type = "JACKET"

elif "JEAN" in product_type:

    product_type = "JEAN_LONG"

elif (
    "PANT" in product_type
    or "TROUSER" in product_type
):

    product_type = "PANT"

elif "SHORT" in product_type:

    product_type = "SHORT"

elif "SHIRT" in product_type:

    product_type = "SHIRT"

elif "POLO" in product_type:

    product_type = "POLO"

elif (
    "TEE" in product_type
    or "TSHIRT" in product_type
):

    product_type = "TSHIRT"


# =====================================================================
# 6. BASE EFFICIENCY
# =====================================================================

try:

    base_efficiency = float(
        st.session_state.get(
            "active_marker_efficiency_value",
            ai_decision.get(
                "marker_efficiency",
                0.82
            )
        )
    )

except (
    TypeError,
    ValueError
):

    base_efficiency = 0.82


base_efficiency = max(
    0.60,
    min(
        base_efficiency,
        0.95
    )
)


# =====================================================================
# 7. PRODUCT COMMERCIAL EFFICIENCY
# =====================================================================
#
# Không dùng efficiency quá thấp để tạo DM cao.
#
# Đây là PACKING efficiency thương mại,
# không phải polygon geometry efficiency.
# =====================================================================

PRODUCT_EFFICIENCY_ADJUSTMENT = {

    "JEAN_LONG": 0.86,
    "JEAN":      0.86,

    "PANT":      0.84,
    "TROUSER":   0.84,
    "KHAKI":     0.84,

    "SHORT":     0.83,

    "JACKET":    0.86,
    "COAT":      0.82,
    "BLAZER":    0.80,

    "SHIRT":     0.86,
    "BLOUSE":    0.86,

    "POLO":      0.84,
    "TEE":       0.84,
    "TSHIRT":    0.84,

    "DRESS":     0.78,
    "SKIRT":     0.80,
}


product_efficiency = (
    PRODUCT_EFFICIENCY_ADJUSTMENT.get(
        product_type,
        base_efficiency
    )
)


# =====================================================================
# 8. NAP / ONE WAY
# =====================================================================

is_nap_mode = bool(
    st.session_state.get(
        "is_nap_fabric",
        False
    )
)


is_one_way_mode = bool(
    st.session_state.get(
        "is_one_way_fabric",
        False
    )
)


if is_one_way_mode:

    product_efficiency -= 0.04

elif is_nap_mode:

    product_efficiency -= 0.02


product_efficiency = max(
    0.60,
    min(
        product_efficiency,
        0.95
    )
)


# ---------------------------------------------------------------------
# MASTER EFFICIENCY COMMIT
# ---------------------------------------------------------------------

st.session_state[
    "active_marker_efficiency_value"
] = float(
    round(
        product_efficiency,
        4
    )
)


ai_decision[
    "marker_efficiency"
] = float(
    round(
        product_efficiency,
        4
    )
)


ctx[
    "ie_detected_type"
] = product_type


# =====================================================================
# 🟢 B2 - MASTER COMMERCIAL MARKER ENGINE
# =====================================================================

summary_grouped_gross = {

    "FABRIC": 0.0,
    "FUSING": 0.0,
    "LINING": 0.0,
    "CONTRAST": 0.0,
    "RIB": 0.0,
    "PADDING": 0.0
}


# =====================================================================
# MASTER ROW DATA
# =====================================================================

prepared_rows = []


if (
    isinstance(
        df_bom,
        pd.DataFrame
    )
    and not df_bom.empty
):

    # -----------------------------------------------------------------
    # MASTER COLUMNS
    # -----------------------------------------------------------------

    if "Gross Consumption" not in df_bom.columns:

        df_bom[
            "Gross Consumption"
        ] = 0.0


    if "Số lượng rập" not in df_bom.columns:

        df_bom[
            "Số lượng rập"
        ] = 1


    if (
        "Khổ vải sản xuất (inch)"
        not in df_bom.columns
    ):

        df_bom[
            "Khổ vải sản xuất (inch)"
        ] = parsed_width


    # -----------------------------------------------------------------
    # USER OVERRIDE
    # -----------------------------------------------------------------

    user_pieces_dict = (
        st.session_state.get(
            "user_edited_pieces",
            {}
        )
    )


    user_material_dict = (
        st.session_state.get(
            "user_edited_materials",
            {}
        )
    )


    # =================================================================
    # FIRST PASS
    #
    # Chỉ thu thập geometry.
    # CHƯA tính consumption từng row.
    # =================================================================

    for idx, r in df_bom.iterrows():

        idx_str = str(
            idx
        ).strip()


        # -------------------------------------------------------------
        # VIRTUAL PIECE
        # -------------------------------------------------------------

        v = stored_virtual_pieces.get(
            idx,
            stored_virtual_pieces.get(
                idx_str,
                {}
            )
        )


        if not isinstance(
            v,
            dict
        ):

            v = {}


        # -------------------------------------------------------------
        # COMPONENT
        # -------------------------------------------------------------

        component_name = str(
            r.get(
                "component_name",
                v.get(
                    "component_name",
                    r.get(
                        "Component Name",
                        ""
                    )
                )
            )
        ).upper().strip()


        # -------------------------------------------------------------
        # MATERIAL CLASS
        # -------------------------------------------------------------

        if idx in user_material_dict:

            p_cls = (
                user_material_dict[idx]
            )

        elif idx_str in user_material_dict:

            p_cls = (
                user_material_dict[
                    idx_str
                ]
            )

        else:

            p_cls = v.get(
                "material_class",
                r.get(
                    "Material Class",
                    "FABRIC"
                )
            )


        p_cls = str(
            p_cls
        ).upper().strip()


        if (
            p_cls
            not in summary_grouped_gross
        ):

            p_cls = "FABRIC"


        # =============================================================
        # GEOMETRY
        # =============================================================

        try:

            p_length = float(
                v.get(
                    "production_l",
                    r.get(
                        "Chiều dài rập (inch)",
                        r.get(
                            "bounding_box_length",
                            0.0
                        )
                    )
                ) or 0.0
            )

        except (
            TypeError,
            ValueError
        ):

            p_length = 0.0


        try:

            p_width = float(
                v.get(
                    "production_w",
                    r.get(
                        "Chiều rộng rập (inch)",
                        r.get(
                            "bounding_box_width",
                            0.0
                        )
                    )
                ) or 0.0
            )

        except (
            TypeError,
            ValueError
        ):

            p_width = 0.0


        try:

            pure_unit_area = float(
                v.get(
                    "polygon_net_area",
                    r.get(
                        "polygon_net_area",
                        0.0
                    )
                ) or 0.0
            )

        except (
            TypeError,
            ValueError
        ):

            pure_unit_area = 0.0


        # -------------------------------------------------------------
        # GEOMETRY FALLBACK
        # -------------------------------------------------------------

        if (
            pure_unit_area <= 0
            and p_length > 0
            and p_width > 0
        ):

            pure_unit_area = (
                p_length
                * p_width
            )


        if pure_unit_area <= 0:

            pure_unit_area = 10.0


        # =============================================================
        # PIECE QUANTITY
        # =============================================================

        try:

            if idx in user_pieces_dict:

                pcs = int(
                    user_pieces_dict[idx]
                )

            elif idx_str in user_pieces_dict:

                pcs = int(
                    user_pieces_dict[
                        idx_str
                    ]
                )

            elif (
                "active_user_pieces"
                in v
            ):

                pcs = int(
                    float(
                        v[
                            "active_user_pieces"
                        ]
                    )
                )

            elif pd.notna(
                r.get(
                    "Số lượng rập"
                )
            ):

                pcs = int(
                    float(
                        r.get(
                            "Số lượng rập"
                        )
                    )
                )

            else:

                pcs = int(
                    float(
                        r.get(
                            "piece_count",
                            r.get(
                                "cut_quantity",
                                1
                            )
                        )
                        or 1
                    )
                )

        except (
            TypeError,
            ValueError
        ):

            pcs = 1


        pcs = max(
            pcs,
            1
        )


        # -------------------------------------------------------------
        # MASTER PIECE COMMIT
        # -------------------------------------------------------------

        df_bom.at[
            idx,
            "Số lượng rập"
        ] = pcs


        if (
            idx
            not in stored_virtual_pieces
        ):

            stored_virtual_pieces[
                idx
            ] = {}


        stored_virtual_pieces[
            idx
        ][
            "active_user_pieces"
        ] = pcs


        stored_virtual_pieces[
            idx
        ][
            "material_class"
        ] = p_cls


        # =============================================================
        # 🔥 SEAM LOGIC
        # =============================================================
        #
        # QUAN TRỌNG:
        #
        # polygon_net_area lấy trực tiếp từ CAD.
        #
        # Mặc định coi polygon area đã bao seam allowance.
        #
        # Chỉ cộng seam nếu upstream EXPLICITLY nói:
        # area_includes_seam = False
        # =============================================================

        area_includes_seam_raw = v.get(
            "area_includes_seam",
            r.get(
                "area_includes_seam",
                True
            )
        )


        if (
            area_includes_seam_raw
            is None
        ):

            area_includes_seam = True

        else:

            area_includes_seam = bool(
                area_includes_seam_raw
            )


        if (
            p_cls
            in [
                "FABRIC",
                "CONTRAST"
            ]
            and not area_includes_seam
        ):

            seam_modifier = 1.06

        else:

            seam_modifier = 1.00


        # =============================================================
        # TOTAL NET COMMERCIAL AREA
        # =============================================================

        total_piece_area = (
            pure_unit_area
            * pcs
            * seam_modifier
        )


        # =============================================================
        # MATERIAL WIDTH
        # =============================================================

        if p_cls == "FUSING":

            current_w = float(
                st.session_state.get(
                    "fusing_width",
                    59.0
                )
            )

        elif p_cls == "LINING":

            current_w = float(
                st.session_state.get(
                    "lining_width",
                    57.0
                )
            )

        elif p_cls == "RIB":

            current_w = float(
                st.session_state.get(
                    "rib_width",
                    40.0
                )
            )

        elif p_cls == "PADDING":

            current_w = float(
                st.session_state.get(
                    "padding_width",
                    60.0
                )
            )

        else:

            current_w = parsed_width


        if current_w <= 0:

            current_w = parsed_width


        if current_w <= 0:

            current_w = 58.0


        df_bom.at[
            idx,
            "Khổ vải sản xuất (inch)"
        ] = current_w


        # =============================================================
        # ROW RECORD
        # =============================================================

        prepared_rows.append({

            "idx": idx,

            "component_name":
                component_name,

            "material_class":
                p_cls,

            "pcs":
                pcs,

            "pure_unit_area":
                pure_unit_area,

            "seam_modifier":
                seam_modifier,

            "total_piece_area":
                total_piece_area,

            "current_width":
                current_w,

        })


    # =================================================================
    # SECOND PASS
    #
    # GROUP MATERIAL
    #
    # Mỗi Material Class = một commercial marker.
    # =================================================================

    grouped_material_area = {}


    for item in prepared_rows:

        p_cls = item[
            "material_class"
        ]


        if p_cls not in grouped_material_area:

            grouped_material_area[
                p_cls
            ] = 0.0


        grouped_material_area[
            p_cls
        ] += item[
            "total_piece_area"
        ]


    # =================================================================
    # MATERIAL EFFICIENCY
    # =================================================================

    material_efficiency = {

        "FABRIC":
            product_efficiency,

        "CONTRAST":
            product_efficiency,

        "FUSING":
            0.65,

        "LINING":
            0.65,

        "RIB":
            0.82,

        "PADDING":
            0.85,
    }


    # =================================================================
    # MASTER SHRINKAGE
    #
    # V30:
    #
    # Không cộng wastage riêng.
    #
    # Shrinkage chỉ được apply ONE TIME ở MASTER MARKER.
    # =================================================================

    shrinkage_multiplier = (
        (1.0 + shrink_v)
        *
        (1.0 + shrink_h)
    )


    # =================================================================
    # THIRD PASS
    #
    # MASTER COMMERCIAL CONSUMPTION
    # =================================================================

    master_material_consumption = {}


    for p_cls, total_area in (
        grouped_material_area.items()
    ):

        if total_area <= 0:

            master_material_consumption[
                p_cls
            ] = 0.0

            continue


        # -------------------------------------------------------------
        # WIDTH
        # -------------------------------------------------------------

        if p_cls == "FUSING":

            marker_width = float(
                st.session_state.get(
                    "fusing_width",
                    59.0
                )
            )

        elif p_cls == "LINING":

            marker_width = float(
                st.session_state.get(
                    "lining_width",
                    57.0
                )
            )

        elif p_cls == "RIB":

            marker_width = float(
                st.session_state.get(
                    "rib_width",
                    40.0
                )
            )

        elif p_cls == "PADDING":

            marker_width = float(
                st.session_state.get(
                    "padding_width",
                    60.0
                )
            )

        else:

            marker_width = parsed_width


        if marker_width <= 0:

            marker_width = parsed_width


        if marker_width <= 0:

            marker_width = 58.0


        # -------------------------------------------------------------
        # EFFICIENCY
        # -------------------------------------------------------------

        row_efficiency = float(
            material_efficiency.get(
                p_cls,
                product_efficiency
            )
        )


        row_efficiency = max(
            0.60,
            min(
                row_efficiency,
                0.95
            )
        )


        # -------------------------------------------------------------
        # MASTER COMMERCIAL AREA
        # -------------------------------------------------------------

        gross_area_sq_inches = (
            total_area
            / row_efficiency
        )


        # -------------------------------------------------------------
        # SHRINKAGE ONE TIME
        # -------------------------------------------------------------

        gross_area_post_shrink = (
            gross_area_sq_inches
            * shrinkage_multiplier
        )


        # -------------------------------------------------------------
        # MASTER LINEAR CONSUMPTION
        # -------------------------------------------------------------

        linear_inches_needed = (
            gross_area_post_shrink
            / marker_width
        )


        # -------------------------------------------------------------
        # 🔥 NO EXTRA 5% WASTAGE
        # -------------------------------------------------------------

        master_yards = (
            linear_inches_needed
            / 36.0
        )


        master_yards = max(
            master_yards,
            0.0
        )


        master_material_consumption[
            p_cls
        ] = master_yards


    # =================================================================
    # FOURTH PASS
    #
    # ALLOCATE MASTER CONSUMPTION BACK TO EACH PIECE
    # =================================================================

    for item in prepared_rows:

        idx = item[
            "idx"
        ]

        p_cls = item[
            "material_class"
        ]

        piece_area = item[
            "total_piece_area"
        ]

        group_total_area = (
            grouped_material_area.get(
                p_cls,
                0.0
            )
        )


        master_yards = (
            master_material_consumption.get(
                p_cls,
                0.0
            )
        )


        # -------------------------------------------------------------
        # PROPORTIONAL ALLOCATION
        # -------------------------------------------------------------

        if (
            group_total_area > 0
            and piece_area > 0
        ):

            allocation_ratio = (
                piece_area
                / group_total_area
            )

        else:

            allocation_ratio = 0.0


        piece_gross_yards = (
            master_yards
            * allocation_ratio
        )


        piece_gross_yards = round(
            max(
                piece_gross_yards,
                0.0
            ),
            4
        )


        # -------------------------------------------------------------
        # COMMIT ROW
        # -------------------------------------------------------------

        df_bom.at[
            idx,
            "Gross Consumption"
        ] = piece_gross_yards


        summary_grouped_gross[
            p_cls
        ] += piece_gross_yards


        # -------------------------------------------------------------
        # DEBUG
        # -------------------------------------------------------------

        print(
            f"[DM V30.0] "
            f"idx={idx} | "
            f"{item['component_name']} | "
            f"class={p_cls} | "
            f"pcs={item['pcs']} | "
            f"area={item['pure_unit_area']:.2f} | "
            f"area_total={piece_area:.2f} | "
            f"width={item['current_width']:.2f}\" | "
            f"eff={material_efficiency.get(p_cls, product_efficiency):.4f} | "
            f"alloc={allocation_ratio:.5f} | "
            f"gross={piece_gross_yards:.4f} Yds"
        )


    # =================================================================
    # ROUND SUMMARY
    # =================================================================

    for k in summary_grouped_gross:

        summary_grouped_gross[
            k
        ] = round(
            summary_grouped_gross[k],
            4
        )


    # =================================================================
    # MASTER COMMIT
    # =================================================================

    df_bom[
        "Gross Consumption"
    ] = (
        pd.to_numeric(
            df_bom[
                "Gross Consumption"
            ],
            errors="coerce"
        )
        .fillna(0.0)
        .round(4)
    )


    st.session_state[
        "summary_grouped_gross"
    ] = summary_grouped_gross


    st.session_state[
        "active_calculated_df_bom"
    ] = df_bom.copy()


    ai_decision[
        "virtual_pieces_layer"
    ] = stored_virtual_pieces


    ai_decision[
        "marker_efficiency"
    ] = float(
        product_efficiency
    )


    ai_decision[
        "master_material_consumption"
    ] = master_material_consumption


    ai_decision[
        "grouped_material_area"
    ] = grouped_material_area


    ctx[
        "ai_expert_decision"
    ] = ai_decision


    st.session_state[
        "bom_data"
    ] = ctx


    # =================================================================
    # 🔍 MASTER AUDIT DEBUG
    # =================================================================

    print(
        "\n"
        "=============================================================\n"
        "🔒 DM MASTER V30.0\n"
        "=============================================================\n"
        f"PRODUCT       = {product_type}\n"
        f"SIZE          = "
        f"{st.session_state.get('current_active_size', 'N/A')}\n"
        f"FABRIC WIDTH  = {parsed_width:.2f}\"\n"
        f"SHRINK V      = {shrink_v_percent:.2f}%\n"
        f"SHRINK H      = {shrink_h_percent:.2f}%\n"
        f"SHRINK MULT   = {shrinkage_multiplier:.4f}\n"
        f"BASE EFF      = {base_efficiency:.4f}\n"
        f"PRODUCT EFF   = {product_efficiency:.4f}\n"
        "-------------------------------------------------------------\n"
        f"FABRIC AREA   = "
        f"{grouped_material_area.get('FABRIC', 0.0):.2f} sq.in\n"
        f"FABRIC DM     = "
        f"{master_material_consumption.get('FABRIC', 0.0):.4f} Yds\n"
        "-------------------------------------------------------------\n"
        f"LINING AREA   = "
        f"{grouped_material_area.get('LINING', 0.0):.2f} sq.in\n"
        f"LINING DM     = "
        f"{master_material_consumption.get('LINING', 0.0):.4f} Yds\n"
        "-------------------------------------------------------------\n"
        f"FUSING AREA   = "
        f"{grouped_material_area.get('FUSING', 0.0):.2f} sq.in\n"
        f"FUSING DM     = "
        f"{master_material_consumption.get('FUSING', 0.0):.4f} Yds\n"
        "-------------------------------------------------------------\n"
        f"TOTAL FABRIC  = "
        f"{summary_grouped_gross['FABRIC']:.4f} Yds\n"
        f"TOTAL LINING  = "
        f"{summary_grouped_gross['LINING']:.4f} Yds\n"
        f"TOTAL FUSING  = "
        f"{summary_grouped_gross['FUSING']:.4f} Yds\n"
        "=============================================================\n"
    )
      # =====================================================================
    # 🟩 ĐOẠN 5.2C (VERSION V27.0): AUTOMATED CORES IGNITION
    # =====================================================================

    if "bom_data" not in st.session_state:
        st.session_state["bom_data"] = {}

    ctx = st.session_state["bom_data"]

    # ================================================================
    # 🔥 KIỂM TRA DỮ LIỆU ĐÃ CÓ NHƯNG CHƯA ĐƯỢC TÍNH DM
    # ================================================================
    has_raw_bom = (
        isinstance(ctx.get("bom_rows"), list)
        and len(ctx.get("bom_rows", [])) > 0
    )

    has_virtual_pieces = (
        isinstance(ctx.get("ai_expert_decision"), dict)
        and isinstance(
            ctx.get("ai_expert_decision", {}).get("virtual_pieces_layer"),
            dict
        )
        and len(
            ctx.get("ai_expert_decision", {})
               .get("virtual_pieces_layer", {})
        ) > 0
    )

    has_calculated_bom = (
        "active_calculated_df_bom" in st.session_state
        and isinstance(
            st.session_state.get("active_calculated_df_bom"),
            pd.DataFrame
        )
        and not st.session_state["active_calculated_df_bom"].empty
    )

    # ================================================================
    # 🔥 CHỈ KÍCH HOẠT KHI CÓ BOM THÔ NHƯNG CHƯA CÓ MASTER DM
    # ================================================================
    if (has_raw_bom or has_virtual_pieces) and not has_calculated_bom:

        if not st.session_state.get(
            "pipeline_auto_run_executed",
            False
        ):

            with st.spinner(
                "⚙️ IE Engine đang tự động tính định mức thương mại..."
            ):

                try:

                    # ------------------------------------------------
                    # 🔒 KHÓA TRẠNG THÁI TRƯỚC KHI RERUN
                    # ------------------------------------------------
                    st.session_state[
                        "pipeline_auto_run_executed"
                    ] = True

                    # ------------------------------------------------
                    # 🔄 KHÔI PHỤC df_bom NẾU BỊ MẤT SAU RERUN
                    # ------------------------------------------------
                    if (
                        "df_bom" not in locals()
                        or df_bom is None
                        or (
                            isinstance(df_bom, pd.DataFrame)
                            and df_bom.empty
                        )
                    ):

                        rows_raw = ctx.get(
                            "bom_rows",
                            st.session_state.get(
                                "processed_display_rows",
                                []
                            )
                        )

                        if rows_raw:
                            df_bom = pd.DataFrame(rows_raw)

                    # ------------------------------------------------
                    # 🔥 KIỂM TRA LẠI THÔNG SỐ CHAT TRƯỚC KHI CHẠY
                    # ------------------------------------------------
                    current_width = float(
                        st.session_state.get(
                            "current_active_width",
                            ctx.get(
                                "fabric_width_inch",
                                58.0
                            )
                        )
                    )

                    current_size = str(
                        st.session_state.get(
                            "current_active_size",
                            ctx.get(
                                "calculated_on_size",
                                "32"
                            )
                        )
                    ).strip()

                    current_warp = float(
                        st.session_state.get(
                            "current_warp_shrinkage",
                            ctx.get(
                                "warp_shrinkage_percent",
                                0.0
                            )
                        )
                    )

                    current_weft = float(
                        st.session_state.get(
                            "current_weft_shrinkage",
                            ctx.get(
                                "weft_shrinkage_percent",
                                0.0
                            )
                        )
                    )

                    # ------------------------------------------------
                    # 🔒 COMMIT LẠI THÔNG SỐ MASTER
                    # ------------------------------------------------
                    st.session_state[
                        "current_active_width"
                    ] = current_width

                    st.session_state[
                        "current_active_size"
                    ] = current_size

                    st.session_state[
                        "current_warp_shrinkage"
                    ] = current_warp

                    st.session_state[
                        "current_weft_shrinkage"
                    ] = current_weft

                    ctx["fabric_width_inch"] = current_width
                    ctx["usable_width_inch"] = current_width
                    ctx["calculated_on_size"] = current_size
                    ctx["warp_shrinkage_percent"] = current_warp
                    ctx["weft_shrinkage_percent"] = current_weft

                    st.session_state["bom_data"] = ctx

                    print(
                        f"[5.2C MASTER SYNC] "
                        f"Size={current_size} | "
                        f"Width={current_width}\" | "
                        f"Warp={current_warp}% | "
                        f"Weft={current_weft}%"
                    )

                    # ------------------------------------------------
                    # 🔄 RERUN 1 LẦN ĐỂ 5.2B1 → 5.2B2 TÍNH DM
                    # ------------------------------------------------
                    st.rerun()

                except Exception as e:

                    st.session_state[
                        "pipeline_auto_run_executed"
                    ] = False

                    st.error(
                        f"❌ Lỗi AUTOMATED CORES IGNITION: {e}"
                    )


        # =====================================================================
    # 🟩 ĐOẠN 6: KHỞI TẠO HÀM XUẤT EXCEL NỘI BỘ (LOCAL EXPORT ENGINE - FIXED)
    # =====================================================================
    def local_export_excel_ppj_format(df_sum, df_det, product_type, bom_ctx, density):
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        from openpyxl.utils import get_column_letter

        output_stream = io.BytesIO()
        workbook = Workbook()
        
        f_family = "Segoe UI"
        f_normal = Font(name=f_family, size=10)
        f_bold = Font(name=f_family, size=10, bold=True)
        f_title = Font(name=f_family, size=14, bold=True, color="0E6251")
        f_header = Font(name=f_family, size=10, bold=True, color="FFFFFF")
        
        fill_header = PatternFill(start_color="0E6251", end_color="0E6251", fill_type="solid")
        fill_meta = PatternFill(start_color="F2F4F4", end_color="F2F4F4", fill_type="solid")
        
        bd_side = Side(style='thin', color='BDC3C7')
        bd_thin = Border(left=bd_side, right=bd_side, top=bd_side, bottom=bd_side)
        
        # 🔥 ĐỒNG BỘ AN TOÀN TRÁNH BẪY LỖI NAMEERROR KHI KẾ THỪA BIẾN NGOÀI TRONG HÀM
        f_width_val = float(st.session_state.get("current_active_width", 58.0))
        w_shrink_val = float(st.session_state.get("current_warp_shrinkage", 0.0))
        h_shrink_val = float(st.session_state.get("current_weft_shrinkage", 0.0))
        s_code_val = str(st.session_state.get("current_active_size", bom_ctx.get("detected_base_size", "32"))).upper().strip()

        # --- TAB 1: BOM SUMMARY ---
        w_s1 = workbook.active
        w_s1.title = "BOM Summary"
        w_s1.sheet_view.showGridLines = True
        
        w_s1.cell(row=1, column=1, value="PHÒNG IE / CẮT CAD - HỆ THỐNG QUẢN LÝ PPJ GROUP").font = Font(name=f_family, size=8, italic=True, color="7F8C8D")
        w_s1.cell(row=2, column=1, value="BẢNG ĐỊNH MỨC CHI TIẾT SẢN XUẤT ĐẠI TRÀ").font = f_title
        w_s1.cell(row=4, column=1, value="THÔNG SỐ ĐẦU VÀO SƠ ĐỒ CAD (TECHNICAL PROFILE)").font = Font(name=f_family, size=11, bold=True)
        
        st_code = str(bom_ctx.get("style_code", "N/A")).upper()
        cust_name = str(bom_ctx.get("customer_name", "FACTORY STANDARD")).upper()
        
        m_data = [
            ("Mã hàng / Style Code:", st_code, "Khách hàng / Đối tác:", cust_name),
            ("Size may mẫu (Sample Size):", s_code_val, "Khổ vải hữu dụng (Width):", f'{f_width_val}"'),
            ("Co rút dọc (Warp Shrinkage):", f'{w_shrink_val}%', "Co rút ngang (Weft Shrinkage):", f'{h_shrink_val}%'),
            ("Chủng loại sản phẩm:", str(product_type).upper(), "Hiệu suất sơ đồ (Density):", f'{density * 100:.1f}%')
        ]
        
        for r_idx, row_data in enumerate(m_data, start=5):
            for c_idx, val in enumerate(row_data, start=1):
                cell = w_s1.cell(row=r_idx, column=c_idx, value=val)
                cell.border = bd_thin
                if c_idx == 1 or c_idx == 3:
                    cell.font = f_bold; cell.fill = fill_meta; cell.alignment = Alignment(horizontal="left", vertical="center")
                else:
                    cell.font = f_normal; cell.alignment = Alignment(horizontal="center", vertical="center")
                    
        w_s1.cell(row=10, column=1, value="BẢNG TỔNG HỢP TIÊU HAO VẬT TƯ (BOM SUMMARY)").font = Font(name=f_family, size=11, bold=True)
        sum_hd = ["Phân loại vật tư", "Mã Vật Liệu Gốc", "Định Mức (Gross Consumption)", "Đơn Vị Tính (UOM)"]
        for c_idx, h_text in enumerate(sum_hd, start=1):
            cell = w_s1.cell(row=11, column=c_idx, value=h_text)
            cell.font = f_header; cell.fill = fill_header; cell.alignment = Alignment(horizontal="center", vertical="center"); cell.border = bd_thin
            
        c_row = 12
        for _, r in df_sum.iterrows():
            w_s1.cell(row=c_row, column=1, value=r.get("Phân loại vật tư", "VẬT TƯ"))
            w_s1.cell(row=c_row, column=2, value=r.get("Material Class", "FABRIC"))
            w_s1.cell(row=c_row, column=3, value=float(r.get("Gross Consumption", 0.0)))
            w_s1.cell(row=c_row, column=4, value=r.get("UOM", "YDS"))
            w_s1.cell(row=c_row, column=3).number_format = '#,##0.0000'
            for c_idx in range(1, 5):
                cell = w_s1.cell(row=c_row, column=c_idx)
                cell.font = f_normal; cell.border = bd_thin
                if c_idx == 2 or c_idx == 4: 
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            c_row += 1

        for col_idx, col in enumerate(w_s1.columns, start=1):
            m_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col_idx)
            w_s1.column_dimensions[col_letter].width = max(m_len + 3, 12)

        # --- TAB 2: DETAILED CAD PIECES ---
        w_s2 = workbook.create_sheet(title="Detailed CAD Pieces")
        w_s2.sheet_view.showGridLines = True
        w_s2.cell(row=1, column=1, value=f"CHI TIẾT CẤU TRÚC ĐA GIÁC RẬP GERBER ACCUMULATION - DÒNG: {str(product_type).upper()}").font = Font(name=f_family, size=11, bold=True)
        
        # 🌟 ĐỒNG BỘ CHUẨN TÊN CỘT LƯỚI HIỂN THỊ TRÊN FILE EXCEL ĐỂ KHÔNG BỊ KHUYẾT SỐ LIỆU
        det_hd = [
            "Component Name", "Material Class", "Role/Piece Type", "Khổ vải sản xuất (inch)", 
            "Size tính toán", "Số lượng rập", "Chiều dài rập (inch)", "Chiều rộng rập (inch)", 
            "polygon_net_area", "Gross Consumption"
        ]
        for c_idx, h_text in enumerate(det_hd, start=1):
            cell = w_s2.cell(row=3, column=c_idx, value=h_text)
            cell.font = f_header; cell.fill = fill_header; cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); cell.border = bd_thin

        c_row = 4
        for _, r in df_det.iterrows():
            for c_idx, h_col in enumerate(det_hd, start=1):
                val = r.get(h_col, "")
                cell = w_s2.cell(row=c_row, column=c_idx, value=val)
                cell.font = f_normal; cell.border = bd_thin
                
                if c_idx == 1 or c_idx == 2 or c_idx == 3:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                elif c_idx == 4 or c_idx == 5 or c_idx == 6:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    if isinstance(val, (int, float)):
                        cell.number_format = '#,##0.0000' if h_col == "Gross Consumption" else '#,##0.00'
            c_row += 1

        for col_idx, col in enumerate(w_s2.columns, start=1):
            m_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col_idx)
            w_s2.column_dimensions[col_letter].width = max(m_len + 3, 12)

        workbook.save(output_stream)
        output_stream.seek(0)
        return output_stream.getvalue() # Trả về mảng dữ liệu byte sạch để st.download_button đọc trực tiếp

        # =====================================================================
       # =====================================================================
    # 🟩 ĐOẠN 7.1 (VERSION V27.0): DISPLAY LAYER - SUMMARY ONLY
    # =====================================================================
    import re
    import pandas as pd
    import streamlit as st

    # ---------------------------------------------------------------------
    # 1. MASTER CONTEXT RECOVERY
    # ---------------------------------------------------------------------
    if "bom_data" not in st.session_state or not isinstance(st.session_state["bom_data"], dict):
        st.session_state["bom_data"] = {}

    ctx = st.session_state["bom_data"]

    if "ai_expert_decision" not in ctx or not isinstance(ctx["ai_expert_decision"], dict):
        ctx["ai_expert_decision"] = {}

    ai_decision_final = ctx["ai_expert_decision"]

    # ---------------------------------------------------------------------
    # 2. SUMMARY GROSS RECOVERY
    # ---------------------------------------------------------------------
    grouped_gross = st.session_state.get(
        "summary_grouped_gross",
        {
            "FABRIC": 0.0,
            "FUSING": 0.0,
            "LINING": 0.0,
            "CONTRAST": 0.0,
            "RIB": 0.0,
            "PADDING": 0.0
        }
    )

    if not isinstance(grouped_gross, dict):
        grouped_gross = {
            "FABRIC": 0.0,
            "FUSING": 0.0,
            "LINING": 0.0,
            "CONTRAST": 0.0,
            "RIB": 0.0,
            "PADDING": 0.0
        }

    # ---------------------------------------------------------------------
    # 3. DEBUG MASTER RAM
    # ---------------------------------------------------------------------
    st.markdown("### 🔬 Hệ Thống Kiểm Toán Dữ Liệu RAM")

    d_c1, d_c2, d_c3 = st.columns(3)

    d_c1.write(
        f"**DEBUG FABRIC:** "
        f"`{float(grouped_gross.get('FABRIC', 0.0)):.4f} Yds`"
    )

    d_c2.write(
        f"**DEBUG LINING:** "
        f"`{float(grouped_gross.get('LINING', 0.0)):.4f} Yds`"
    )

    d_c3.write(
        f"**DEBUG FUSING:** "
        f"`{float(grouped_gross.get('FUSING', 0.0)):.4f} Yds`"
    )

    st.divider()

    # ---------------------------------------------------------------------
    # 4. AI AUDIT REPORT
    # ---------------------------------------------------------------------
    st.header(
        "📋 AI AUDIT REPORT "
        "(BÁO CÁO KIỂM TOÁN ĐỊNH MỨC TỰ ĐỘNG)"
    )

    virtual_pieces = ai_decision_final.get(
        "virtual_pieces_layer",
        {}
    )

    if not isinstance(virtual_pieces, dict):
        virtual_pieces = {}

    # ---------------------------------------------------------------------
    # 5. COMPLEXITY
    # ---------------------------------------------------------------------
    try:
        comp_score_val = float(
            ai_decision_final.get("complexity_score", 45.0)
        )
    except (TypeError, ValueError):
        comp_score_val = 45.0

    comp_score_val = max(
        0.0,
        min(comp_score_val, 100.0)
    )

    ui_complexity_tier = (
        "COMPLEX"
        if comp_score_val >= 50
        else "NORMAL"
    )

    ui_complexity_icon = (
        "🔴"
        if comp_score_val >= 75
        else (
            "🟡"
            if comp_score_val >= 45
            else "🟢"
        )
    )

    # ---------------------------------------------------------------------
    # 6. PRODUCT TYPE
    # ---------------------------------------------------------------------
    real_sync_product_type = str(
        ctx.get(
            "ie_product_type_friendly",
            ai_decision_final.get(
                "product_type_friendly",
                "JEAN_LONG (Quần dài Jeans/Pants)"
            )
        )
    ).strip()

    # ---------------------------------------------------------------------
    # 7. MARKER EFFICIENCY
    # ---------------------------------------------------------------------
    try:
        marker_efficiency = float(
            st.session_state.get(
                "active_marker_efficiency_value",
                ai_decision_final.get(
                    "marker_efficiency",
                    0.7400
                )
            )
        )
    except (TypeError, ValueError):
        marker_efficiency = 0.7400

    marker_efficiency = max(
        0.0,
        min(marker_efficiency, 1.0)
    )

    # ---------------------------------------------------------------------
    # 8. FABRIC WIDTH MASTER
    # ---------------------------------------------------------------------
    try:
        chat_width_override = float(
            st.session_state.get(
                "current_active_width",
                st.session_state.get(
                    "fabric_width_inch",
                    ctx.get("fabric_width_inch", 58.0)
                )
            )
        )
    except (TypeError, ValueError):
        chat_width_override = 58.0

    if chat_width_override <= 0:
        chat_width_override = 58.0

    # ---------------------------------------------------------------------
    # 9. MATERIAL WIDTH MASTER
    # ---------------------------------------------------------------------
    def get_width_value(session_keys, default_value):
        for key in session_keys:
            value = st.session_state.get(key)

            try:
                value = float(value)
                if value > 0:
                    return value
            except (TypeError, ValueError):
                pass

        return float(default_value)

    fusing_w_audit = get_width_value(
        ["fusing_width_inch", "fusing_width"],
        59.0
    )

    lining_w_audit = get_width_value(
        ["lining_width_inch", "lining_width"],
        57.0
    )

    rib_w_audit = get_width_value(
        ["rib_width_inch", "rib_width"],
        40.0
    )

    padding_w_audit = get_width_value(
        ["padding_width_inch", "padding_width"],
        60.0
    )

    # ---------------------------------------------------------------------
    # 10. TECHNICAL WIDTH AUDIT
    # ---------------------------------------------------------------------
    st.caption(
        f"🔗 **Bảng tra cứu khổ vải kỹ thuật đang áp dụng:** "
        f"Chính (Chat): `{chat_width_override:.1f}\"` | "
        f"Keo (Fusing): `{fusing_w_audit:.1f}\"` | "
        f"Lót (Lining): `{lining_w_audit:.1f}\"` | "
        f"Bo (Rib): `{rib_w_audit:.1f}\"` | "
        f"Gòn (Padding): `{padding_w_audit:.1f}\"`"
    )

    # ---------------------------------------------------------------------
    # 11. MASTER KPI
    # ---------------------------------------------------------------------
    try:
        confidence_value = float(
            ctx.get("confidence", 0.95)
        )
    except (TypeError, ValueError):
        confidence_value = 0.95

    # Hỗ trợ cả trường hợp AI trả 95 thay vì 0.95
    if confidence_value > 1.0:
        confidence_value = confidence_value / 100.0

    confidence_value = max(
        0.0,
        min(confidence_value, 1.0)
    )

    m1, m2, m3, m4 = st.columns(4)

    m1.metric(
        "🤖 Chủng Loại Nhận Diện (IE)",
        real_sync_product_type
    )

    m2.metric(
        f"{ui_complexity_icon} Mức Độ Phức Tạp",
        f"{ui_complexity_tier} ({comp_score_val:.0f}/100)"
    )

    m3.metric(
        "📐 Mật Độ Sơ Đồ Chỉ Định",
        f"{marker_efficiency * 100:.2f}%"
    )

    m4.metric(
        "🎯 Độ Tin Cậy AI (Confidence)",
        f"{confidence_value * 100:.1f}%"
    )

    # ---------------------------------------------------------------------
    # 12. LẤY DATAFRAME MASTER ĐÃ TÍNH
    # ---------------------------------------------------------------------
    if (
        "active_calculated_df_bom" in st.session_state
        and isinstance(
            st.session_state["active_calculated_df_bom"],
            pd.DataFrame
        )
    ):
        df_bom_display = (
            st.session_state["active_calculated_df_bom"].copy()
        )

    elif (
        "df_bom" in locals()
        and isinstance(df_bom, pd.DataFrame)
    ):
        df_bom_display = df_bom.copy()

    else:
        df_bom_display = pd.DataFrame()

    # ---------------------------------------------------------------------
    # 13. MASTER TOTAL GROSS DM
    # ---------------------------------------------------------------------
    if "Gross Consumption" not in df_bom_display.columns:
        df_bom_display["Gross Consumption"] = 0.0

    df_bom_display["Gross Consumption"] = pd.to_numeric(
        df_bom_display["Gross Consumption"],
        errors="coerce"
    ).fillna(0.0)

    _debug_total_dm = float(
        df_bom_display["Gross Consumption"].sum()
    )

    st.caption(
        f"🔒 **MASTER ENGINE DATA:** "
        f"{len(df_bom_display)} pieces | "
        f"Total Gross DM = **{_debug_total_dm:.4f} Yds**"
    )

    # ---------------------------------------------------------------------
    # 14. SUMMARY GROSS BY MATERIAL CLASS
    # ---------------------------------------------------------------------
    total_fabric = float(
        grouped_gross.get("FABRIC", 0.0)
    )

    total_fusing = float(
        grouped_gross.get("FUSING", 0.0)
    )

    total_lining = float(
        grouped_gross.get("LINING", 0.0)
    )

    total_contrast = float(
        grouped_gross.get("CONTRAST", 0.0)
    )

    total_rib = float(
        grouped_gross.get("RIB", 0.0)
    )

    total_padding = float(
        grouped_gross.get("PADDING", 0.0)
    )

    # ---------------------------------------------------------------------
    # 15. BUILD SUMMARY TABLE
    # ---------------------------------------------------------------------
    summary_data = {
        "Phân loại vật tư": ["VẢI CHÍNH"],
        "Material Class": ["FABRIC"],
        "Gross Consumption": [
            round(total_fabric, 4)
        ],
        "UOM": ["Yds"]
    }

    material_summary_map = [
        ("VẢI PHỐI", "CONTRAST", total_contrast),
        ("MÉC / KEO", "FUSING", total_fusing),
        ("VẢI LÓT", "LINING", total_lining),
        ("BO / RIB", "RIB", total_rib),
        ("GÒN LÓT THÂN", "PADDING", total_padding)
    ]

    for display_name, material_class, value in material_summary_map:

        if value > 0.0:

            summary_data["Phân loại vật tư"].append(
                display_name
            )

            summary_data["Material Class"].append(
                material_class
            )

            summary_data["Gross Consumption"].append(
                round(value, 4)
            )

            summary_data["UOM"].append(
                "Yds"
            )

    # ---------------------------------------------------------------------
    # 16. FINAL SUMMARY UI
    # ---------------------------------------------------------------------
    df_summary = pd.DataFrame(summary_data)

    st.subheader(
        "📊 BẢNG TỔNG HỢP BOM SUMMARY (YARDS)"
    )

    st.dataframe(
        df_summary,
        use_container_width=True,
        hide_index=True
    )
       # =====================================================================
       # =====================================================================
    # 🟩 ĐOẠN 7.2 (VERSION V28.0): BOM EDITOR PIPELINE
    # =====================================================================
    import pandas as pd
    import streamlit as st

    # ---------------------------------------------------------------------
    # 1. RECOVER MASTER DATAFRAME
    # ---------------------------------------------------------------------
    if (
        "active_calculated_df_bom" in st.session_state
        and isinstance(
            st.session_state["active_calculated_df_bom"],
            pd.DataFrame
        )
    ):
        df_bom_display_final = (
            st.session_state["active_calculated_df_bom"].copy()
        )

        # -------------------------------------------------------------
        # 2. SAFE NUMERIC NORMALIZATION
        # -------------------------------------------------------------
        numeric_columns = [
            "Chiều dài rập (inch)",
            "Chiều rộng rập (inch)",
            "polygon_net_area",
            "Gross Consumption",
            "Khổ vải sản xuất (inch)",
            "Số lượng rập"
        ]

        for col in numeric_columns:
            if col in df_bom_display_final.columns:
                df_bom_display_final[col] = pd.to_numeric(
                    df_bom_display_final[col],
                    errors="coerce"
                ).fillna(0.0)

        # -------------------------------------------------------------
        # 3. COMPONENT NAME MASTER
        # -------------------------------------------------------------
        c_name_master = next(
            (
                c
                for c in [
                    "component_name",
                    "Component Name",
                    "Component_Name"
                ]
                if c in df_bom_display_final.columns
            ),
            None
        )

        if c_name_master:
            df_bom_display_final["Component Name"] = (
                df_bom_display_final[c_name_master]
                .astype(str)
                .str.upper()
                .str.strip()
            )
        else:
            df_bom_display_final["Component Name"] = (
                "CHI TIẾT RẬP THÔ"
            )

        # -------------------------------------------------------------
        # 4. SIZE MASTER
        # -------------------------------------------------------------
        current_size = str(
            st.session_state.get(
                "current_active_size",
                st.session_state.get(
                    "target_size",
                    "32"
                )
            )
        ).upper().strip()

        df_bom_display_final["Size tính toán"] = current_size

        # -------------------------------------------------------------
        # 5. MATERIAL CLASS MASTER
        # -------------------------------------------------------------
        if "user_edited_materials" not in st.session_state:
            st.session_state["user_edited_materials"] = {}

        user_edited_materials = st.session_state[
            "user_edited_materials"
        ]

        clean_mats = []

        for idx, row in df_bom_display_final.iterrows():

            idx_str = str(idx).strip()

            if idx in user_edited_materials:
                p_cls = user_edited_materials[idx]

            elif idx_str in user_edited_materials:
                p_cls = user_edited_materials[idx_str]

            else:
                p_cls = row.get(
                    "Material Class",
                    row.get(
                        "material_class",
                        "FABRIC"
                    )
                )

            p_cls = str(
                p_cls
            ).upper().strip()

            if p_cls not in [
                "FABRIC",
                "LINING",
                "FUSING",
                "CONTRAST",
                "RIB",
                "PADDING"
            ]:
                p_cls = "FABRIC"

            clean_mats.append(p_cls)

        df_bom_display_final["Material Class"] = clean_mats

        # -------------------------------------------------------------
        # 6. PHYSICAL MASTER INDEX
        # -------------------------------------------------------------
        df_bom_display_final["Mã Chi Tiết"] = (
            df_bom_display_final.index.astype(str)
        )

        # -------------------------------------------------------------
        # 7. ORDER DISPLAY COLUMNS
        # -------------------------------------------------------------
        ordered_cols = [
            "Mã Chi Tiết",
            "Component Name",
            "Material Class",
            "Chiều dài rập (inch)",
            "Chiều rộng rập (inch)",
            "Khổ vải sản xuất (inch)",
            "Size tính toán",
            "Số lượng rập",
            "polygon_net_area",
            "Gross Consumption"
        ]

        display_final_cols = [
            c
            for c in ordered_cols
            if c in df_bom_display_final.columns
        ]

        df_bom_display_final = (
            df_bom_display_final[display_final_cols]
        )

        # -------------------------------------------------------------
        # 8. BOM EDITOR HEADER
        # -------------------------------------------------------------
        st.subheader(
            "📐 ĐỊNH MỨC CHI TIẾT TỪNG RẬP "
            "VÀ ĐIỀU CHỈNH VẬT TƯ (BOM EDITOR)"
        )

        col_t1, col_t2 = st.columns(2)

        with col_t1:
            st.caption(
                "🟢 **Hướng dẫn:** "
                "Click trực tiếp vào ô **FABRIC** của chi tiết "
                "để đổi sang **LINING**, **FUSING**, **CONTRAST**, "
                "**RIB** hoặc **PADDING**."
            )

        # -------------------------------------------------------------
        # 9. EXCEL EXPORT
        # -------------------------------------------------------------
        with col_t2:

            try:

                if "local_export_excel_ppj_format" in locals():

                    excel_file = local_export_excel_ppj_format(
                        df_summary
                        if "df_summary" in locals()
                        else pd.DataFrame(),

                        df_bom_display_final.drop(
                            columns=["Mã Chi Tiết"],
                            errors="ignore"
                        ),

                        str(
                            ctx.get(
                                "product_category",
                                ctx.get(
                                    "ie_detected_type",
                                    "JEAN"
                                )
                            )
                        ),

                        ctx,

                        float(
                            st.session_state.get(
                                "active_marker_efficiency_value",
                                0.74
                            )
                        )
                    )

                    style_name_clean = str(
                        ctx.get(
                            "style_code",
                            ai_decision_final.get(
                                "style_code",
                                "Style"
                            )
                        )
                    ).strip()

                    style_name_clean = (
                        style_name_clean
                        .replace("/", "_")
                        .replace("\\", "_")
                        .replace(" ", "_")
                    )

                    st.download_button(
                        "🟢 DOWNLOAD EXCEL ĐỊNH MỨC THƯƠNG MẠI",
                        data=excel_file,
                        mime=(
                            "application/vnd.openxmlformats-"
                            "officedocument.spreadsheetml.sheet"
                        ),
                        file_name=(
                            f"PPJ_BOM_{style_name_clean}.xlsx"
                        ),
                        use_container_width=True
                    )

            except Exception:
                pass

        # -------------------------------------------------------------
        # 10. USER EDIT MEMORY
        # -------------------------------------------------------------
        if "user_edited_pieces" not in st.session_state:
            st.session_state["user_edited_pieces"] = {}

        if "user_edited_materials" not in st.session_state:
            st.session_state["user_edited_materials"] = {}

        # -------------------------------------------------------------
        # 11. DATA EDITOR
        # -------------------------------------------------------------
        edited_df = st.data_editor(
            df_bom_display_final,

            key="bom_data_editor_matrix_fixed_v28_0",

            use_container_width=True,

            hide_index=True,

            column_config={

                "Mã Chi Tiết": None,

                "Component Name":
                    st.column_config.TextColumn(
                        "📋 Component Name (Tên Chi Tiết)",
                        disabled=True,
                        width="large"
                    ),

                "Material Class":
                    st.column_config.SelectboxColumn(
                        "🧵 Material Class (Click Chọn Sửa)",

                        options=[
                            "FABRIC",
                            "LINING",
                            "FUSING",
                            "CONTRAST",
                            "RIB",
                            "PADDING"
                        ],

                        required=True,
                        disabled=False,
                        width="medium"
                    ),

                "Chiều dài rập (inch)":
                    st.column_config.NumberColumn(
                        "📏 Chiều dài (inch)",
                        format="%.2f",
                        disabled=True
                    ),

                "Chiều rộng rập (inch)":
                    st.column_config.NumberColumn(
                        "📐 Chiều rộng (inch)",
                        format="%.2f",
                        disabled=True
                    ),

                "Khổ vải sản xuất (inch)":
                    st.column_config.NumberColumn(
                        "Khổ vải (inch)",
                        format="%.1f",
                        disabled=True
                    ),

                "Size tính toán":
                    st.column_config.TextColumn(
                        "Size",
                        disabled=True
                    ),

                "Số lượng rập":
                    st.column_config.NumberColumn(
                        "🔢 Số lượng",
                        format="%d",
                        min_value=1,
                        disabled=False
                    ),

                "polygon_net_area":
                    st.column_config.NumberColumn(
                        "polygon_net_area",
                        format="%.2f",
                        disabled=True
                    ),

                "Gross Consumption":
                    st.column_config.NumberColumn(
                        "Gross Consumption (Yds)",
                        format="%.4f",
                        disabled=True
                    )
            }
        )

        # -------------------------------------------------------------
        # 12. EDIT EVENT LISTENER
        # -------------------------------------------------------------
        editor_key = "bom_data_editor_matrix_fixed_v28_0"

        if (
            edited_df is not None
            and editor_key in st.session_state
        ):

            editor_state = st.session_state[
                editor_key
            ]

            changes = editor_state.get(
                "edited_rows",
                {}
            )

            if changes:

                has_updates = False

                for row_key_str, updated_cols in changes.items():

                    # -------------------------------------------------
                    # SAFETY CHECK: edited_rows dùng POSITIONAL INDEX
                    # -------------------------------------------------
                    try:
                        display_row_position = int(
                            row_key_str
                        )
                    except (
                        TypeError,
                        ValueError
                    ):
                        continue

                    if (
                        display_row_position < 0
                        or display_row_position >= len(
                            df_bom_display_final
                        )
                    ):
                        continue

                    # -------------------------------------------------
                    # TRUY NGƯỢC INDEX GỐC
                    # -------------------------------------------------
                    orig_idx = (
                        df_bom_display_final.index[
                            display_row_position
                        ]
                    )

                    master_target_idx = (
                        df_bom_display_final.at[
                            orig_idx,
                            "Mã Chi Tiết"
                        ]
                    )

                    target_key = str(
                        master_target_idx
                    ).strip()

                    # Nếu index gốc là số nguyên thì giữ integer
                    if target_key.isdigit():
                        target_key = int(target_key)

                    # -------------------------------------------------
                    # USER EDIT PIECE COUNT
                    # -------------------------------------------------
                    if "Số lượng rập" in updated_cols:

                        try:
                            new_pcs = int(
                                float(
                                    updated_cols[
                                        "Số lượng rập"
                                    ]
                                )
                            )

                            new_pcs = max(
                                1,
                                new_pcs
                            )

                            st.session_state[
                                "user_edited_pieces"
                            ][target_key] = new_pcs

                            has_updates = True

                        except (
                            TypeError,
                            ValueError
                        ):
                            pass

                    # -------------------------------------------------
                    # USER EDIT MATERIAL CLASS
                    # -------------------------------------------------
                    if "Material Class" in updated_cols:

                        new_material = str(
                            updated_cols[
                                "Material Class"
                            ]
                        ).upper().strip()

                        allowed_materials = [
                            "FABRIC",
                            "LINING",
                            "FUSING",
                            "CONTRAST",
                            "RIB",
                            "PADDING"
                        ]

                        if new_material in allowed_materials:

                            st.session_state[
                                "user_edited_materials"
                            ][target_key] = (
                                new_material
                            )

                            has_updates = True

                # -----------------------------------------------------
                # 13. RE-TRIGGER MASTER ENGINE
                # -----------------------------------------------------
                if has_updates:

                    st.session_state[
                        "pipeline_auto_run_executed"
                    ] = False

                    st.session_state[
                        "active_calculated_df_bom"
                    ] = edited_df.copy()

                    st.rerun()
