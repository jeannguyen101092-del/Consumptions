import streamlit as st
import re
import json
import copy
import streamlit as st
import pandas as pd  # <--- Bắt buộc phải có dòng này
import threading

from typing import List, Optional
from pydantic import BaseModel, Field

class SpecMetaSchema(BaseModel):
    warp_shrink: float = Field(default=3.0, description="Độ co rút dọc (%) trích xuất từ Techpack")
    weft_shrink: float = Field(default=3.0, description="Độ co rút ngang (%) trích xuất từ Techpack")
    gather_ratio: float = Field(default=1.0, description="Tỷ lệ nhún vải (Ví dụ: 1.45 nếu có nhún sườn)")
    has_stripe: bool = Field(default=False, description="True nếu vải có vân sọc, kẻ caro, plaid")
    fabric_group: str = Field(default="WOVEN", description="Nhóm vải chính: DENIM, WOVEN, hoặc KNIT")

class BomRowSchema(BaseModel):
    component_name: str = Field(description="Tên chi tiết rập (Ví dụ: FRONT PANEL, POCKET...)")
    material_class: str = Field(description="Phân loại nguyên liệu: FABRIC, LINING, FUSING, ELASTIC, THREAD")
    piece_count: int = Field(default=1, description="Tổng số lượng chi tiết thực tế khi sản xuất")
    polygon_net_area: Optional[float] = Field(default=0.0, description="Diện tích đa giác từ Gerber/Lectra nếu có")
    polygon_area_mode: Optional[str] = Field(default="PER_PIECE", description="TOTAL hoặc PER_PIECE")
    polygon_unit: Optional[str] = Field(default="IN2", description="CM2 hoặc IN2")
    bounding_box_length: Optional[float] = Field(default=0.0, description="Chiều dài hộp bao khối rập thô")
    bounding_box_width: Optional[float] = Field(default=0.0, description="Chiều rộng hộp bao khối rập thô")
    fabric_width_inch: Optional[float] = Field(default=None, description="Khổ rộng thực tế của vật tư từ BOM")
    # 🎯 BỔ SUNG 2 DÒNG NÀY ĐỂ PYDANTIC CHẤP NHẬN DỮ LIỆU ĐỊNH MỨC DO AI TỰ TÍNH:
    marker_efficiency: Optional[str] = Field(default="82.5%", description="Hiệu suất sơ đồ do AI tự lập luận")
    gross_consumption: Optional[float] = Field(default=0.0, description="Định mức Yards do AI tự tính toán ra số thực")

class AgentOutputSchema(BaseModel):
    spec_meta: SpecMetaSchema
    bom_rows: List[BomRowSchema]

# Danh sách từ khóa tĩnh để tự động loại trừ phụ liệu đếm chiếc khỏi vải cuộn
EXCLUDE_HARDWARE_KEYS = (
    "CHỈ", "THREAD", "ZIPPER", "DÂY KÉO", "BUTTON", "NÚT", "SHANK", "RIVET", 
    "LABEL", "MÁC", "TAG", "EYELETS", "SNAP", "VELCRO", "HOOK", "LOOP", 
    "STOPPER", "TOGGLE", "BUCKLE", "GROMMET", "STICKER", "CARE WHITE", 
    "HEAT STAMP", "HANGTAG", "POLYBAG", "BAO BÌ"
)

def convert_to_sq_inches(area: float, unit: str) -> float:
    """Bộ chuyển đổi đơn vị đo lường vạn năng bám sát hệ thống Gerber/Lectra"""
    u = str(unit).upper().strip()
    if u in ["CM2", "CMSQ", "SQUARE_CM"]:
        return area / 6.4516
    if u in ["MM2", "MMSQ", "SQUARE_MM"]:
        return area / 645.16
    return area


























import streamlit as st

# =====================================================================
# ĐOẠN 6a: KHỞI TẠO BỘ NHỚ STATE & CẤU HÌNH CSS PHẲNG NATIVE CHUẨN ERP
# =====================================================================

# 1. Cấu hình trang rộng toàn màn hình chuẩn hệ thống SaaS/ERP Văn phòng
st.set_page_config(layout="wide", page_title="AI Fabric Consumption Matrix")

# 2. Khởi tạo an toàn cấu trúc trạng thái bộ nhớ hệ thống (Session State)
if "bom_data" not in st.session_state: st.session_state.bom_data = None
if "chat_history" not in st.session_state: st.session_state.chat_history = []
if "pdf_bytes" not in st.session_state: st.session_state.pdf_bytes = None
if "pdf_name" not in st.session_state: st.session_state.pdf_name = ""
if "pdf_text_cache" not in st.session_state: st.session_state.pdf_text_cache = None
if "accumulated_bom_rows" not in st.session_state: st.session_state.accumulated_bom_rows = []

# 3. Tự động phân tách trích xuất văn bản và hình ảnh trang đầu từ tài liệu PDF
if st.session_state.pdf_bytes is not None and (st.session_state.pdf_text_cache is None or st.session_state.get("pdf_page_one_image") is None):
    try:
        import fitz
        doc = fitz.open(stream=st.session_state.pdf_bytes, filetype="pdf")
        
        # Trích xuất văn bản chữ
        if st.session_state.pdf_text_cache is None:
            full_text_extract = ""
            for page_num in range(len(doc)):
                full_text_extract += f"\n--- TRANG THỨ {page_num + 1} ---\n" + doc.load_page(page_num).get_text("text")
            st.session_state.pdf_text_cache = full_text_extract
            
        # Trích xuất hình ảnh trang đầu tiên làm Sketch bản vẽ
        if "pdf_page_one_image" not in st.session_state or st.session_state.pdf_page_one_image is None:
            if len(doc) > 0:
                page = doc.load_page(0)
                pix = page.get_pixmap(dpi=150)
                st.session_state.pdf_page_one_image = pix.tobytes("png")
    except Exception: 
        pass

# 4. Engine đồng bộ dữ liệu KPIs động biến thiên theo thời gian thực trên đỉnh trần
kpi_style_id = "N/A"
total_materials = len(st.session_state.accumulated_bom_rows) if st.session_state.accumulated_bom_rows else 0
main_fabric_cons = "0.000 Yds"
active_size_kpi = "AUTOMATIC"

if st.session_state.get("bom_data") and "bom_rows" in st.session_state.bom_data:
    kpi_style_id = str(st.session_state.bom_data.get("style_code", "R09-500778")).upper()
    active_size_kpi = str(st.session_state.bom_data.get("calculated_on_size", "MEDIAN")).upper()
    if total_materials == 0: total_materials = len(st.session_state.bom_data["bom_rows"])
    for row in st.session_state.bom_data["bom_rows"]:
        if not row: continue
        if "MAIN" in str(row.get("material_class", "")).upper() or "FABRIC" in str(row.get("material_class", "")).upper():
            val_gross = row.get("gross_consumption", 0.0)
            if val_gross > 0.0:
                main_fabric_cons = f"{val_gross:.3f} Yds"
                break


# 5. Bộ cấu hình định dạng CSS phẳng triệt tiêu vĩnh viễn mọi ô trống khổng lồ
st.markdown("""
<style>
    /* Trả màu nền ứng dụng về màu xám trắng dịu mắt chuẩn văn phòng */
    .stApp {
        background-color: #f8fafc !important;
    }
    header[data-testid="stHeader"] {
        background-color: #f8fafc !important;
    }
    
    /* Ép khoảng đệm trần Streamlit về mặc định, triệt tiêu vĩnh viễn khoảng hở */
    .block-container {
        padding-top: 1.5rem !important; 
        margin-top: 0px !important;
        max-width: 100% !important;
    }
    
    /* Ép tất cả các hàng chia cột mặc định phải co khít sát lên trên cùng */
    div[data-testid="stHorizontalBlock"] {
        margin-top: 0px !important;
        padding-top: 0px !important;
    }

    /* Thẻ chỉ số KPIs sắc màu rực rỡ chữ trắng hiển thị rõ nét vĩnh viễn */
    .kpi-box-flat-matrix {
        border-radius: 6px 6px 0 0 !important;
        padding: 10px 12px !important;
        text-align: center !important;
        box-shadow: 0 2px 4px rgba(0,0,0,0.02) !important;
        box-sizing: border-box !important;
    }
    .kpi-num-flat-matrix {
        font-size: 16px !important;
        font-weight: 700 !important;
        color: #ffffff !important; 
        font-family: 'Segoe UI', sans-serif !important;
        line-height: 1.2 !important;
    }
    .kpi-lbl-flat-matrix {
        font-size: 9px !important;
        font-weight: 600 !important;
        color: #ffffff !important;
        opacity: 0.95 !important;
        text-transform: uppercase !important;
        margin-top: 2px !important;
    }

    /* Đóng gói dải màu phân hệ động sắc nét */
    .bg-style-erp { background: linear-gradient(135deg, #334155 0%, #1e293b 100%) !important; }
    .bg-items-erp { background: linear-gradient(135deg, #0d9488 0%, #0f766e 100%) !important; }
    .bg-cons-erp  { background: linear-gradient(135deg, #ea580c 0%, #c2410c 100%) !important; }
    .bg-size-erp  { background: linear-gradient(135deg, #16a34a 0%, #15803d 100%) !important; }

    /* Hộp trắng bao bọc hình vẽ rập vector hình học gọn gàng 140px */
    .image-placeholder-box-flat {
        border: 1px solid #cbd5e1 !important;
        border-top: none !important; 
        border-radius: 0 0 6px 6px !important;
        padding: 10px 5px !important;
        height: 140px !important; 
        display: flex !important;
        align-items: center !important;
        justify-content: center !important;
        box-sizing: border-box !important;
        margin-bottom: 25px !important;
        background-color: #ffffff !important;
    }
    .image-placeholder-box-flat img {
        max-height: 110px !important;
        width: auto !important;
        object-fit: contain !important;
        display: block !important;
        margin: auto !important;
    }

    /* 🌟 FIX TRIỆT ĐỂ: SỬA LỖI ẨN ẢNH VÀ TRẢ LẠI HIỂN THỊ TỰ ĐỘNG CHO SKETCH 🌟 */
    div[data-testid="stImage"] img {
        width: 100% !important;
        height: auto !important;
    }
    
    .cad-header-text-flat {
        font-family: 'Segoe UI', sans-serif !important;
        font-size: 14px !important;
        font-weight: 700 !important;
        color: #0369a1 !important; 
        margin-bottom: 15px !important;
        padding-bottom: 6px !important;
        border-bottom: 2px solid #e2e8f0 !important;
    }

    .meta-box-light-flat {
        background-color: #f8fafc !important; 
        border-left: 4px solid #0284c7 !important;
        padding: 8px 12px !important;
        margin-bottom: 8px !important;
        border-radius: 0 6px 6px 0 !important;
    }
    .meta-label-flat { font-size: 11px !important; font-weight: 700 !important; color: #64748b !important; text-transform: uppercase !important; }
    .meta-value-flat { font-size: 13px !important; font-weight: 600 !important; color: #0f172a !important; margin-top: 1px !important; }

    /* Khóa chết và ép ẩn toàn diện mọi class ghim đỉnh hoặc hàng rỗng cũ bị dính đệm */
    .main-body-spacer, 
    .sticky-top-container, 
    div[smart-fixed-container], 
    div[data-testid="stHorizontalBlock"]:empty {
        display: none !important;
        height: 0px !important;
        margin: 0 !important;
        padding: 0 !important;
    }
</style>
""", unsafe_allow_html=True)





import streamlit as st
import re

# =====================================================================
# KHỞI TẠO DỮ LIỆU ĐỂ TRÁNH LỖI BIẾN CHƯA ĐỊNH NGHĨA (NAMEERROR)
# =====================================================================
kpi_style_id = st.session_state.get("style_id", "N/A")
total_materials = 0
main_fabric_cons = "0.00"
active_size_kpi = "M"

# Khởi tạo các giá trị session state mặc định nếu chưa có
if "pdf_name" not in st.session_state: st.session_state.pdf_name = ""
if "pdf_bytes" not in st.session_state: st.session_state.pdf_bytes = None
if "pdf_text_cache" not in st.session_state: st.session_state.pdf_text_cache = None

# =====================================================================
# ĐOẠN B: GIAO DIỆN HIỂN THỊ KPIs MÀU SẮC ĐỘNG & GRID THÂN TRANG HỢP NHẤT
# =====================================================================

# 🌟 TIÊU ĐỀ ĐÃ ĐỔI SANG MÀU XANH THEME ERP SANG TRỌNG 🌟
st.markdown(
    """
    <div style="background: linear-gradient(135deg, #0f766e 0%, #115e59 100%); border-radius: 6px; padding: 14px 20px; margin-bottom: 20px; box-shadow: 0 4px 6px -1px rgba(15, 118, 110, 0.1), 0 2px 4px -1px rgba(15, 118, 110, 0.06); text-align: center;">
        <h2 style="font-family: 'Segoe UI', sans-serif; font-size: 16px; font-weight: 700; color: #ffffff; margin: 0; text-transform: uppercase; letter-spacing: 0.8px;">
            🚀 AUTOMATED CAD CONSUMPTION & INDUSTRIAL COSTING ENGINE
        </h2>
    </div>
    """, 
    unsafe_allow_html=True
)

# Chuỗi mã hóa hình ảnh vector đồ họa gốc của 4 ô trang phục
encoded_ao = "data:image/svg+xml;utf8,%3Csvg%20xmlns%3D%27http%3A%2F%2Fwww.w3.org%2F2000%2Fsvg%27%20width%3D%27100%27%20height%3D%27100%27%20viewBox%3D%270%200%2024%2024%27%20fill%3D%27none%27%20stroke%3D%27%23334155%27%20stroke-width%3D%271.25%27%20stroke-linecap%3D%27round%27%20stroke-linejoin%3D%27round%27%3E%3Cpath%20d%3D%27M20.38%203.46L16%202a4%204%200%200%200-8%200l-4.38%201.46a2%202%200%200%200-1.37%202l.35%2011.23a2%202%200%200%200%202%201.94h14.8a2%202%200%200%200%202-1.94l.35-11.23a2%202%200%200%200-1.37-2z%27%2F%3E%3Cpath%20d%3D%27M12%205v16%27%2F%3E%3Cpath%20d%3D%27M4%2010h16%27%2F%3E%3C%2Fsvg%3E"
encoded_quan = "data:image/svg+xml;utf8,%3Csvg%20xmlns%3D%27http%3A%2F%2Fwww.w3.org%2F2000%2Fsvg%27%20width%3D%27100%27%20height%3D%27100%27%20viewBox%3D%270%200%2024%2024%27%20fill%3D%27none%27%20stroke%3D%27%230f766e%27%20stroke-width%3D%271.25%27%20stroke-linecap%3D%27round%27%20stroke-linejoin%3D%27round%27%3E%3Cpath%20d%3D%27M4%202h16l-2%2020H6L4%202z%27%2F%3E%3Cpath%20d%3D%27M12%202v20%27%2F%3E%3Cpath%20d%3D%27M5%208h14%27%2F%3E%3C%2Fsvg%3E"
encoded_vest = "data:image/svg+xml;utf8,%3Csvg%20xmlns%3D%27http%3A%2F%2Fwww.w3.org%2F2000%2Fsvg%27%20width%3D%27100%27%20height%3D%27100%27%20viewBox%3D%270%200%2024%2024%27%20fill%3D%27none%27%20stroke%3D%27%23c2410c%27%20stroke-width%3D%271.25%27%20stroke-linecap%3D%27round%27%20stroke-linejoin%3D%27round%27%3E%3Cpath%20d%3D%27M4%202v20l8-4%208%204V2l-8%204-8-4z%27%2F%3E%3Cpath%20d%3D%27M12%206v12%27%2F%3E%3Cpath%20d%3D%27M4%208h16%27%2F%3E%3C%2Fsvg%3E"
encoded_vay = "data:image/svg+xml;utf8,%3Csvg%20xmlns%3D%27http%3A%2F%2Fwww.w3.org%2F2000%2Fsvg%27%20width%3D%27100%27%20height%3D%27100%27%20viewBox%3D%270%200%2024%2024%27%20fill%3D%27none%27%20stroke%3D%27%2315803d%27%20stroke-width%3D%271.25%27%20stroke-linecap%3D%27round%27%20stroke-linejoin%3D%27round%27%3E%3Cpath%20d%3D%27M6%202h12l3%207-9%2013-9-7%203-7z%27%2F%3E%3Cpath%20d%3D%27M6%209h12%27%2F%3E%3Cpath%20d%3D%27M12%202v7%27%2F%3E%3C%2Fsvg%3E"

# Phân bổ lưới 4 ô KPIs Native gốc của Streamlit
k_col1, k_col2, k_col3, k_col4 = st.columns(4)

with k_col1: 
    st.markdown(f'<div class="kpi-box-flat-matrix bg-style-erp"><div class="kpi-num-flat-matrix">{kpi_style_id}</div><div class="kpi-lbl-flat-matrix">Mã hàng đang xử lý</div></div>', unsafe_allow_html=True)
    st.markdown(f'<div class="image-placeholder-box-flat"><img src="{encoded_ao}" alt="Ao"></div>', unsafe_allow_html=True)

with k_col2: 
    st.markdown(f'<div class="kpi-box-flat-matrix bg-items-erp"><div class="kpi-num-flat-matrix">{total_materials} Item(s)</div><div class="kpi-lbl-flat-matrix">Tổng số vật tư kết xuất</div></div>', unsafe_allow_html=True)
    st.markdown(f'<div class="image-placeholder-box-flat"><img src="{encoded_quan}" alt="Quan"></div>', unsafe_allow_html=True)

with k_col3: 
    st.markdown(f'<div class="kpi-box-flat-matrix bg-cons-erp"><div class="kpi-num-flat-matrix">{main_fabric_cons}</div><div class="kpi-lbl-flat-matrix">Định mức vải chính dự kiến</div></div>', unsafe_allow_html=True)
    st.markdown(f'<div class="image-placeholder-box-flat"><img src="{encoded_vest}" alt="Vest"></div>', unsafe_allow_html=True)

with k_col4: 
    st.markdown(f'<div class="kpi-box-flat-matrix bg-size-erp"><div class="kpi-num-flat-matrix">{active_size_kpi}</div><div class="kpi-lbl-flat-matrix">Cỡ hạt tính định mức</div></div>', unsafe_allow_html=True)
    st.markdown(f'<div class="image-placeholder-box-flat"><img src="{encoded_vay}" alt="Vay"></div>', unsafe_allow_html=True)

# --- BẢNG ĐIỀU KHIỂN SIDEBAR MÁY CHỦ ---
st.sidebar.markdown("### ⚙️ ENGINE CONTROLS")
if st.sidebar.button("🗑️ CLEAR SYSTEM MEMORY", use_container_width=True):
    # 🚨 ĐÃ SỬA: Chuyển về dict trống thay vì None để tránh lỗi crash hệ thống
    st.session_state.bom_data = {}
    st.session_state.chat_history = []
    st.session_state.pdf_bytes = None
    st.session_state.pdf_name = ""
    st.session_state.pdf_text_cache = None
    
    # 🚨 ĐÃ THÊM: Xóa sạch dữ liệu bảng chi tiết CAD và tổng hợp định mức
    if "processed_display_rows" in st.session_state: 
        st.session_state.processed_display_rows = []
    if "accumulated_bom_rows" in st.session_state: 
        st.session_state.accumulated_bom_rows = []
        
    if "last_active_blueprint" in st.session_state: st.session_state.last_active_blueprint = None
    if "raw_ai_debug_payload" in st.session_state: st.session_state.raw_ai_debug_payload = None
    if "pdf_page_one_image" in st.session_state: st.session_state.pdf_page_one_image = None
    
    # Ép hệ thống render lại giao diện trống sạch sẽ ngay lập tức
    st.rerun()



# ------------------------------------------------------------------------------
# LƯỚI CHIA ĐÔI CỘT CHÍNH THỰC TẾ (SỬ DỤNG HEIGHT NATIVE CỦA STREAMLIT)
# ------------------------------------------------------------------------------
col_left, col_right = st.columns(2)

# --- CỘT TRÁI: BỘ TẢI FILE & HỒ SƠ TÓM TẮT MÃ HÀNG MÀU XANH ---
with col_left:
    # Ép chiều cao native bằng tham số height, tự động sinh thanh cuộn nếu tràn nội dung
    with st.container(border=True, height=520):
        st.markdown("### 📂 TECHPACK UPLOADER & PROFILE SUMMARY")
        
        uploaded_file = st.file_uploader("Upload PDF", type=["pdf"], label_visibility="collapsed")
        
        if uploaded_file is not None:
            if st.session_state.pdf_name != uploaded_file.name:
                st.session_state.pdf_text_cache = None
                if "pdf_page_one_image" in st.session_state: st.session_state.pdf_page_one_image = None
                if "accumulated_bom_rows" in st.session_state: st.session_state.accumulated_bom_rows = []
            st.session_state.pdf_bytes = uploaded_file.read()
            st.session_state.pdf_name = uploaded_file.name

        if st.session_state.pdf_text_cache is not None:
            st.markdown("<div style='margin-top: 5px;'></div>", unsafe_allow_html=True)
            txt = st.session_state.pdf_text_cache
            
            def get_meta(pattern, default="N/A"):
                m = re.search(pattern, txt, re.IGNORECASE)
                return m.group(1).strip() if m else default

            style_id = get_meta(r'(?:Style ID|Style_ID|Mã hàng)\s*[:\-=\s]*([\w\d\-]+)', st.session_state.pdf_name.replace(".pdf",""))
            short_desc = get_meta(r'(?:Short Desc|Description|Tên sản phẩm)\s*[:\-=\s]*([^\n]+)', "THE RUCHED MINI DRESS")
            customer = get_meta(r'(?:Customer|Khách hàng|Brand)\s*[:\-=\s]*([^\n]+)', "FACTORY STANDARD")
            season = get_meta(r'(?:Season|Mùa hàng)\s*[:\-=\s]*([^\n]+)', "FALL Winter 2026")
            fabric_type = get_meta(r'(?:Long Description|Chất liệu gốc)\s*[:\-=\s]*([^\n]+)', "POPLIN FABRIC COTTON - SP26")

            m_col1, m_col2 = st.columns(2)
            with m_col1:
                st.markdown(f'<div class="meta-box-light-flat"><div class="meta-label-flat">Style Code / Mã hàng</div><div class="meta-value-flat"><b>{style_id}</b></div></div>', unsafe_allow_html=True)
                st.markdown(f'<div class="meta-box-light-flat"><div class="meta-label-flat">Customer / Đối tác</div><div class="meta-value-flat">{customer}</div></div>', unsafe_allow_html=True)
                st.markdown(f'<div class="meta-box-light-flat"><div class="meta-label-flat">Season / Mùa sản xuất</div><div class="meta-value-flat">{season}</div></div>', unsafe_allow_html=True)
            with m_col2:
                st.markdown(f'<div class="meta-box-light-flat"><div class="meta-label-flat">Garment Type / Kiểu dáng</div><div class="meta-value-flat">{short_desc}</div></div>', unsafe_allow_html=True)
                st.markdown(f'<div class="meta-box-light-flat"><div class="meta-label-flat">Material Spec / Mô tả vải</div><div class="meta-value-flat">{fabric_type[:25]}...</div></div>', unsafe_allow_html=True)
                st.markdown(f'<div class="meta-box-light-flat"><div class="meta-label-flat">Techpack Status</div><div class="meta-value-light" style="color: #16a34a; font-weight: bold;">🟢 READY TO BOM</div></div>', unsafe_allow_html=True)
        else:
            if st.session_state.pdf_bytes is None:
                st.markdown("<div style='margin-top: 20px; text-align: center; color: #64748b; font-size: 13px;'>Bảng tóm tắt hồ sơ trống. Vui lòng tải tài liệu lên hệ thống.</div>", unsafe_allow_html=True)

# --- CỘT PHẢI: KHÔNG GIAN HIỂN THỊ THÔNG TIN HÌNH ẢNH SKETCH ---
with col_right:
    with st.container(border=True, height=520):
        st.markdown("### 🎨 TECHPACK SKETCH VISUALIZER")
        
        # Hiển thị hình vẽ phác thảo nguyên bản mượt mà
        if "pdf_page_one_image" in st.session_state and st.session_state.pdf_page_one_image is not None:
            st.image(st.session_state.pdf_page_one_image, use_container_width=True)
        else:
            st.markdown("<div style='margin-top: 60px; text-align: center; color: #64748b; font-size: 13px;'>Chưa có hình ảnh phác thảo. Vui lòng tải Techpack PDF để trích xuất hệ thống.</div>", unsafe_allow_html=True)











import hashlib
import json
import re
import fitz
import google.generativeai as genai
import streamlit as st


# =====================================================================
# 🧠 ĐOẠN A (NÂNG CẤP QUET TOÀN DIỆN BOM): KHỐI HÀM CACHE AI (ĐÃ SỬA LỖI KHỔ VẢI)
# =====================================================================
@st.cache_data(
    show_spinner=False,
    ttl=60,
    hash_funcs={bytes: lambda b: hashlib.sha256(b).hexdigest()},
)
def execute_cached_gemini_scan(
    pdf_bytes,
    current_query,
    active_width,
    target_size_cmd,
    raw_json_schema,
    prompt_agent_2,
):
    """Hàm gọi AI quét TOÀN BỘ các trang trong file Techpack để bóc tách trọn

    vẹn cấu trúc Vải chính, Vải lót và Keo lót (Fusing).
    """
    import copy
    import hashlib

    if hasattr(pdf_bytes, "getvalue"):
        pdf_bytes = pdf_bytes.getvalue()

    if not isinstance(pdf_bytes, bytes):
        raise TypeError("Dữ liệu PDF đầu vào không đúng định dạng bytes hợp lệ!")

    full_pdf_raw_text = ""
    image_payloads = []

    with fitz.open(stream=pdf_bytes, filetype="pdf") as doc_recovery:
        total_pages = len(doc_recovery)

        # 🚨 ĐÃ SỬA: Quét toàn bộ số trang của file Techpack để tìm sạch linh kiện keo/phụ liệu ở trang sau
        for idx in range(total_pages):
            page_text = doc_recovery[idx].get_text("text")
            full_pdf_raw_text += f"\n--- PAGE {idx + 1} ---\n{page_text}"

            # Chỉ render ảnh cho 5 trang đầu hoặc trang chứa hình vẽ rập để tối ưu hóa dung lượng gửi đi
            if len(image_payloads) < 5:
                try:
                    pix = doc_recovery[idx].get_pixmap(
                        dpi=72, colorspace=fitz.csRGB
                    )
                    image_payloads.append(
                        {
                            "mime_type": "image/jpeg",
                            "data": pix.tobytes("jpeg"),
                        }
                    )
                except Exception:
                    continue

    gemini_inputs = copy.deepcopy(image_payloads)
    gemini_inputs.insert(
        0,
        f"=== USER CHAT COMMAND ===\n{current_query}\n\n=== TECHPACK TEXT ===\n{full_pdf_raw_text}\n",
    )

    # 🚨 ĐÃ CẬP NHẬT PROMPT ÉP ĐẦU RA TOÀN DIỆN NGUYÊN PHỤ LIỆU
    extended_prompt = (
        prompt_agent_2
        + """
    CRITICAL MULTI-MATERIAL EXTRACTION RULES:
    - You MUST extract EVERY SINGLE component listed in the document, not just FABRIC.
    - Carefully scan for pocket linings, waist linings, and fusing/interfacing descriptors.
    - If a component name contains "FUSING", "INTERLINING", "MEX", "DỰNG", "KEO LOT", classify its material_class strictly as "FUSING".
    - If a component name contains "LINING", "POCKET BAG", "LOT TUI", classify its material_class strictly as "LINING".
    - Do not omit any minor panels or components from the final JSON structure.
    """
    )
    gemini_inputs.append(extended_prompt)

    model = genai.GenerativeModel("gemini-2.5-flash")
    response = model.generate_content(
        gemini_inputs,
        generation_config={
            "response_mime_type": "application/json",
            "response_schema": raw_json_schema,
            "temperature": 0.0,
        },
        request_options={"timeout": 120.0},
    )

    if not response or not response.text:
        raise RuntimeError("Mô hình Gemini trả về kết quả rỗng!")

    txt = response.text.strip()
    if txt.startswith("```"):
        txt = re.sub(r"^```json\s*", "", txt)
        txt = re.sub(r"^```\s*", "", txt)
        txt = re.sub(r"\s*```$", "", txt)
    txt = txt.strip()

    try:
        blueprint_worker = json.loads(txt)
    except json.JSONDecodeError as json_err:
        raise RuntimeError(
            f"Mô hình Gemini trả về cấu trúc chuỗi JSON không hợp lệ:\n\n{txt}"
        ) from json_err

    if blueprint_worker and "bom_rows" in blueprint_worker:
        blueprint_worker["calculated_on_size"] = target_size_cmd
        for row in blueprint_worker.get("bom_rows", []):
            if "component_name" in row:
                row["component_name"] = " ".join(
                    str(row["component_name"]).upper().split()
                )
            try:
                row["bounding_box_length"] = round(
                    float(row.get("bounding_box_length", 0.0)), 2
                )
            except Exception:
                row["bounding_box_length"] = 0.0
            try:
                row["bounding_box_width"] = round(
                    float(row.get("bounding_box_width", 0.0)), 2
                )
            except Exception:
                row["bounding_box_width"] = 0.0
            try:
                row["polygon_net_area"] = float(row.get("polygon_net_area", 0.0))
            except Exception:
                row["polygon_net_area"] = 0.0
            try:
                row["piece_count"] = int(float(row.get("piece_count", 1)))
            except Exception:
                row["piece_count"] = 1
            try:
                row["gross_consumption"] = round(
                    float(row.get("gross_consumption", 0.0415)), 4
                )
            except Exception:
                row["gross_consumption"] = 0.0415
            try:
                row["marker_efficiency"] = str(
                    row.get("marker_efficiency", "82.5%")
                ).strip()
            except Exception:
                row["marker_efficiency"] = "82.5%"
            
            # 🛠️ ĐÃ SỬA: ÉP ĐÈ KHỔ VẢI THEO Ô CHAT VÀO TỪNG DÒNG RẬP CHỐNG KẸT CACHE 56 CŨ
            try:
                forced_width = float(active_width)
                if current_query:
                    width_match = re.search(r"(khổ\s*vải|khổ)\s*(\d+(\.\d+)?)", str(current_query), re.IGNORECASE)
                    if width_match:
                        forced_width = float(width_match.group(2))
                
                # Gán thẳng khổ vải vừa quét động được vào dữ liệu chi tiết rập
                row["fabric_width_inch"] = forced_width
            except Exception:
                row["fabric_width_inch"] = float(active_width)

    return blueprint_worker





import streamlit as st

# =====================================================================
# 🟩 ĐOẠN 1: CHAT WORKSPACE LAYER (CHỐNG KẸT LUỒNG & PHÁT LỆNH)
# =====================================================================

# 1. Khởi tạo an toàn bộ nhớ đệm hệ thống (Session State)
if "chat_history" not in st.session_state:
    st.session_state.chat_history = []
if "ai_processing" not in st.session_state:
    st.session_state.ai_processing = False
if "last_submitted_query" not in st.session_state:
    st.session_state.last_submitted_query = ""

# 2. Tạo một khung Container riêng độc lập để chứa lịch sử hội thoại cũ
chat_history_container = st.container()
with chat_history_container:
    st.markdown('<br><div class="cad-card"><div class="cad-header">💬 CHATGPT IE COLLABORATION WORKSPACE</div></div>', unsafe_allow_html=True)
    if st.session_state.get("chat_history"):
        for msg in st.session_state.chat_history:
            st.chat_message("user").write(msg["user"])
            st.chat_message("assistant").write(msg["ai"])

# 🚨 ĐÃ SỬA: Đặt sát lề trái ngoài cùng, đổi key sang _v8 mới tinh để giải phóng hoàn toàn bộ nhớ đệm kẹt cũ
safe_user_prompt = st.chat_input(
    "Gõ lệnh tính toán (Ví dụ: tính định mức cỡ 32 khổ 56 co rút dọc 3 ngang 14)...",
    key="ie_workspace_fixed_dynamic_chat_final_patch_v8"
)

# 3. Kích hoạt cờ hiệu xử lý và ép tải lại luồng chính khi người dùng gửi thành công
if safe_user_prompt:
    st.session_state["last_submitted_query"] = str(safe_user_prompt).strip()
    st.session_state.ai_processing = True
    st.rerun()

# =====================================================================
# 🟩 ĐOẠN 2 (BẢN UPDATE PROMPT CAD HÌNH HỌC): AI CORE ENGINE
# =====================================================================

if st.session_state.ai_processing:
    current_query = st.session_state["last_submitted_query"]

    # Bộ quét tự động tìm file PDF dự phòng từ bộ nhớ đệm dùng chung khi chuyển tab
    active_pdf = st.session_state.get("pdf_bytes")
    if active_pdf is None:
        active_pdf = (
            st.session_state.get("uploaded_file")
            or st.session_state.get("current_pdf")
            or st.session_state.get("pdf_data")
        )

    # 🛠️ TRÍCH XUẤT KHỔ VẢI ĐỘNG TỪ Ô CHAT ĐỂ BẺ GÃY SỐ 56 CỐ ĐỊNH
    dynamic_width = 56.0  # Giá trị mặc định phòng hờ
    target_size = "32"    # Cỡ mẫu mặc định
    
    if current_query:
        import re
        width_match = re.search(r"(khổ\s*vải|khổ)\s*(\d+(\.\d+)?)", str(current_query), re.IGNORECASE)
        if width_match:
            dynamic_width = float(width_match.group(2))
            
        size_match = re.search(r"(cỡ|size)\s*(\d+)", str(current_query), re.IGNORECASE)
        if size_match:
            target_size = str(size_match.group(2))

    if active_pdf is not None:
        with st.spinner(
            "🧠 AI Vision đang trích xuất và gắn nhãn rập cấu trúc kỹ thuật..."
        ):
            try:
                # 1. JSON SCHEMA MỞ RỘNG MÁ TRẬN ĐA GIÁC CAD
                raw_json_schema = {
                    "type": "OBJECT",
                    "properties": {
                        "detected_product_type": {
                            "type": "STRING",
                            "description": "Kiểu dáng sản phẩm, ví dụ: JEANS, JACKET, SHIRT",
                        },
                        "detected_base_size": {
                            "type": "STRING",
                            "description": "Size mẫu trích xuất, ví dụ: 32",
                        },
                        "bom_rows": {
                            "type": "ARRAY",
                            "description": "Danh sách chi tiết thông số hình học thô bóc tách từ Techpack",
                            "items": {
                                "type": "OBJECT",
                                "properties": {
                                    "component_name": {
                                        "type": "STRING",
                                        "description": "Tên chi tiết rập gốc từ tài liệu",
                                    },
                                    "material_class": {
                                        "type": "STRING",
                                        "description": "Allowed: FABRIC, LINING, FUSING, TAPE, ELASTIC, RIB, TRIM, THREAD, ACCESSORY",
                                    },
                                    "geometry_role": {
                                        "type": "STRING",
                                        "description": "Allowed: MAJOR_PANEL hoặc MINOR_COMPONENT",
                                    },
                                    "piece_type": {
                                        "type": "STRING",
                                        "description": "BẮT BUỘC phân loại nhãn rập chuẩn ngành.",
                                    },
                                    "uom": {
                                        "type": "STRING",
                                        "description": "Cố định: YDS",
                                    },
                                    "piece_count": {
                                        "type": "INTEGER",
                                        "description": "Số lượng rập chi tiết gốc (Pcs).",
                                    },
                                    "bounding_box_length": {
                                        "type": "NUMBER",
                                        "description": "Chiều dài chi tiết rập L-inch.",
                                    },
                                    "bounding_box_width": {
                                        "type": "NUMBER",
                                        "description": "Chiều rộng chi tiết rập W-inch.",
                                    },
                                    "data_confidence": {
                                        "type": "STRING",
                                        "description": "Allowed: HIGH, LOW",
                                    },
                                    "calculation_status": {
                                        "type": "STRING",
                                        "description": "Allowed: READY, MISSING_INPUT",
                                    },
                                },
                                "required": [
                                    "component_name",
                                    "material_class",
                                    "geometry_role",
                                    "piece_type",
                                    "uom",
                                    "piece_count",
                                    "bounding_box_length",
                                    "bounding_box_width",
                                    "data_confidence",
                                    "calculation_status",
                                ],
                            },
                        },
                    },
                    "required": [
                        "detected_product_type",
                        "detected_base_size",
                        "bom_rows",
                    ],
                }
                # 2. PROMPT CHUYÊN GIA CAD TRÍCH XUẤT & SUY LUẬN KHÔNG GIAN RẬP
                prompt_agent_2 = """
                You are a senior Industrial Garment IE & CAD Pattern Engineering Intelligence. Your absolute priority is to extract or intelligently estimate the physical dimensions (Length and Width in INCHES) for EVERY garment component found in the Techpack.
                
                🚨 CRITICAL DIMENSION RETRIEVAL & ESTIMATION DIRECTIVES (ANTI-ZERO RULE):
                1. PRIMARY SOURCE (TABLES): Search all pages for spec tables, graded measurement sheets, or marker detail blocks. Extract the exact 'bounding_box_length' and 'bounding_box_width' for the base size (Size 32).
                
                2. SECONDARY SOURCE (PATTERN SKETCH ESTIMATION): If numerical dimensions are NOT explicitly written in a structured table, you MUST estimate the bounding rectangle directly from the technical flat sketches, drawings, or pattern diagrams.
                   - Use the overall garment proportions and drawing scale to approximate the bounding box.
                   - NEVER output 0.0 or null for 'bounding_box_width' or 'bounding_box_length' if the component physically exists in the design sketch. A pant leg or waistband CANNOT have a width of 0.
                
                3. INDUSTRIAL GARMENT HEURISTIC BOUNDS (GUARDRAILS FOR JEANS/PANTS):
                   If you cannot find exact metrics and must estimate from the sketch, apply these standard apparel industry geometric boundaries to prevent mathematically impossible zeros:
                   - TROUSER_FRONT / TROUSER_BACK (Main Leg Panels): Length typically ranges between 35.0 to 45.0 inches. Width MUST be estimated based on the leg opening/thigh ratio, typically between 10.0 to 16.0 inches. NEVER leave width as 0.
                   - WAISTBAND / LƯNG QUẦN: Length is tied to waist size (e.g., Size 32 approx 32.0-35.0 inches if flat/curved). Width is standard waistband height, typically 1.5 to 2.5 inches.
                   - POCKET BAG / LÓT TÚI: Length typically 10.0 to 13.0 inches, Width typically 6.0 to 8.0 inches.
                   - COIN POCKET / FLAP / MINOR PIECES: Length 3.0 to 5.0 inches, Width 3.0 to 5.0 inches.
                
                4. DATA CONFIDENCE & STATUS LOGIC:
                   - Set data_confidence = "HIGH" and calculation_status = "READY" ONLY if the numbers are explicitly extracted from text/tables.
                   - Set data_confidence = "LOW" and calculation_status = "READY" if the dimensions are mathematically estimated/reconstructed from the pattern sketch or garment guardrails.
                   - ONLY set calculation_status = "MISSING_INPUT" if the component name is mentioned but there is absolutely zero visual representation or context to infer size.
                
                Ensure your output strictly adheres to the requested JSON structure. Every valid component must have non-zero geometric properties to allow proper 2D packing area calculation.
                """

                # 3. GỌI HÀM QUÉT AI CACHE VÀ TRUYỀN KHỔ VẢI ĐỘNG ĐÃ TRÍCH XUẤT
                bom_data = execute_cached_gemini_scan(
                    pdf_bytes=active_pdf,
                    current_query=current_query,
                    active_width=dynamic_width,  # ✅ ĐÃ SỬA: Thay số 56.0 cố định cũ bằng biến động dynamic_width
                    target_size_cmd=target_size,
                    raw_json_schema=raw_json_schema,
                    prompt_agent_2=prompt_agent_2
                )
                
                # Lưu kết quả bóc tách vào session_state
                st.session_state["bom_data"] = bom_data
                if bom_data and "bom_rows" in bom_data:
                    st.session_state["accumulated_bom_rows"] = bom_data["bom_rows"]
                
                st.session_state.ai_processing = False
                st.success("✅ AI Core đã đồng bộ cấu trúc rập thành công!")
                st.rerun()

            except Exception as e:
                st.session_state.ai_processing = False
                st.error(f"❌ Lỗi xử lý AI Core Engine: {str(e)}")
                import traceback
                st.text(traceback.format_exc())





import pandas as pd
import streamlit as st
import re

def initialize_and_sync_parameters():
    """Khối 1: Trích xuất và đồng bộ thông số vải, co rút, kích cỡ thời gian thực"""
    if not (st.session_state.get("bom_data") or st.session_state.get("accumulated_bom_rows")):
        return None, None
        
    bom_source = st.session_state.get("bom_data", {})
    
    # 1. Trích xuất text từ ô chat câu lệnh người dùng
    user_query_text = ""
    if st.session_state.get("last_submitted_query"): 
        user_query_text = str(st.session_state.get("last_submitted_query"))
    elif st.session_state.get("ie_workspace_static_chat_input_key"): 
        user_query_text = str(st.session_state.get("ie_workspace_static_chat_input_key"))
    if not user_query_text and st.session_state.get("chat_history"): 
        user_query_text = str(st.session_state.chat_history[-1]["user"])

    # 2. Thiết lập thông số mặc định ban đầu từ file gốc
    fabric_width = bom_source.get("fabric_width_inch", 56.0)
    warp_shrinkage = bom_source.get("warp_shrinkage_percent", 0.0)
    weft_shrinkage = bom_source.get("weft_shrinkage_percent", 0.0)
    
    detected_size = bom_source.get("detected_base_size", bom_source.get("calculated_on_size", "32"))
    target_size = str(detected_size).upper()

    # 3. Quét nhanh thông số ép buộc từ chat bằng Regex (nếu có)
    if user_query_text:
        w_match = re.search(r"(khổ\s*vải|khổ)\s*(\d+(\.\d+)?)", user_query_text, re.IGNORECASE)
        if w_match: fabric_width = float(w_match.group(2))
        
        warp_match = re.search(r"(co\s*rút\s*dọc|dọc)\s*(\d+(\.\d+)?)", user_query_text, re.IGNORECASE)
        if warp_match: warp_shrinkage = float(warp_match.group(2))
        
        weft_match = re.search(r"(co\s*rút\s*ngang|ngang)\s*(\d+(\.\d+)?)", user_query_text, re.IGNORECASE)
        if weft_match: weft_shrinkage = float(weft_match.group(2))
        
        size_match = re.search(r"(cỡ|size)\s*([a-zA-Z0-9]+)", user_query_text, re.IGNORECASE)
        if size_match: target_size = str(size_match.group(2)).upper()

    # 4. Ghi đè đồng bộ các thông số vào bộ nhớ hệ thống
    bom_source["fabric_width_inch"] = fabric_width
    bom_source["usable_width_inch"] = fabric_width  
    bom_source["warp_shrinkage_percent"] = warp_shrinkage
    bom_source["weft_shrinkage_percent"] = weft_shrinkage
    bom_source["calculated_on_size"] = target_size
    
    st.session_state["bom_data"] = bom_source
    return bom_source, user_query_text
import re
import streamlit as st

def extract_cutting_instructions_from_pdf(component_name, raw_pdf_text):
    """Thuật toán quét Callout Văn bản PDF: Tự động phân tích các lệnh kỹ thuật 
    (CUT 2, PAIR, SELF, FUSE, MIRROR, FOLD) trực tiếp từ file PDF thay vì gán cứng.
    """
    if not raw_pdf_text:
        return {"layer_multiplier": 1, "is_paired": False, "calc_log": "Không tìm thấy dữ liệu văn bản thô PDF."}
        
    # Chuẩn hóa chuỗi văn bản để tìm kiếm không gian lân cận chi tiết rập
    text_clean = " ".join(str(raw_pdf_text).lower().split())
    comp_clean = str(component_name).lower().strip()
    
    # Thiết lập cấu trúc mặc định theo quy chuẩn dệt may
    layer_multiplier = 1
    is_paired = False
    calc_log = "AI đọc văn bản PDF: Mặc định hệ số kết cấu đơn."
    
    # 1. Thuật toán quét vùng lân cận (Window Scanning): Tìm kiếm Callout kỹ thuật xung quanh tên rập
    match_index = text_clean.find(comp_clean)
    if match_index != -1:
        # Cắt một đoạn văn bản xung quanh tên chi tiết (Phạm vi 150 ký tự) để tìm Callout chỉ định cắt
        window_start = max(0, match_index - 50)
        window_end = min(len(text_clean), match_index + 150)
        scan_window = text_clean[window_start:window_end]
        
        # ➔ A. Quét lệnh số lượng cắt trực tiếp (Ví dụ: CUT 2, CUT 4, CUT 6, SELF X2)
        cut_match = re.search(r'(cut|cắt|self|shell)\s*(x\s*|\s*|\s*=\s*)(\d+)', scan_window)
        if cut_match:
            detected_qty = int(cut_match.group(3))
            layer_multiplier = detected_qty
            calc_log = f"Trích xuất trực tiếp PDF Callout: Tìm thấy lệnh cắt {detected_qty} chi tiết."
            
        # ➔ B. Quét lệnh đối xứng / cặp đôi (PAIR, MIRROR, X2)
        if any(k in scan_window for k in ["pair", "cặp", "đối", "mirror", "đối xứng", "x2"]):
            is_paired = True
            # Nếu lệnh cắt chưa nhân đôi, tự động gán kết cấu cặp
            if layer_multiplier == 1:
                layer_multiplier = 2
                calc_log = f"Trích xuất trực tiếp PDF Callout: Phát hiện cấu trúc đối xứng cặp (PAIR/MIRROR)."
                
        # ➔ C. Quét lệnh gập đôi vải (FOLD, GẬP ĐÔI)
        if any(k in scan_window for k in ["fold", "gập", "gap doi", "gập đôi"]):
            layer_multiplier = max(layer_multiplier, 2)
            calc_log += " | Phát hiện chi tiết đi biên gập đôi (FOLD)."
            
    return {
        "layer_multiplier": layer_multiplier,
        "is_paired": is_paired,
        "calc_log": calc_log
    }


import numpy as np

import numpy as np

def calculate_skyline_2d_metrics(bom_rows_list, user_query_text):
    """Khối 2b Siêu Cấp: Mô phỏng hình học phi tuyến tính Gerber Core Engine.
    ĐÃ ĐỒNG BỘ: Kết nối hoàn hảo với trường diện tích ước lượng của Khối 2a,
    giúp giải phóng hàm toán học để định mức tự động nhảy lên con số thực tế 1.5 - 2.6 YDS.
    """
    ctx = classify_pieces_and_products(bom_rows_list, user_query_text)
    if not ctx or not ctx.get("stable_bom_list"):
        return {"product_segmented": "GENERIC_TOP", "fabric_pattern": "SOLID", "actual_packing_density": 0.80, "global_gross_fabric_yds": 1.85, "major_shape_area": 0.0}

    fabric_pattern = ctx["fabric_pattern"]
    fabric_width = ctx["fabric_width"]
    stable_bom = ctx["stable_bom_list"]

    # =====================================================================
    # 1. ĐỌC DỮ LIỆU DIỆN TÍCH ĐỘNG TỪ BỘ PARSER KHỐI 2A
    # =====================================================================
    total_net_area = 0.0
    total_bbox_area = 0.0
    total_piece_count = 0.0
    all_expanded_pieces = []
    
    for r in stable_bom:
        try:
            pcs = float(r.get("piece_count", r.get("Số lượng rập", 1.0)))
            if pcs <= 0: pcs = 1.0
        except:
            pcs = 1.0
            
        l_inch = float(r.get("bounding_box_length", r.get("Dài (L-inch)", 0.0)))
        w_inch = float(r.get("bounding_box_width", r.get("Rộng (W-inch)", 0.0)))
        bbox_a = l_inch * w_inch
        
        # Đọc diện tích tịnh vừa được Khối 2a sinh ra
        net_a = float(r.get("polygon_net_area", 0.0))
        if net_a <= 0:
            net_a = bbox_a * 0.74 # Fallback an toàn nếu rập trống diện tích
            
        total_net_area += net_a * pcs
        total_bbox_area += bbox_a * pcs
        total_piece_count += pcs
        
        for _ in range(int(pcs)):
            all_expanded_pieces.append({
                "net_area": net_a, "bbox_area": bbox_a, "length": l_inch, "width": w_inch
            })

    # =====================================================================
    # 2. TRÍCH XUẤT ĐẶC TRƯNG HÌNH HỌC PHI TUYẾN TÍNH CHUẨN GERBER ENGINE
    # =====================================================================
    # Tự động phân loại rập lớn (Major) động dựa theo tỷ lệ đóng góp diện tích (>8%)
    major_threshold_area = total_net_area * 0.08 if total_net_area > 0 else 50.0
    major_pieces_list = [p for p in all_expanded_pieces if p["net_area"] > major_threshold_area]
    minor_pieces_list = [p for p in all_expanded_pieces if p["net_area"] <= major_threshold_area]
    
    # Tính toán chính xác tỷ lệ phân mảnh dựa trên số lượng chiếc rập phụ thực tế
    fragmentation_ratio = len(minor_pieces_list) / total_piece_count if total_piece_count > 0 else 0.20
    bounding_box_fill = total_net_area / total_bbox_area if total_bbox_area > 0 else 0.72

    if major_pieces_list:
        avg_aspect_ratio = sum(max(p["length"], p["width"]) / max(min(p["length"], p["width"]), 0.1) for p in major_pieces_list) / len(major_pieces_list)
        avg_major_width = sum(p["width"] for p in major_pieces_list) / len(major_pieces_list)
        width_occupancy_ratio = avg_major_width / fabric_width
    else:
        avg_aspect_ratio = 1.8
        width_occupancy_ratio = 0.28

    convexity_score = bounding_box_fill  
    rotation_freedom_factor = 0.95 if "one-way" in str(user_query_text).lower() else 1.0
    compactness_score = max(min(1.0 - (abs(avg_aspect_ratio - 1.0) * 0.05), 1.0), 0.60)
    
    minor_area_sum = sum(p["net_area"] for p in minor_pieces_list)
    small_piece_ratio = minor_area_sum / total_net_area if total_net_area > 0 else 0.15
    marker_fragmentation = total_piece_count / (total_net_area / 100.0) if total_net_area > 0 else 1.0
    edge_irregularity = 1.0 - convexity_score

    # Hàm Logistic Curve tính toán bộ phạt không gian khi rập to chiếm khổ vải lớn (>32%)
    logistic_midpoint = 0.32
    logistic_k = 18.0  
    width_penalty_logistic = 0.08 / (1.0 + np.exp(-logistic_k * (width_occupancy_ratio - logistic_midpoint)))

    # =====================================================================
    # 3. TÍNH TOÁN MẬT ĐỘ NÈN ĐỘNG (DYNAMIC NESTING DENSITY)
    # =====================================================================
    calculated_density = 0.68 + (bounding_box_fill * 0.12) + (compactness_score * 0.04)
    nesting_efficiency_bonus = (small_piece_ratio * 0.05) + (fragmentation_ratio * 0.03)
    actual_packing_density = (calculated_density + nesting_efficiency_bonus - width_penalty_logistic) * rotation_freedom_factor
    actual_packing_density = max(min(actual_packing_density, 0.92), 0.62)

    # =====================================================================
    # 4. TÍNH CHIỀU DÀI SƠ ĐỒ VÀ BỘ HAO HỤT KHÔNG GIAN SẢN XUẤT ĐỘNG
    # =====================================================================
    if total_net_area <= 0:
        total_net_area = ctx.get("major_shape_area", 0.0) + ctx.get("minor_shape_area", 0.0)
        
    simulated_length = (total_net_area / fabric_width) / actual_packing_density
    simulated_length *= (1.0 + (edge_irregularity * 0.04)) * ctx.get("constraint_penalty", 1.0)

    # Hệ số hao hụt dạt đầu bàn cắt phi tuyến tính (Logistic) dựa trên chiều dài sơ đồ
    length_logistic_mid = 45.0  
    length_k = -0.08
    wastage_curve_factor = 0.01 + (0.15 / (1.0 + np.exp(-length_k * (simulated_length - length_logistic_mid))))
    fabric_wastage_multiplier = 1.015 + wastage_curve_factor
    
    end_loss_inch = 1.5 + (marker_fragmentation * 0.05) + (width_occupancy_ratio * 1.5)
    global_gross_fabric = (simulated_length / 36.0) * fabric_wastage_multiplier + (end_loss_inch / 36.0)

    # =====================================================================
    # 5. XỬ LÝ CHU KỲ VÂN VẢI ĐỘNG (NAP / PLAID)
    # =====================================================================
    fabric_repeat_inch = float(ctx.get("fabric_repeat_inch", 4.0)) 

    if fabric_pattern == "NAP":
        global_gross_fabric += (fabric_repeat_inch * 0.35 * (1.0 - small_piece_ratio)) / 36.0
    elif fabric_pattern in ["PLAID", "STRIPE"]:
        plaid_loss_ratio = (fabric_repeat_inch * 1.35) / simulated_length if simulated_length > 0 else 0.05
        global_gross_fabric *= (1.0 + min(plaid_loss_ratio, 0.35))

    # Ép định mức tối thiểu thực tế cho dòng hàng Jacket người lớn phòng trường hợp rập bị thiếu chi tiết
    if "JACKET" in str(ctx.get("product_type", "")).upper() and global_gross_fabric < 1.2:
        global_gross_fabric = 2.25

    major_area_sum = sum(p["net_area"] for p in major_pieces_list) if major_pieces_list else total_net_area

    return {
        "product_segmented": ctx.get("product_type", "JACKET"), 
        "fabric_pattern": fabric_pattern,
        "actual_packing_density": actual_packing_density, 
        "global_gross_fabric_yds": global_gross_fabric,
        "major_shape_area": major_area_sum  
    }



import re
import streamlit as st

def extract_cutting_instructions_from_pdf(component_name, raw_pdf_text):
    """Thuật toán quét Callout văn bản PDF: Tự động phân tích các lệnh kỹ thuật 
    (CUT 2, PAIR, SELF, FUSE, MIRROR, FOLD) trực tiếp từ dữ liệu văn bản thô của Techpack.
    Loại bỏ hoàn toàn cơ chế phán đoán gán cứng thủ công theo tên.
    """
    if not raw_pdf_text:
        return {"layer_multiplier": 1, "is_paired": False, "calc_log": "CAD Fallback: Không tìm thấy dữ liệu văn bản thô PDF."}
        
    # Chuẩn hóa chuỗi văn bản để tìm kiếm không gian lân cận chi tiết rập
    text_clean = " ".join(str(raw_pdf_text).lower().split())
    comp_clean = str(component_name).lower().strip()
    
    # Thiết lập cấu trúc mặc định theo quy chuẩn dệt may
    layer_multiplier = 1
    is_paired = False
    calc_log = "AI Engine: Mặc định hệ số kết cấu đơn (Cut 1)."
    
    # Tìm vị trí xuất hiện của tên chi tiết rập trong file văn bản PDF Techpack
    match_index = text_clean.find(comp_clean)
    if match_index != -1:
        # Cắt một đoạn văn bản xung quanh tên chi tiết (Phạm vi trước 50 và sau 150 ký tự) để tìm Callout kỹ thuật
        window_start = max(0, match_index - 50)
        window_end = min(len(text_clean), match_index + 150)
        scan_window = text_clean[window_start:window_end]
        
        # ➔ A. Quét lệnh số lượng cắt trực tiếp (Ví dụ: CUT 2, CUT 4, CUT 6, SELF X2, SHELL X4)
        cut_match = re.search(r'(cut|cắt|self|shell)\s*(x\s*|\s*|\s*=\s*)(\d+)', scan_window)
        if cut_match:
            detected_qty = int(cut_match.group(3))
            layer_multiplier = detected_qty
            calc_log = f"Trích xuất trực tiếp PDF Callout: Tìm thấy lệnh cắt {detected_qty} chi tiết."
            
        # ➔ B. Quét lệnh đối xứng / cặp đôi (PAIR, MIRROR, X2)
        if any(k in scan_window for k in ["pair", "cặp", "đối", "mirror", "đối xứng", "x2", "1 pair"]):
            is_paired = True
            # Nếu lệnh cắt chưa nhân đôi, tự động gán kết cấu cặp sản xuất
            if layer_multiplier == 1:
                layer_multiplier = 2
                calc_log = f"Trích xuất trực tiếp PDF Callout: Phát hiện cấu trúc đối xứng cặp (PAIR/MIRROR)."
                
        # ➔ C. Quét lệnh gập đôi vải bàn cắt (FOLD, GẬP ĐÔI, OPEN FOLD)
        if any(k in scan_window for k in ["fold", "gập", "gap doi", "gập đôi"]):
            layer_multiplier = max(layer_multiplier, 2)
            calc_log += " | Phát hiện chi tiết đi biên gập đôi (FOLD)."
            
    return {
        "layer_multiplier": layer_multiplier,
        "is_paired": is_paired,
        "calc_log": calc_log
    }

def process_pieces_layer_and_areas(bom_rows_list, product_segmented, warp_shrinkage, weft_shrinkage):
    """Khối 3 hoàn chỉnh ổn định thế mới: Bóc tách lớp cắt tự động từ PDF Callout.
    ĐÃ SỬA LỖI GỐC: Ép ghi đè kích thước ĐÃ CỘNG CO RÚT vào tất cả các cấu trúc dữ liệu chuyển giao.
    """
    total_fabric_piece_area = 0.0
    piece_calculated_data = []
    raw_pdf_context = st.session_state.get("raw_pdf_text_extracted", "")

    for r in bom_rows_list:
        if not r or not isinstance(r, dict): continue
        raw_l = safe_float(r.get("bounding_box_length", r.get("Dài (L-inch)", 0.0)))
        raw_w = safe_float(r.get("bounding_box_width", r.get("Rộng (W-inch)", 0.0)))
        
        pcs = safe_int(r.get("original_piece_count", r.get("piece_count", r.get("Số lượng rập", 1))))
        if "original_piece_count" not in r:
            r["original_piece_count"] = pcs
        
        mat_class_raw = str(r.get("material_class", r.get("Material Class", "FABRIC"))).upper().strip()
        comp_name_raw = str(r.get("component_name", "UNNAMED")).upper().strip()
        geo_role_raw = str(r.get("geometry_role", "MINOR_COMPONENT")).upper().strip()
        piece_type_ai = str(r.get("piece_type", geo_role_raw)).upper().strip()
        
        combined_str_item = f" {comp_name_raw} {piece_type_ai} ".lower().replace("_", " ")
        is_button = any(k in combined_str_item for k in ["button", "nút", "nut", "khuy"])

        if raw_l > 0:
            # 🔴 ĐOẠN TÍNH CO RÚT QUAN TRỌNG:
            adj_l = raw_l * (1 + safe_float(warp_shrinkage) / 100.0)
            adj_w = raw_w * (1 + safe_float(weft_shrinkage) / 100.0) if raw_w > 0 else raw_w
            
            # Ép lưu vết trực tiếp vào dictionary dữ liệu gốc r để Khối 5 bốc ra hiển thị
            r["production_length"] = adj_l
            r["production_width"] = adj_w
            
            pdf_callout = extract_cutting_instructions_from_pdf(comp_name_raw, raw_pdf_context)
            layer_multiplier = safe_int(pdf_callout.get("layer_multiplier", 1), default=1)
            calc_chain_log = pdf_callout.get("calc_log", "")

            # Thiết lập hệ số hình dạng xén góc rập (Shape Factor)
            is_belt_loop = any(k in combined_str_item for k in ["beltloop", "đỉa", "dia"])
            if any(k in combined_str_item for k in ["panel", "front", "back", "thân", "body", "sleeve", "tay"]):
                shape_factor = 0.92 if "back" in combined_str_item else 0.85
                if "DRESS" in str(product_segmented).upper() and "flare" in combined_str_item: shape_factor = 0.52
                elif "TROUSER" in str(product_segmented).upper(): shape_factor = 0.63
            elif any(k in combined_str_item for k in ["waistband", "lưng", "collar", "cổ", "belt"]) or is_belt_loop:
                shape_factor = 0.96
            else:
                shape_factor = 0.78

            # Tính toán diện tích thực tế dựa trên kích thước ĐÃ CỘNG CO RÚT
            seamed_l = adj_l + 0.88
            seamed_w = adj_w + 0.88 if raw_w > 0 else adj_w
            
            total_pcs_final = pcs * layer_multiplier
            item_area = seamed_l * seamed_w * shape_factor * total_pcs_final
            
            if "FABRIC" in mat_class_raw: 
                total_fabric_piece_area += item_area
            
            r["piece_count"] = total_pcs_final
            r["Số lượng rập"] = total_pcs_final
            r["polygon_net_area"] = round(seamed_l * seamed_w * shape_factor, 2)
            r["calculation_status"] = "PROCESSED"
            r["cad_algorithm"] = calc_chain_log
            
            # 🔴 ĐỒNG BỘ SANG MẢNG TRUNG GIAN KHỐI 5: Thay raw_l bằng adj_l, raw_w bằng adj_w
            piece_calculated_data.append({
                "row_ref": r, "item_area": item_area, "is_button": is_button, "pcs_display": f"{total_pcs_final} Pcs",
                "layer_multiplier": layer_multiplier, "mat_class_raw": mat_class_raw, "combined_str": combined_str_item, 
                "is_belt_loop": is_belt_loop, 
                "raw_l": adj_l,  # Đưa kích thước đã co rút làm kích thước hiển thị chính
                "raw_w": adj_w,  # Đưa kích thước đã co rút làm kích thước hiển thị chính
                "pcs_val": pcs, "custom_name": comp_name_raw
            })
            
        elif is_button:
            r["production_length"] = 0.0
            r["production_width"] = 0.0
            r["calculation_status"] = "PROCESSED"
            piece_calculated_data.append({
                "row_ref": r, "item_area": 0.0, "is_button": True, "pcs_display": f"{pcs} Pcs",
                "layer_multiplier": 1, "mat_class_raw": mat_class_raw, "combined_str": combined_str_item,
                "is_belt_loop": False, "raw_l": 0.0, "raw_w": 0.0, "pcs_val": pcs, "custom_name": comp_name_raw
            })
    
    # Lưu mảng đã đồng bộ co rút ngược vào session_state phòng khi re-run Streamlit
    st.session_state["piece_calculated_data"] = piece_calculated_data
    return round(total_fabric_piece_area, 4), piece_calculated_data



def allocate_gerber_share_consumption(piece_calculated_data, total_fabric_piece_area, skyline_results):
    """Khối 4 hoàn chỉnh nâng cấp: Phân bổ định mức Gerber chuẩn xác.
    Đã vá lỗi lệch cấu trúc dict, bẫy object tham chiếu và đồng bộ cache Skyline gốc.
    """
    base_gross_fabric = skyline_results.get("global_gross_fabric_yds", 0.0)
    if base_gross_fabric == 0.0:
        base_gross_fabric = skyline_results.get("global_gross_fabric_consumption", 0.0)
    if base_gross_fabric == 0.0:
        base_gross_fabric = skyline_results.get("global_gross_fabric", 0.0)
        
    product_segmented = skyline_results.get("product_segmented", "CASUAL_TOP")
    fabric_pattern_raw = skyline_results.get("fabric_pattern", "SOLID")
    actual_packing_density = skyline_results.get("actual_packing_density", 0.85)
    if actual_packing_density <= 0: actual_packing_density = 0.85
    
    bom_source = st.session_state.get("bom_data", {})
    usable_width = bom_source.get("fabric_width_inch", 56.0)
    if not isinstance(usable_width, (int, float)) or usable_width <= 0: usable_width = 56.0
    
    layout_mapping = {"SOLID": "SOLID LAYOUT", "STRIPE": "STRIPE LAYOUT", "PLAID": "PLAID LAYOUT", "NAP": "NAP LAYOUT (CẮT 1 CHIỀU)"}
    current_layout_text = layout_mapping.get(fabric_pattern_raw, f"{fabric_pattern_raw} LAYOUT")
    
    processed_rows = []

    for item in piece_calculated_data:
        if "row_ref" not in item: continue
        r = item["row_ref"]
        item_area = item["item_area"]
        is_button = item["is_button"]
        layer_multiplier = item["layer_multiplier"]
        
        mat_class_raw = str(item["mat_class_raw"]).upper().strip()
        combined_str_curr = item["combined_str"]
        
        # Đọc trực tiếp kích thước sản xuất đã cộng co rút từ r
        raw_l = r.get("production_length", item.get("raw_l", 0.0))
        raw_w = r.get("production_width", item.get("raw_w", 0.0))
        
        pcs = item["pcs_val"]
        custom_name = item["custom_name"]

        display_name = custom_name if custom_name else str(r.get("component_name", "UNNAMED")).upper().strip()
        status_raw = str(r.get("calculation_status", "READY")).upper().strip()
        confidence = str(r.get("data_confidence", "HIGH")).upper().strip()

        if is_button:
            total_button_units = pcs * layer_multiplier
            gross_consumption = round((total_button_units * 1.03), 2)
            calc_chain = f"Đếm chiếc phụ liệu mẫu hàng {product_segmented}: {total_button_units} cái"
            pcs_display = f"{total_button_units} Cái"
        else:
            pcs_display = item.get("pcs_display", f"{pcs * layer_multiplier} Pcs")
            is_roll_trim = any(k in combined_str_curr for k in ["elastic", "thun", "zipper", "khóa", "hanger", "loop", "label"])

            if is_roll_trim and "FABRIC" not in mat_class_raw:
                gross_consumption = round(((raw_l * pcs * layer_multiplier) / 36.0 * 1.04), 4)
                calc_chain = f"Dải cuộn phụ liệu ({product_segmented}): L-inch / 36.0 + 4% hao hụt"
            else:
                if "FABRIC" in mat_class_raw:
                    geo_role = str(r.get('geometry_role', '')).upper()
                    piece_type = str(r.get('piece_type', '')).upper()
                    is_major = ("MAJOR" in geo_role or "MAJOR" in piece_type) or \
                               any(k in combined_str_curr for k in ["front", "back", "body", "thân", "sleeve", "tay", "panel", "leg"])
                    
                    if total_fabric_piece_area > 0 and base_gross_fabric > 0:
                        share_ratio = item_area / total_fabric_piece_area
                        gross_consumption = round(base_gross_fabric * share_ratio, 4)
                        calc_chain = f"Gerber Major Panel: Gánh nền sơ đồ ({base_gross_fabric:.3f} yds)" if is_major else f"Gerber Minor Component: Phân bổ diện tích phụ ({base_gross_fabric:.3f} yds)"
                    else:
                        estimated_base = ((item_area / usable_width) / 36.0) / actual_packing_density
                        gross_consumption = round(estimated_base * 1.045, 4)
                        calc_chain = f"CAD Geometry Fallback: Giả lập hình học phẳng ({gross_consumption:.3f} yds)"
                            
                elif mat_class_raw in ["FUSING", "LINING"]:
                    gross_consumption = round(((item_area / usable_width) / 36.0 / 0.82), 4)
                    calc_chain = f"Sơ đồ {mat_class_raw} độc lập"
                else:
                    gross_consumption, calc_chain = 0.0, f"Vật tư dòng {product_segmented}."

        # Chặn đứng rủi ro bất đồng bộ con trỏ ô nhớ bằng cách ghi đồng thời vào r và lớp bọc ngoài
        r["Gross Consumption"] = gross_consumption
        item["row_ref"]["Gross Consumption"] = gross_consumption
        r["Số lượng rập"] = pcs_display
        
        # Đẩy dữ liệu dòng phẳng đã làm giàu vào mảng kết quả cuối
        processed_rows.append(r)

    # Lưu cứng định mức tổng và mật độ nén vào bom_data làm chân lý cho Khối sau
    ctx = st.session_state.get("bom_data", {})
    ctx["global_gross_fabric_yds"] = base_gross_fabric
    ctx["actual_packing_density"] = actual_packing_density
    st.session_state["bom_data"] = ctx

    # Lưu mảng phẳng sạch, loại bỏ việc lưu lồng cấu trúc dict phức tạp
    st.session_state["processed_display_rows"] = processed_rows
    return processed_rows


import io
import re
import numpy as np
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# =====================================================================
# 🟩 ĐOẠN 1: BỐC TÁCH THAM SỐ VÀ ĐỒNG BỘ CHÍNH XÁC SIZE MẪU (ĐÃ FIX KHỔ VẢI 58)
# =====================================================================
chat_input_text = str(st.session_state.get("last_submitted_query", "")).lower()

def extract_param(pattern, text, session_key, default_val):
    match = re.search(pattern, text)
    if match:
        val = float(match.group(2) if len(match.groups()) >= 2 else match.group(1))
        st.session_state[session_key] = val
        return val
    return float(st.session_state.get(session_key, default_val))

# 1. Bốc tách tỷ lệ co rút vải
warp_shrink = extract_param(r'(co rút dọc|dọc)\s*(-?\d+\.?\d*)', chat_input_text, "warp_shrinkage", 0.0)
weft_shrink = extract_param(r'(co rút ngang|ngang)\s*(-?\d+\.?\d*)', chat_input_text, "weft_shrinkage", 0.0)

ctx = st.session_state.get("bom_data", {})
if not isinstance(ctx, dict): 
    ctx = {}

# 2. Sửa lỗi luồng bốc size mẫu (Ưu tiên bộ nhớ hệ thống trước)
detected_size_code = ""
if ctx.get("detected_base_size"):
    detected_size_code = str(ctx.get("detected_base_size")).upper().strip()
elif ctx.get("base_size"):
    detected_size_code = str(ctx.get("base_size")).upper().strip()
else:
    size_match = re.search(r'size\s*(\d+x\d+)', chat_input_text)
    if size_match:
        detected_size_code = size_match.group(1).upper()
    else:
        detected_size_code = "32X33" 

ctx["detected_base_size"] = detected_size_code
st.session_state["detected_base_size"] = detected_size_code

# 🚨 3. ĐỒNG BỘ KHỔ VẢI YÊU CẦU: Ép cứng mặc định về 58.0 inch chuẩn thương mại
fabric_width = extract_param(r'(?:khổ\s*vải|vải\s*khổ|khổ)\s*[:=-]?\s*(\d+(?:\.\d+)?)', chat_input_text, "fabric_width_inch", 58.0) 
if fabric_width <= 0 or fabric_width == 55.0: fabric_width = 58.0
ctx["fabric_width_inch"] = fabric_width

# Khổ keo và khổ lót
fusing_width = extract_param(r'(?:khổ\s*keo|keo\s*khổ|dựng\s*khổ|khổ\s*dựng)\s*[:=-]?\s*(\d+(?:\.\d+)?)', chat_input_text, "fusing_width_inch", 59.0)
if fusing_width <= 0: fusing_width = 59.0
ctx["fusing_width_inch"] = fusing_width

lining_width = extract_param(r'(?:khổ\s*lót|lót\s*khổ|vải\s*lót\s*khổ)\s*[:=-]?\s*(\d+(?:\.\d+)?)', chat_input_text, "lining_width_inch", 57.0)
if lining_width <= 0: lining_width = 57.0
ctx["lining_width_inch"] = lining_width

# =====================================================================
# 🟩 ĐOẠN 2: CHUẨN HÓA DỮ LIỆU ĐẦU VÀO VÀ ĐỒNG BỘ SỐ LƯỢNG RẬP CHI TIẾT
# =====================================================================
rows = ctx.get("bom_rows", [])
if not rows:
    rows = st.session_state.get("processed_display_rows", [])

if rows is not None and (isinstance(rows, list) and len(rows) > 0 or isinstance(rows, pd.DataFrame) and not rows.empty):
    df_bom = pd.DataFrame(rows) if isinstance(rows, list) else rows.copy()
    df_bom = df_bom.loc[:, ~df_bom.columns.duplicated()].copy()
    
    prod = str(ctx.get("detected_product_type", ctx.get("product_segmented", "JACKET"))).upper().strip()
    fabric_pattern_raw = str(ctx.get("fabric_pattern", "SOLID")).upper()
    
    m_col = next((c for c in ["Material Class", "material_class"] if c in df_bom.columns), "material_class")
    pcs_col = next((c for c in ["Số lượng rập", "piece_count"] if c in df_bom.columns), "piece_count")
    orig_l_col = next((c for c in ["bounding_box_length", "Dài (L-inch)"] if c in df_bom.columns), "bounding_box_length")
    orig_w_col = next((c for c in ["bounding_box_width", "Rộng (W-inch)"] if c in df_bom.columns), "bounding_box_width")
    
    df_bom[orig_l_col] = pd.to_numeric(df_bom[orig_l_col], errors='coerce').fillna(0.0)
    df_bom[orig_w_col] = pd.to_numeric(df_bom[orig_w_col], errors='coerce').fillna(0.0)
    
    # Trích xuất giữ lại cột số liệu gốc sạch trước khi giải toán hình học
    target_orig_gross_col = next((c for c in ["Gross Consumption", "gross_consumption", "allocated_gross"] if c in df_bom.columns), None)
    if target_orig_gross_col:
        df_bom["original_raw_gross"] = pd.to_numeric(df_bom[target_orig_gross_col], errors='coerce').fillna(0.0)
    else:
        df_bom["original_raw_gross"] = 0.0

    # Khởi tạo bộ đệm lưu trữ chỉnh sửa loại vật liệu của người dùng
    if "user_edited_materials" not in st.session_state:
        st.session_state["user_edited_materials"] = {}
    if "user_edited_pieces" not in st.session_state:
        st.session_state["user_edited_pieces"] = {}

    # Ghi đè loại vật tư nếu người dùng tự sửa tay trên lưới UI
    for idx, row in df_bom.iterrows():
        if idx in st.session_state["user_edited_materials"]:
            df_bom.at[idx, m_col] = st.session_state["user_edited_materials"][idx]

    # THUẬT TOÁN ĐỊNH DANH SỐ LƯỢNG RẬP CHUẨN CÔNG NGHIỆP
    def clean_precise_piece_count(row):
        comp_name = str(row.get("component_name", row.get("Component Name", ""))).upper().strip()
        pcs_raw_str = str(row.get(pcs_col, "1"))
        pcs_extracted = re.search(r'(\d+)', pcs_raw_str)
        pcs_val = float(pcs_extracted.group(1)) if pcs_extracted else 1.0
        
        if any(k in comp_name for k in ["POCKET BAG", "TÚI LÓT", "LÓT TÚI"]):
            return max(pcs_val, 4.0)
        return pcs_val

    df_bom["pcs_numeric"] = [
        float(st.session_state["user_edited_pieces"][idx]) if idx in st.session_state["user_edited_pieces"]
        else clean_precise_piece_count(row) for idx, row in df_bom.iterrows()
    ]
    df_bom[pcs_col] = df_bom["pcs_numeric"]

       # =====================================================================
        # =====================================================================
       # =====================================================================
    # =====================================================================
    # 🟩 ĐOẠN 3: STATISTICAL PRIOR ENGINE & PRODUCTION FEATURING (TÍCH HỢP MA TRẬN ÉP KEO)
    # =====================================================================
    import numpy as np
    import pandas as pd

    # 🧠 1. MA TRẬN TRI THỨC ÉP KEO CHUẨN CÔNG TY (KNOWLEDGE BASE)
    FUSING_STRICT_RULES = {
        "SHIRT": ["COLLAR", "STAND", "FRONT PLACKET", "UNDER PLACKET", "CUFF", "SLEEVE PLACKET", "FLAP"],
        "TOPS_KNIT": ["POLO PLACKET", "PLACKET"], 
        "JEAN_LONG": ["WAISTBAND", "FACING", "FLY", "SHIELD", "ZIP", "POCKET FACING", "COIN", "FLAP"],
        "SHORT": ["WAISTBAND", "FLY", "FACING", "POCKET FACING"],
        "SKIRT": ["WAISTBAND", "WAIST FACING", "ZIP FACING"],
        "DRESS_FLARE": ["WAISTBAND", "NECK FACING", "ARMHOLE", "PLACKET", "ZIP FACING"],
        "JACKET": ["COLLAR", "STAND", "LAPEL", "FRONT FACING", "FRONT PANEL", "POCKET FACING", "FLAP", "WELT", "CUFF", "TAB"],
        "VEST": ["FRONT PANEL", "LAPEL", "COLLAR", "STAND", "FRONT FACING", "POCKET FACING", "FLAP", "WELT", "CUFF"]
    }

    # Barem mật độ cơ sở của công ty đóng vai trò là "Prior" (Khoảng kỳ vọng ban đầu)
    COMPANY_DENSITY_PRIOR = {
        "SHIRT": 0.83, "JEAN_LONG": 0.875, "SHORT": 0.88, 
        "JACKET": 0.68, "VEST": 0.87, "TOPS_KNIT": 0.80, 
        "SKIRT": 0.87, "DRESS_FLARE": 0.74
    }

    # Định vị các cột dữ liệu hệ thống
    comp_col_check = next((c for c in ["Component Name", "component_name", "Component_Name"] if c in df_bom.columns), "component_name")
    l_prod_col_check = "Dài sản xuất (L-inch)" if "Dài sản xuất (L-inch)" in df_bom.columns else (orig_l_col if 'orig_l_col' in locals() else "")
    w_prod_col_check = "Rộng sản xuất (W-inch)" if "Rộng sản xuất (W-inch)" in df_bom.columns else (orig_w_col if 'orig_w_col' in locals() else "")
    
    # Đọc cấu hình sản xuất từ Streamlit UI
    fabric_width = float(st.session_state.get("fabric_width_inch", 58.0))
    rotation_freedom = st.session_state.get("allow_rotation_90", True)      # Cho phép xoay 90°
    one_way_flag = st.session_state.get("is_one_way_fabric", False)          # Vải tuyết / Vải 1 chiều
    stripe_plaid_flag = st.session_state.get("is_stripe_plaid", False)       # Vải canh sọc / caro
    fabric_type = st.session_state.get("fabric_material_type", "WOVEN")       # WOVEN, KNIT...

    # 🧠 BƯỚC QUÉT TRƯỚC DANH MỤC SẢN PHẨM ĐỂ PHỤC VỤ SÀNG LỌC ÉP KEO HÌNH HỌC
    prod_upper_name = str(prod).upper().strip()
    product_category = "JEAN_LONG"
    for k in COMPANY_DENSITY_PRIOR.keys():
        if k in prod_upper_name or (k == "DRESS_FLARE" and any(d in prod_upper_name for d in ["DRESS", "FLARE", "ĐẦM", "XÒE"])):
            product_category = k
            break

    # Khởi tạo mảng thu thập đặc trưng hình học cấp chi tiết (Piece-Level Features)
    piece_areas = []
    piece_aspect_ratios = []
    piece_void_ratios = []
    piece_convex_hull_ratios = [] 
    
    total_pattern_pieces = 0.0
    total_pocket_pieces = 0.0
    max_piece_length = 0.0
    symmetry_pieces_count = 0.0

    for idx, r in df_bom.iterrows():
        mat_class_clean = str(r[m_col]).upper().strip()
        comp_name_clean = str(r.get(comp_col_check, "")).upper().strip()
        
        # Nhận diện cấu trúc phức tạp cụm túi (Welt Pocket có trọng số khó gấp đôi Patch Pocket)
        if any(k in comp_name_clean for k in ["POCKET", "TÚI", "WELT"]):
            weight = 2.0 if "WELT" in comp_name_clean else 1.0
            total_pocket_pieces += float(r["pcs_numeric"]) * weight

        # Khử nhiễu bộ đặc trưng: Loại trừ chỉ may, phụ liệu ảo và CÁC CHI TIẾT TỰ ĐỘNG ÉP KEO THEO MA TRẠN ĐỘNG
        is_fusing_by_rule = False
        if product_category in FUSING_STRICT_RULES and any(k in comp_name_clean for k in FUSING_STRICT_RULES[product_category]):
            is_fusing_by_rule = True

        # Bộ đặc trưng này chỉ tính toán diện tích thớ rập của VẢI CHÍNH để dự báo mật độ (Fabric Density Prior)
        if not any(k in mat_class_clean for k in ["THREAD", "CHI", "ACCESSORY", "PHU LIEU", "BUTTON", "ZIPPER", "LABEL", "FUSING", "MEC", "KEO"]) and not is_fusing_by_rule:
            current_pcs = float(st.session_state.get("user_edited_pieces", {}).get(idx, r["pcs_numeric"]))
            total_pattern_pieces += current_pcs
            
            net_area = float(r.get("polygon_net_area", 0.0))
            l_val = float(r.get(l_prod_col_check, 0.0))
            w_val = float(r.get(w_prod_col_check, 0.0))
            bbox_area = l_val * w_val
            
            if current_pcs >= 2:
                symmetry_pieces_count += current_pcs

            if l_val > max_piece_length:
                max_piece_length = l_val

            if net_area > 0 and bbox_area > 0:
                raw_ratio = l_val / w_val if w_val > 0 else 1.0
                best_ratio = min(raw_ratio, 1.0 / raw_ratio) if rotation_freedom else raw_ratio
                sim_convex_ratio = min(1.0, round(net_area / (bbox_area * 0.95), 4))

                for _ in range(int(current_pcs)):
                    piece_areas.append(net_area)
                    piece_aspect_ratios.append(best_ratio)
                    piece_void_ratios.append((bbox_area - net_area) / bbox_area)
                    piece_convex_hull_ratios.append(sim_convex_ratio)

    # Tổng hợp ma trận 15 đặc trưng cốt lõi nuôi AI
    features = {}
    if len(piece_areas) > 0:
        features["total_pieces"] = float(total_pattern_pieces)
        features["largest_piece_area"] = float(max(piece_areas))
        features["mean_piece_area"] = float(np.mean(piece_areas))
        features["std_piece_area"] = float(np.std(piece_areas))
        features["avg_aspect_ratio"] = float(np.mean(piece_aspect_ratios))
        features["max_aspect_ratio"] = float(max(piece_aspect_ratios))
        features["avg_void_ratio"] = float(np.mean(piece_void_ratios))
        features["convex_hull_ratio"] = float(np.mean(piece_convex_hull_ratios))
        features["width_utilization"] = float(max_piece_length / fabric_width)
        features["rotation_freedom"] = 1.0 if rotation_freedom else 0.0
        features["symmetry_ratio"] = float(symmetry_pieces_count / total_pattern_pieces)
        features["fabric_width"] = float(fabric_width)
        features["one_way_flag"] = 1.0 if one_way_flag else 0.0
        features["stripe_plaid_flag"] = 1.0 if stripe_plaid_flag else 0.0
        features["pocket_complexity"] = float(total_pocket_pieces)
    else:
        features = {k: 0.0 for k in ["total_pieces", "largest_piece_area", "mean_piece_area", "std_piece_area", "avg_aspect_ratio", "max_aspect_ratio", "avg_void_ratio", "convex_hull_ratio", "width_utilization", "rotation_freedom", "symmetry_ratio", "fabric_width", "one_way_flag", "stripe_plaid_flag", "pocket_complexity"]}

    # ĐỒNG BỘ ĐỒNG LOẠT BIẾN CHUỖI GIAO DIỆN CŨ ĐỂ KHỬ SẠCH LỖI NAMEERROR
    if product_category == "VEST": ai_product_type = "VEST (Áo Vest/Blazer)"
    elif product_category == "JACKET": ai_product_type = "JACKET (Áo khoác Jacket)"
    elif product_category == "DRESS_FLARE": ai_product_type = "DRESS_FLARE (Đầm xòe/Thời trang)"
    elif product_category == "SKIRT": ai_product_type = "SKIRT (Chân váy)"
    elif product_category == "TOPS_KNIT": ai_product_type = "TOPS_KNIT (Áo thun/Polo)"
    elif product_category == "SHIRT": ai_product_type = "SHIRT (Áo sơ mi)"
    elif product_category == "SHORT": ai_product_type = "SHORT (Quần short)"
    else: ai_product_type = "JEAN_LONG (Quần dài Jeans/Pants)"
    
    base_prior = COMPANY_DENSITY_PRIOR[product_category]

    # Hàm hồi quy tinh chỉnh mật độ kỳ vọng ban đầu (Regression Matrix)
    density_delta = (
        (features["total_pieces"] * 0.0004)
        - (features["avg_void_ratio"] * 0.08)
        - (features["one_way_flag"] * 0.035)
        - (features["stripe_plaid_flag"] * 0.05)
        + (features["rotation_freedom"] * 0.02)
        - (features["width_utilization"] * 0.015)
    )
    estimated_density = max(0.50, min(0.94, base_prior + density_delta))

    # Bộ giải toán toán tính điểm phức tạp liên tục (Complexity Score)
    complexity_score = min(100.0, max(1.0, (total_pattern_pieces * 1.2) + (features["avg_void_ratio"] * 80) + (total_pocket_pieces * 2.0)))
    ai_complexity = "COMPLEX" if complexity_score >= 50 else "NORMAL"

    # Bộ giải toán hao hụt sản xuất thực tế (Đầu cây, nối cây, co rút nghỉ vải, canh sọc)
    base_waste = 1.015 
    production_penalties = (
        (features["stripe_plaid_flag"] * 0.025)
        + (features["one_way_flag"] * 0.01)
        + (1.02 if fabric_type in ["KNIT", "THUN"] else 1.0) - 1.0
        + (0.01 if features["pocket_complexity"] > 10 else 0.0)
    )
    calculated_wastage = base_waste + production_penalties

    # ĐỒNG BỘ BIẾN SỐ CHO LAYOUT BÁO CÁO CŨ
    target_density = estimated_density
    target_wastage = calculated_wastage

    # Lưu trữ thông tin an toàn vào hệ thống ctx dưới dạng ƯỚC TÍNH (ESTIMATED)
    if "ai_expert_decision" not in ctx or not isinstance(ctx["ai_expert_decision"], dict):
        ctx["ai_expert_decision"] = {}

    ctx["ai_expert_decision"]["product_category"] = product_category
    ctx["ai_expert_decision"]["assigned_marker_density"] = round(estimated_density, 4) 
    ctx["ai_expert_decision"]["estimated_density_prior"] = round(estimated_density, 4)
    ctx["ai_expert_decision"]["dynamic_wastage_factor"] = round(calculated_wastage, 4)
    ctx["ai_expert_decision"]["wastage_factor"] = round(calculated_wastage, 4) 
    ctx["ai_expert_decision"]["geometry_features"] = features
    ctx["ai_expert_decision"]["longest_piece_length"] = max_piece_length
    
    st.session_state["bom_data"] = ctx

    # Xuất thông tin giải trình trực quan lên giao diện UI chính
    st.subheader("🧠 Hệ Thống Trích Xuất Đặc Trưng Hình Học AI CAD Engine")
    m1, m2, m3 = st.columns(3)
    with m1:


    
        st.metric(label="📊 Ước lượng Mật độ Tiên nghiệm (Regression)", value=f"{estimated_density*100:.2f}%", delta=f"{(estimated_density - base_prior)*100:+.2f}% vs Barem")
    with m2:
        st.metric(label="✂️ Định mức Hao hụt Sản xuất Động", value=f"{((calculated_wastage-1)*100):.2f}%")
    with m3:
        st.metric(label="🧩 Tổng số mảnh rập thực tế", value=f"{features['total_pieces']:.0f} Pcs")


        # =====================================================================
    # 🟩 ĐOẠN 4: VIRTUAL PIECE INFERENCE ENGINE & GEOMETRIC CLASSIFIER
    # =====================================================================
    import numpy as np

    # 🚨 KHÔNG SỬA DF_BOM GỐC - KHỞI TẠO LỚP CHI TIẾT ẢO TRONG CONTEXT
    if "virtual_pieces_layer" not in ctx:
        ctx["virtual_pieces_layer"] = {}
        
    comp_col_check = next((c for c in ["Component Name", "component_name", "Component_Name"] if c in df_bom.columns), "component_name")
    
    # Bộ sniffer dò tìm chính xác cột kích thước hình học từ CAD đầu vào
    detected_l_col = next((c for c in ["Dài gốc (inch)", "orig_l", "original_length", "length_inch", "Length"] if c in df_bom.columns), None)
    detected_w_col = next((c for c in ["Rộng gốc (inch)", "orig_w", "original_width", "width_inch", "Width"] if c in df_bom.columns), None)
    actual_l_col = detected_l_col if detected_l_col else (orig_l_col if 'orig_l_col' in locals() else "Dài gốc (inch)")
    actual_w_col = detected_w_col if detected_w_col else (orig_w_col if 'orig_w_col' in locals() else "Rộng gốc (inch)")

    # Đồng bộ thông số co rút sợi sản xuất phục vụ nắn kích thước phôi ảo
    fusing_warp_shrink = float(st.session_state.get("fusing_warp_shrink", 0.0))
    fusing_weft_shrink = float(st.session_state.get("fusing_weft_shrink", 0.0))
    lining_warp_shrink = float(st.session_state.get("lining_warp_shrink", 0.0))
    lining_weft_shrink = float(st.session_state.get("lining_weft_shrink", 0.0))

    # ⚙️ HÀM PHÂN LOẠI CHẤT LIỆU CHỈ DỰA TRÊN ĐẶC TRƯNG HÌNH HỌC PHẲNG TRƯỚC VĂN BẢN
    def geometric_material_classifier(row, net_area, bbox_area, slenderness, void_ratio):
        mat_str = str(row[m_col]).upper().strip()
        comp_str = str(row.get(comp_col_check, row.get("component_name", ""))).upper().strip()
        
        # Đặc trưng 1: Nếu chi tiết mảnh, dài (slenderness >= 6.0) và rỗng cực ít (void_ratio <= 0.12) -> Khả năng cao là nẹp / dây viền / keo lót rib
        if slenderness >= 5.0 and void_ratio <= 0.15:
            if not any(k in mat_str for k in ["THREAD", "CHI", "ACCESSORY"]):
                return "FUSING", 0.92
                
        # Đặc trưng 2: Nếu chi tiết nhỏ nhưng độ rỗng lớn (void_ratio > 0.3) và nằm gần cụm từ pocket -> Pocket Facing / Lining
        if net_area < 150.0 and any(k in comp_str for k in ["POCKET", "TÚI", "WELT", "BAG"]):
            if "LINING" in mat_str or "LÓT" in comp_str:
                return "LINING", 0.88
            return "FABRIC", 0.75
            
        # Đặc trưng từ khóa cứng trong file gốc
        if any(k in mat_str or k in comp_str for k in ["FUSING", "MEC", "KEO", "INTERLINING"]): return "FUSING", 1.0
        if any(k in mat_str or k in comp_str for k in ["LINING", "LOT", "POCKETING", "VAI LOT"]): return "LINING", 1.0
        if any(k in mat_str or k in comp_str for k in ["THREAD", "CHI", "ACCESSORY", "BUTTON", "ZIPPER"]): return "ACCESSORY", 1.0
        
        # Mặc định là vải chính thân lớn nếu diện tích bao bì đáng kể
        if net_area > 200.0:
            return "FABRIC", 0.95
        return "FABRIC", 0.70

    # 📊 DUYỆT TỪNG DÒNG BOM GỐC ĐỂ PHÂN TÍCH VÀ TRÍCH XUẤT PHÔI ẢO (VIRTUAL PIECE MODEL)
    virtual_pieces_registry = {}

    for idx, row in df_bom.iterrows():
        comp_name_raw = str(row.get(comp_col_check, row.get("component_name", "")))
        comp_name_upper = comp_name_raw.upper().strip()
        
        l_orig = float(row.get(actual_l_col, 0.0))
        w_orig = float(row.get(actual_w_col, 0.0))
        net_area_raw = float(row.get("polygon_net_area", l_orig * w_orig * 0.78))
        bbox_area_raw = l_orig * w_orig
        
        # Trích xuất 4 chỉ số hình học phẳng lõi cấp chi tiết (Piece Features)
        slenderness = l_orig / w_orig if w_orig > 0 else 1.0
        void_ratio = (bbox_area_raw - net_area_raw) / bbox_area_raw if bbox_area_raw > 0 else 0.0
        
        # 1. PIECE CLASSIFICATION: Suy luận vai trò vật tư bằng thuật toán hình học
        p_class, class_confidence = geometric_material_classifier(row, net_area_raw, bbox_area_raw, slenderness, void_ratio)
        
        # Nắn thông số kích thước sản xuất co rút sợi dựa trên nhãn phân loại ảo vừa tìm được
        if p_class == "FABRIC":
            w_prod = round(w_orig * (1 + weft_shrink / 100.0), 3)
            l_prod = round(l_orig * (1 + warp_shrink / 100.0), 3)
        elif p_class == "FUSING":
            w_prod = round(w_orig * (1 + fusing_weft_shrink / 100.0), 3)
            l_prod = round(l_orig * (1 + fusing_warp_shrink / 100.0), 3)
        elif p_class == "LINING":
            w_prod = round(w_orig * (1 + lining_weft_shrink / 100.0), 3)
            l_prod = round(l_orig * (1 + lining_warp_shrink / 100.0), 3)
        else:
            w_prod, l_prod = w_orig, l_orig

        bbox_area_prod = l_prod * w_prod
        # Diện tích đa giác thực tế tỷ lệ thuận phình to theo hệ số co rút sản xuất
        net_area_prod = round(net_area_raw * (bbox_area_prod / bbox_area_raw) if bbox_area_raw > 0 else net_area_raw, 2)

        # 2. PIECE QUANTITY INFERENCE: Suy luận số lượng mảnh thực nghiệm rải bàn cắt
        raw_pcs = float(row.get("pcs_numeric", 1.0))
        inferred_pcs = raw_pcs
        qty_confidence = 1.0
        
        # Phát hiện đặc trưng rập mở phẳng siêu to (Full-panel Detection)
        # Nếu thớ rộng đơn lẻ w_prod vượt quá 18 inch, khả năng cao rập đã đối xứng chẵn trục trên CAD
        is_full_panel = (w_prod > 18.0) or (net_area_prod > 450.0 and (l_prod / w_prod) < 1.8)
        
        # Phân tích định hướng cấu trúc vòng thân qua từ khóa metadata của rập
        has_front = any(k in comp_name_upper for k in ["FRONT", "TRƯỚC"])
        has_back = any(k in comp_name_upper for k in ["BACK", "SAU"])
        has_split = any(k in comp_name_upper for k in ["SPLIT", "RÃ", "MỔ", "CENTER", "SỐNG"])
        
        if p_class in ["FABRIC", "LINING"] and net_area_prod > 150.0:
            if has_front and is_full_panel:
                # Thân trước mở phẳng, nếu file gốc ghi bằng 2 thì có dấu hiệu lỗi trùng dòng
                inferred_pcs = 1.0
                qty_confidence = 0.94
            elif has_back and not is_full_panel:
                # Thân sau mảnh gầy đơn lẻ xẻ sống tra khóa -> Cần 2 mảnh đối xứng Trái/Phải
                inferred_pcs = 2.0
                qty_confidence = 0.89
            elif not has_front and not has_back:
                # 🚨 THỚ GỘP VÔ ĐỊNH DANH (Ví dụ tên chỉ ghi vỏn vẹn chữ BODY / MAIN CHUNG CHUNG)
                # Tính toán điểm tin cậy (Confidence). Nếu rập to phẳng béo, khả năng cao là rập gộp thân trước mở phẳng
                if is_wide_piece := (w_prod > 20.0):
                    inferred_pcs = raw_pcs # Giữ nguyên cấu trúc rập của file CAD
                    qty_confidence = 0.55  # Đánh dấu độ tin cậy THẤP, không tự động bẻ dòng lung tung
                else:
                    # Rập nhỏ gầy, không hướng tính -> Giữ nguyên cấu trúc CAD gốc
                    inferred_pcs = raw_pcs
                    qty_confidence = 0.60

        # Ép tay áo (Sleeve) về cặp 2 mảnh độc lập nếu file gốc ghi thiếu
        if any(k in comp_name_upper for k in ["SLEEVE", "TAY", "SIEEVE"]):
            if raw_pcs == 1.0:
                inferred_pcs = 2.0
                qty_confidence = 0.98

        # 3. ĐÓNG GÓI ĐỐI TƯỢNG MẢNH ẢO VÀO RECOGNITION REGISTRY (BOM GỐC GIỮ NGUYÊN)
        virtual_pieces_registry[idx] = {
            "original_index": idx,
            "component_name": comp_name_raw,
            "inferred_class": p_class,
            "class_confidence": class_confidence,
            "qty_confidence": qty_confidence,
            "production_l": l_prod,
            "production_w": w_prod,
            "production_net_area": net_area_prod,
            "inferred_pieces": float(st.session_state.get("user_edited_pieces", {}).get(idx, inferred_pcs))
        }

    # Đẩy lớp chi tiết ảo xuống context bộ nhớ để bàn giao cho Đoạn 5.1 và 5.2 xử lý
    ctx["ai_expert_decision"]["virtual_pieces_layer"] = virtual_pieces_registry
    st.session_state["bom_data"] = ctx

    # Thiết lập các cột sản xuất đồng bộ tạm thời cho df_bom phục vụ lưới Grid Đoạn 7 không bị gãy thớ
    df_bom["Dài sản xuất (L-inch)"] = [virtual_pieces_registry[idx]["production_l"] for idx in df_bom.index]
    df_bom["Rộng sản xuất (W-inch)"] = [virtual_pieces_registry[idx]["production_w"] for idx in df_bom.index]
    df_bom["polygon_net_area"] = [virtual_pieces_registry[idx]["production_net_area"] for idx in df_bom.index]

    # Bộ định tuyến phương pháp giải toán (Solver Router Engine)
    def rule_engine_coordinator(row):
        comp_name = str(row.get(comp_col_check, row.get("component_name", ""))).upper().strip()
        if any(k in comp_name for k in ["ELASTIC", "DRAWCORD", "WEBBING", "CHUN", "DÂY LUỒN"]): return "LengthSolver"
        if any(k in comp_name for k in ["BELT_LOOP", "LOOP", "ĐỈA", "STRIP", "BINDING", "VIỀN"]): return "StripSolver"
        return "AreaSolver"

    df_bom["assigned_solver"] = df_bom.apply(rule_engine_coordinator, axis=1)





       # =====================================================================
    # 🟩 ĐOẠN 5.1: GEOMETRIC MARKER ENGINE (MÔ PHỎNG XẾP SƠ ĐỒ HÌNH HỌC)
    # =====================================================================
    ai_decision = ctx.get("ai_expert_decision", {})
    if not isinstance(ai_decision, dict): 
        ai_decision = {}
        
    estimated_density_prior = float(ai_decision.get("estimated_density_prior", 0.78))
    target_wastage = float(ai_decision.get("dynamic_wastage_factor", 1.03))
    features = ai_decision.get("geometry_features", {})
    max_piece_length = float(ai_decision.get("longest_piece_length", 0.0))
    prod_cat_d5 = str(ai_decision.get("product_category", "JEAN_LONG")).upper().strip()

    l_prod_col = "Dài sản xuất (L-inch)" if "Dài sản xuất (L-inch)" in df_bom.columns else orig_l_col
    w_prod_col = "Rộng sản xuất (W-inch)" if "Rộng sản xuất (W-inch)" in df_bom.columns else orig_w_col
    
    current_fabric_width = float(st.session_state.get("fabric_width_inch", 58.0))
    lining_width = float(st.session_state.get("lining_width_inch", 57.0))    
    fusing_width = float(st.session_state.get("fusing_width_inch", 59.0))    

    # Hàm phân loại chất liệu đồng bộ tuyệt đối ma trận ép keo Đoạn 3
    def local_strict_classify(row, idx):
        if "user_edited_materials" in st.session_state and idx in st.session_state["user_edited_materials"]:
            return str(st.session_state["user_edited_materials"][idx]).upper().strip()
        mat_str = str(row[m_col]).upper().strip()
        comp_str = str(row.get("Component Name", row.get("component_name", ""))).upper().strip()
        role_str = str(row.get("Role/Piece Type", row.get("geometry_role", ""))).upper().strip()
        
        fusing_keywords = ["FUSING", "INTERLINING", "KEO", "MEC", "RIB", "BOND", "TAPE", "ADHESIVE", "COLLAR", "CUFF", "WAISTBAND", "LOT KEO", "TAPE"]
        lining_keywords = ["LINING", "LOT", "POCKETING", "MESH", "TAFFETA", "VAI LOT"]
        
        if any(k in mat_str or k in comp_str for k in fusing_keywords): return "FUSING"
        if any(k in mat_str or k in comp_str or k in role_str for k in lining_keywords): return "LINING"
        if any(k in mat_str or k in comp_str for k in ["ACCESSORY", "THREAD", "CHI", "BUTTON", "ZIPPER", "RIVET"]): return "ACCESSORY"
        
        if 'FUSING_STRICT_RULES' in globals() and prod_cat_d5 in FUSING_STRICT_RULES:
            if any(k in comp_str for k in FUSING_STRICT_RULES[prod_cat_d5]):
                if not any(x in mat_str for x in ["THREAD", "CHI", "ACCESSORY", "BUTTON", "ZIPPER"]):
                    return "FUSING"
        return "FABRIC"

    # LUỒNG VẢI CHÍNH (FABRIC)
    total_fabric_net_area = 0.0
    fabric_pieces_to_nest = []

    for idx, r in df_bom.iterrows():
        if local_strict_classify(r, idx) == "FABRIC":
            current_pcs = float(st.session_state.get("user_edited_pieces", {}).get(idx, r["pcs_numeric"]))
            net_area = float(r["polygon_net_area"])
            l_val = float(r[l_prod_col])
            w_val = float(r[w_prod_col])
            total_fabric_net_area += net_area * current_pcs
            for _ in range(int(current_pcs)):
                fabric_pieces_to_nest.append({"l": l_val, "w": w_val, "area": net_area})

    if len(fabric_pieces_to_nest) > 0 and current_fabric_width > 0:
        fabric_pieces_to_nest.sort(key=lambda x: x["area"], reverse=True)
        simulated_marker_length = max_piece_length 
        accumulated_width_used = 0.0
        for piece in fabric_pieces_to_nest:
            p_len = min(piece["l"], piece["w"]) if features.get("rotation_freedom", 1.0) == 1.0 else piece["l"]
            p_wid = max(piece["l"], piece["w"]) if features.get("rotation_freedom", 1.0) == 1.0 else piece["w"]
            if accumulated_width_used + p_wid <= current_fabric_width:
                accumulated_width_used += p_wid
                if p_len > simulated_marker_length: simulated_marker_length = p_len
            else:
                simulated_marker_length += p_len
                accumulated_width_used = p_wid

        total_marker_bounding_area = simulated_marker_length * current_fabric_width
        real_fabric_density = total_fabric_net_area / total_marker_bounding_area if total_marker_bounding_area > 0 else estimated_density_prior
        real_fabric_density = max(estimated_density_prior - 0.05, min(estimated_density_prior + 0.04, real_fabric_density))
        fabric_sim_length = total_fabric_net_area / current_fabric_width / real_fabric_density
        total_fabric_gross_yds = (fabric_sim_length / 36.0) * target_wastage
    else:
        real_fabric_density = estimated_density_prior
        total_fabric_gross_yds = 1.18 if len(fabric_pieces_to_nest) > 0 else 0.0

    # LUỒNG VẢI LÓT (LINING)
    total_lining_net_area = 0.0
    for idx, r in df_bom.iterrows():
        current_pcs = float(st.session_state.get("user_edited_pieces", {}).get(idx, r["pcs_numeric"]))
        if local_strict_classify(r, idx) == "LINING":
            total_lining_net_area += float(r["polygon_net_area"]) * current_pcs

    if total_lining_net_area > 0 and lining_width > 0:
        lining_sim_length = total_lining_net_area / lining_width / 0.76
        total_lining_gross_yds = (lining_sim_length / 36.0) * target_wastage
    else:
        total_lining_gross_yds = 0.0

    ctx["ai_expert_decision"]["real_fabric_density"] = round(real_fabric_density, 4)
    ctx["ai_expert_decision"]["total_fabric_gross_yds"] = round(total_fabric_gross_yds, 4)
    ctx["ai_expert_decision"]["total_lining_gross_yds"] = round(total_lining_gross_yds, 4)
    # =====================================================================
       # =====================================================================
    # 🟩 ĐOẠN 5.2: CONSUMPTION ROUTER & PUBLISHING (PHÂN BỔ ĐỊNH MỨC AN TOÀN)
    # =====================================================================
    ai_decision = ctx.get("ai_expert_decision", {})
    real_fabric_density = float(ai_decision.get("real_fabric_density", 0.78))
    total_fabric_gross_yds = float(ai_decision.get("total_fabric_gross_yds", 0.0))
    total_lining_gross_yds = float(ai_decision.get("total_lining_gross_yds", 0.0))
    target_wastage = float(ai_decision.get("dynamic_wastage_factor", 1.03))

    current_fabric_width = float(st.session_state.get("fabric_width_inch", 58.0))
    lining_width = float(st.session_state.get("lining_width_inch", 57.0))    
    fusing_width = float(st.session_state.get("fusing_width_inch", 59.0))    

    # Bộ giải toán Keo/Méc độc lập bảo vệ biên
    def dynamic_fusing_solver(l_prod, w_prod, net_area, pcs):
        if fusing_width <= 0: return 0.0
        bounding_box_area = l_prod * w_prod
        void_ratio = (bounding_box_area - net_area) / bounding_box_area if bounding_box_area > 0 else 0.0
        slenderness = l_prod / w_prod if w_prod > 0 else 1.0
        if slenderness >= 6.0 and void_ratio <= 0.12:
            fusing_efficiency_calc = 0.65  
            fusing_wastage_calc = 1.08     
            direct_fusing_sim_length = bounding_box_area * pcs / fusing_width / fusing_efficiency_calc
            return (direct_fusing_sim_length / 36.0) * fusing_wastage_calc
        else:
            fusing_efficiency_calc = round(0.72 - (void_ratio * 0.40), 3)
            fusing_wastage_calc = round(1.08 + (void_ratio * 0.25), 3)
            if fusing_efficiency_calc <= 0: fusing_efficiency_calc = 0.5
            direct_fusing_sim_length = (net_area * pcs) / fusing_width / fusing_efficiency_calc
            return (direct_fusing_sim_length / 36.0) * fusing_wastage_calc

    # ⚙️ BỘ ĐIỀU PHỐI VÀ PHÂN BỔ ĐỊNH MỨC AN TOÀN TUYỆT ĐỐI (PROTECTED ROUTER)
    def core_engine_router(row, idx):
        try:
            p_class = local_strict_classify(row, idx) if 'local_strict_classify' in locals() else "FABRIC"
            if p_class == "ACCESSORY": 
                return 0.0
                
            if st.session_state.get("lock_original_techpack", False):
                if "original_raw_gross" in df_bom.columns and float(row.get("original_raw_gross", 0.0)) > 0:
                    return round(float(row["original_raw_gross"]), 4)

            # Đồng bộ an toàn số lượng rập sửa tay
            pcs = float(st.session_state.get("user_edited_pieces", {}).get(idx, row.get("pcs_numeric", 1.0)))
            net_area = float(row.get("polygon_net_area", 0.0))
            l_prod = float(row.get(l_prod_col, 0.0))
            w_prod = float(row.get(w_prod_col, 0.0))
            
            if p_class == "FUSING": 
                return round(dynamic_fusing_solver(l_prod, w_prod, net_area, pcs), 4)
            elif p_class == "FABRIC":
                line_net_area_sum = net_area * pcs
                if total_fabric_net_area > 0: 
                    return round(total_fabric_gross_yds * (line_net_area_sum / total_fabric_net_area), 4)
            elif p_class == "LINING":
                line_net_area_sum = net_area * pcs
                if total_lining_net_area > 0: 
                    return round(total_lining_gross_yds * (line_net_area_sum / total_lining_net_area), 4)
                elif lining_width > 0: 
                    return round((((line_net_area_sum / lining_width) / 36.0 / 0.76) * target_wastage), 4)
        except Exception:
            return 0.0 # Chống sập ngầm dòng rập lỗi, giữ thông suốt hệ thống
        return 0.0

    # 🚨 THỰC THI ÉP BUỘC XUẤT BẢN CỘT DỮ LIỆU CHỐNG LỖI KEYERROR ĐOẠN 7
    gross_list = []
    for idx, row in df_bom.iterrows():
        gross_list.append(core_engine_router(row, idx))
    df_bom["Gross Consumption"] = gross_list
    
    # Đóng gói khổ rộng hiển thị hiển thị UI
    def map_calculated_width(row, idx):
        p_class = local_strict_classify(row, idx) if 'local_strict_classify' in locals() else "FABRIC"
        if p_class == "FABRIC": return current_fabric_width
        if p_class == "LINING": return lining_width
        if p_class == "FUSING": return fusing_width
        return 0.0

    df_bom["Calculated Width (Inch)"] = [map_calculated_width(row, idx) for idx, row in df_bom.iterrows()]
    
    if len(fabric_pieces_to_nest) > 0:
        st.success(f"🧩 **GEOMETRIC SOLVER KẾT QUẢ** | Mật độ thực nghiệm sơ đồ (Real Density): `{real_fabric_density*100:.2f}%` | Định mức tổng vải chính phân bổ: `{total_fabric_gross_yds:.3f} Yds` (Đã đồng bộ kiểm tra va chạm xếp chồng)")
 




          # =====================================================================
    # 🟩 ĐOẠN 6: KHỞI TẠO HÀM XUẤT EXCEL NỘI BỘ (LOCAL EXPORT ENGINE)
    # =====================================================================
    def local_export_excel_ppj_format(df_sum, df_det, product_type, bom_ctx, density):
        output_stream = io.BytesIO()
        workbook = Workbook()
        
        f_family = "Segoe UI"
        f_normal = Font(name=f_family, size=10)
        f_bold = Font(name=f_family, size=10, bold=True)
        f_title = Font(name=f_family, size=14, bold=True, color="0E6251")
        f_header = Font(name=f_family, size=10, bold=True, color="FFFFFF")
        
        fill_header = PatternFill(start_color="0E6251", end_color="0E6251", fill_type="solid")
        fill_meta = PatternFill(start_color="F2F4F4", end_color="F2F4F4", fill_type="solid")
        
        bd_side = Side(style='thin', color='BDC3C7')
        bd_thin = Border(left=bd_side, right=bd_side, top=bd_side, bottom=bd_side)
        
        # --- TAB 1: BOM SUMMARY ---
        w_s1 = workbook.active
        w_s1.title = "BOM Summary"
        w_s1.sheet_view.showGridLines = True
        
        w_s1.cell(row=1, column=1, value="PHÒNG IE / CẮT CAD - HỆ THỐNG QUẢN LÝ PPJ GROUP").font = Font(name=f_family, size=8, italic=True, color="7F8C8D")
        w_s1.cell(row=2, column=1, value="BẢNG ĐỊNH MỨC CHI TIẾT SẢN XUẤT ĐẠI TRÀ").font = f_title
        w_s1.cell(row=4, column=1, value="THÔNG SỐ ĐẦU VÀO SƠ ĐỒ CAD (TECHNICAL PROFILE)").font = Font(name=f_family, size=11, bold=True)
        
        st_code = str(bom_ctx.get("style_code", "N/A")).upper()
        cust_name = str(bom_ctx.get("customer_name", "FACTORY STANDARD")).upper()
        
        m_data = [
            ("Mã hàng / Style Code:", st_code, "Khách hàng / Đối tác:", cust_name),
            ("Size may mẫu (Sample Size):", str(detected_size_code), "Khổ vải hữu dụng (Width):", f'{fabric_width}"'),
            ("Co rút dọc (Warp Shrinkage):", f'{warp_shrink}%', "Co rút ngang (Weft Shrinkage):", f'{weft_shrink}%'),
            ("Chủng loại sản phẩm:", str(product_type).upper(), "Hiệu suất sơ đồ (Density):", f'{density * 100:.1f}%')
        ]
        
        for r_idx, row_data in enumerate(m_data, start=5):
            for c_idx, val in enumerate(row_data, start=1):
                cell = w_s1.cell(row=r_idx, column=c_idx, value=val)
                cell.border = bd_thin
                # SỬA LỖI CÚ PHÁP: Cố định đúng chỉ số mảng cột Tiêu đề
                if c_idx == 1 or c_idx == 3:
                    cell.font = f_bold; cell.fill = fill_meta; cell.alignment = Alignment(horizontal="left", vertical="center")
                else:
                    cell.font = f_normal; cell.alignment = Alignment(horizontal="center", vertical="center")
                    
        w_s1.cell(row=10, column=1, value="BẢNG TỔNG HỢP TIÊU HAO VẬT TƯ (BOM SUMMARY)").font = Font(name=f_family, size=11, bold=True)
        sum_hd = ["Phân loại vật tư", "Mã Vật Liệu Gốc", "Định Mức (Gross Consumption)", "Đơn Vị Tính (UOM)"]
        for c_idx, h_text in enumerate(sum_hd, start=1):
            cell = w_s1.cell(row=11, column=c_idx, value=h_text)
            cell.font = f_header; cell.fill = fill_header; cell.alignment = Alignment(horizontal="center", vertical="center"); cell.border = bd_thin
            
        c_row = 12
        for _, r in df_sum.iterrows():
            w_s1.cell(row=c_row, column=1, value=r.get("Phân loại vật tư", "VẬT TƯ"))
            w_s1.cell(row=c_row, column=2, value=r.get("Material Class", "FABRIC"))
            w_s1.cell(row=c_row, column=3, value=float(r.get("Gross Consumption", 0.0)))
            w_s1.cell(row=c_row, column=4, value=r.get("UOM", "YDS"))
            w_s1.cell(row=c_row, column=3).number_format = '#,##0.0000'
            for c_idx in range(1, 5):
                cell = w_s1.cell(row=c_row, column=c_idx)
                cell.font = f_normal; cell.border = bd_thin
                # SỬA LỖI CÚ PHÁP: Cố định đúng chỉ số mảng cột dữ liệu căn giữa
                if c_idx == 2 or c_idx == 4: 
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            c_row += 1

        for col_idx, col in enumerate(w_s1.columns, start=1):
            m_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col_idx)
            w_s1.column_dimensions[col_letter].width = max(m_len + 3, 12)

        # --- TAB 2: DETAILED CAD PIECES ---
        w_s2 = workbook.create_sheet(title="Detailed CAD Pieces")
        w_s2.sheet_view.showGridLines = True
        w_s2.cell(row=1, column=1, value=f"CHI TIẾT CẤU TRÚC ĐA GIÁC RẬP GERBER ACCUMULATION - DÒNG: {str(product_type).upper()}").font = Font(name=f_family, size=11, bold=True)
        
        det_hd = ["Component Name", "Material Class", "Role/Piece Type", "Khổ vải sản xuất (inch)", "Size tính toán", "Số lượng rập", "Dài sản xuất (L-inch)", "Rộng sản xuất (W-inch)", "polygon_net_area", "Gross Consumption"]
        for c_idx, h_text in enumerate(det_hd, start=1):
            cell = w_s2.cell(row=3, column=c_idx, value=h_text)
            cell.font = f_header; cell.fill = fill_header; cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); cell.border = bd_thin

        c_row = 4
        for _, r in df_det.iterrows():
            for c_idx, h_col in enumerate(det_hd, start=1):
                val = r.get(h_col, "")
                cell = w_s2.cell(row=c_row, column=c_idx, value=val)
                cell.font = f_normal; cell.border = bd_thin
                
                # SỬA LỖI CÚ PHÁP: Cố định đúng chỉ số mảng cột căn lề bảng chi tiết
                if c_idx == 1 or c_idx == 2 or c_idx == 3:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                elif c_idx == 4 or c_idx == 5 or c_idx == 6:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    if isinstance(val, (int, float)):
                        cell.number_format = '#,##0.0000' if h_col == "Gross Consumption" else '#,##0.00'
            c_row += 1

        for col_idx, col in enumerate(w_s2.columns, start=1):
            m_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col_idx)
            w_s2.column_dimensions[col_letter].width = max(m_len + 3, 12)

        workbook.save(output_stream)
        output_stream.seek(0)
        return output_stream
       # =====================================================================
       # =====================================================================
       # =====================================================================
    #    # =====================================================================
        # =====================================================================
       # =====================================================================
    # 🟩 ĐOẠN 7: REAL-TIME AUDIT INTERFACE & INTERACTIVE CONTROL (UI ENGINE)
    # =====================================================================
    st.header("📋 AI AUDIT REPORT (BÁO CÁO KIỂM TOÁN ĐỊNH MỨC TỰ ĐỘNG)")
    ai_decision_final = ctx.get("ai_expert_decision", {})
    estimated_prior_val = float(ai_decision_final.get("estimated_density_prior", 0.78))
    ui_display_density = float(ai_decision_final.get("real_fabric_density", estimated_prior_val))
    comp_score_val = float(ai_decision_final.get("complexity_score", 45.0))
    ui_complexity_tier = "COMPLEX" if comp_score_val >= 50 else "NORMAL"
    ui_complexity_icon = "🔴" if comp_score_val >= 75 else ("🟡" if comp_score_val >= 45 else "🟢")
    prod_cat_ui = str(ai_decision_final.get("product_category", "JEAN_LONG")).upper().strip()

    m1, m2, m3, m4 = st.columns(4)
    m1.metric("🤖 Loại Hàng Nhận Diện", ai_product_type if 'ai_product_type' in locals() else "JEAN_LONG")
    m2.metric(f"{ui_complexity_icon} Mức Độ Phức Tạp", f"{ui_complexity_tier} ({comp_score_val:.0f}/100)")
    m3.metric("📐 Mật Độ Sơ Đồ Chỉ Định", f"{ui_display_density*100:.2f}%")
    m4.metric("🎯 Độ Tin Cậy AI (Confidence)", f"{float(ctx.get('confidence', 0.95))*100:.1f}%")

    # Hàm phân loại chất liệu layer hiển thị đồng bộ tri thức ép keo Đoạn 3
    def ui_layer_material_classify(row, idx):
        if "user_edited_materials" in st.session_state and idx in st.session_state["user_edited_materials"]:
            return str(st.session_state["user_edited_materials"][idx]).upper().strip()
        mat_str = str(row[m_col]).upper().strip()
        comp_str = str(row.get("Component Name", row.get("component_name", ""))).upper().strip()
        role_str = str(row.get("Role/Piece Type", row.get("geometry_role", ""))).upper().strip()
        
        fusing_kws = ["FUSING", "INTERLINING", "KEO", "MEC", "RIB", "BOND", "TAPE", "ADHESIVE", "COLLAR", "CUFF", "WAISTBAND", "LOT KEO", "TAPE"]
        lining_kws = ["LINING", "LOT", "POCKETING", "MESH", "TAFFETA", "VAI LOT"]
        
        if any(k in mat_str or k in comp_str for k in fusing_kws): return "FUSING"
        if any(k in mat_str or k in comp_str or k in role_str for k in lining_kws): return "LINING"
        if any(k in mat_str or k in comp_str for k in ["ACCESSORY", "THREAD", "CHI", "BUTTON", "ZIPPER", "RIVET"]): return "ACCESSORY"
        
        if 'FUSING_STRICT_RULES' in globals() and prod_cat_ui in FUSING_STRICT_RULES:
            if any(k in comp_str for k in FUSING_STRICT_RULES[prod_cat_ui]):
                if not any(x in mat_str for x in ["THREAD", "CHI", "ACCESSORY", "BUTTON", "ZIPPER"]):
                    return "FUSING"
        return "FABRIC"

    # Gán nhãn tạm thời cho cột phân loại vật tư
    df_bom["_temp_class"] = [ui_layer_material_classify(r, idx) for idx, r in df_bom.iterrows()]
    
    # 🚨 BỘ PHÒNG VỆ CHỐNG SẬP APP: Nếu chưa có cột định mức thương mại, tự động tạo giá trị bằng 0 để không bị lỗi KeyError
    if "Gross Consumption" not in df_bom.columns:
        df_bom["Gross Consumption"] = 0.0

    summary_grouped = df_bom.groupby(["_temp_class"]).agg({"Gross Consumption": "sum"}).reset_index()
    cls_map = {"FABRIC": "VẢI CHÍNH", "FUSING": "MÉC / KEO", "LINING": "VẢI LÓT", "THREAD": "CHỈ MAY", "ACCESSORY": "PHỤ LIỆU", "UNKNOWN": "VẬT TƯ KHÁC"}
    
    df_summary = pd.DataFrame({
        "Phân loại vật tư": summary_grouped["_temp_class"].map(cls_map).fillna("VẬT TƯ KHÁC"),
        "Material Class": summary_grouped["_temp_class"],
        "Gross Consumption": summary_grouped["Gross Consumption"].round(4),
        "UOM": "YDS"
    })

    st.markdown("##### 📊 Bảng Tổng Hợp Tiêu Hao Vật Tư Đại Trà (BOM Summary)")
    st.dataframe(df_summary, use_container_width=True, hide_index=True)

    df_bom_display = df_bom.copy()
    if "Calculated Width (Inch)" in df_bom_display.columns:
        df_bom_display["Khổ vải sản xuất (inch)"] = df_bom_display["Calculated Width (Inch)"].round(1)
    else:
        df_bom_display["Khổ vải sản xuất (inch)"] = current_fabric_width if 'current_fabric_width' in locals() else 58.0
        
    df_bom_display["Size tính toán"] = detected_size_code if 'detected_size_code' in locals() else "M"
    df_bom_display["material_class"] = df_bom_display["_temp_class"]
    df_bom_display = df_bom_display.rename(columns={"component_name": "Component Name", "material_class": "Material Class", "geometry_role": "Role/Piece Type"})
    df_bom_display["Số lượng rập"] = [float(st.session_state.get("user_edited_pieces", {}).get(idx, r["pcs_numeric"])) for idx, r in df_bom.iterrows()]
    df_bom_display["_original_row_index"] = df_bom.index

    ordered_cols = ["_original_row_index", "Component Name", "Material Class", "Role/Piece Type", "Khổ vải sản xuất (inch)", "Size tính toán", "Số lượng rập", "Dài sản xuất (L-inch)", "Rộng sản xuất (W-inch)", "polygon_net_area", "Gross Consumption"]
    display_final_cols = [c for c in ordered_cols if c in df_bom_display.columns]
    df_bom_display = df_bom_display[display_final_cols]

    col_t1, col_t2 = st.columns(2)
    col_t1.subheader("📋 Bảng Kế Hoạch Định Mức Rải Sơ Đồ Chi Tiết")
    
    with col_t2:
        try:
            if 'local_export_excel_ppj_format' in locals():
                excel_file = local_export_excel_ppj_format(df_summary, df_bom_display.drop(columns=["_original_row_index"], errors="ignore"), prod, ctx, ui_display_density)
                style_name_clean = str(ctx.get('style_code', 'Style')).strip().replace('/', '_').replace('\\', '_')
                st.download_button("🟢 DOWNLOAD EXCEL ĐỊNH MỨC THƯƠNG MẠI", data=excel_file, mime="application/vnd.openpyxl_formats-officedocument.spreadsheetml.sheet", file_name=f"PPJ_BOM_{prod}_{style_name_clean}.xlsx", use_container_width=True)
        except Exception as e: st.error(f"Lỗi kết xuất Excel: {e}")

    edited_df = st.data_editor(
        df_bom_display, 
        column_config={
            "_original_row_index": None, 
            "Số lượng rập": st.column_config.NumberColumn("Số lượng rập", min_value=1.0, max_value=40.0, step=1.0),
            "Material Class": st.column_config.SelectboxColumn(
                "Material Class", help="Chọn lại nhóm vật tư nếu AI nhận diện sai",
                options=["FABRIC", "FUSING", "LINING", "ACCESSORY", "THREAD"], required=True
            )
        }, use_container_width=True, hide_index=True, key="bom_data_editor_grid_final_v8" 
    )

    has_changed = False
    for _, row in edited_df.iterrows():
        orig_idx = int(row["_original_row_index"])
        old_pcs = float(df_bom.at[orig_idx, "pcs_numeric"])
        new_pcs = float(row["Số lượng rập"])
        if old_pcs != new_pcs:
            st.session_state["user_edited_pieces"][orig_idx] = new_pcs
            has_changed = True
        old_mat = str(df_bom.at[orig_idx, "_temp_class"]).upper().strip()
        new_mat = str(row["Material Class"]).upper().strip()
        if old_mat != new_mat:
            st.session_state["user_edited_materials"][orig_idx] = new_mat
            has_changed = True
            
    if has_changed:
        st.session_state["processed_display_rows"] = df_bom.to_dict(orient="records")
        st.rerun()
